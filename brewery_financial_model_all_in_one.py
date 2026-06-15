"""
All-in-one Python financial model for a microbrewery-style business.

Contains:
1) A transparent monthly financial model engine:
   - SKU x channel volumes
   - cost-plus pricing + channel price factors + inflation
   - direct costs + fixed OPEX + inflation
   - CAPEX + straight-line depreciation
   - working capital (DSO/DIO/DPO + % add-ons)
   - multi-facility debt schedule (linear / annuity / specified)
   - dividend policy (cash sweep or share-of-profits)
   - monthly + annual financial statements
   - DCF valuation (FCFF + EV/EBITDA terminal multiple)
   - investor IRR + MOIC

2) A runnable example (main) that:
   - creates sample assumptions
   - runs the model
   - exports an Excel output workbook

Dependencies:
- pandas
- numpy
- openpyxl (optional, only for Excel export)

Usage:
    python brewery_financial_model_all_in_one.py

Notes:
- This is an "engine" (model logic), not a cell-by-cell conversion of an Excel template.
- Replace the example assumptions with your own inputs.
"""

from __future__ import annotations

from dataclasses import dataclass, field
import re
from typing import Dict, Iterable, List, Literal, Optional, Tuple

import numpy as np
import pandas as pd
from finmodel.allocation import allocate_opex_by_drivers
from finmodel.operations_schedule import plan_brewery_operations
from finmodel.opex_defaults import build_default_opex_cost_pools
from finmodel.opex_schemas import OpexCostPool, SKUCostContext


# =============================
# Utility functions
# =============================
RepaymentType = Literal["linear", "annuity", "interest_only_then_linear", "specified"]
SalesPlanPropagationMode = Literal["manual", "repeat_first_year", "grow_first_year"]


def annual_to_monthly_rate(annual_rate: float) -> float:
    """Convert an effective annual rate to an effective monthly rate."""
    return (1.0 + float(annual_rate)) ** (1.0 / 12.0) - 1.0


def safe_div(n: float, d: float, default: float = 0.0) -> float:
    return float(n) / float(d) if float(d) != 0.0 else float(default)


def irr(cashflows: Iterable[float], guess: float = 0.1) -> float:
    """
    Internal rate of return (per period) via Newton method.
    Returns np.nan if it fails to converge.
    """
    cfs = np.array(list(cashflows), dtype=float)
    if np.allclose(cfs, 0.0):
        return np.nan

    def f(r: float) -> float:
        return np.sum(cfs / ((1.0 + r) ** np.arange(len(cfs))))

    def fprime(r: float) -> float:
        t = np.arange(len(cfs))
        return np.sum(-t * cfs / ((1.0 + r) ** (t + 1.0)))

    r = float(guess)
    for _ in range(200):
        fr = f(r)
        fpr = fprime(r)
        if abs(fpr) < 1e-12:
            break
        new_r = r - fr / fpr
        if abs(new_r - r) < 1e-10:
            return new_r
        r = new_r
    return np.nan


# =============================
# Data structures
# =============================
@dataclass(frozen=True)
class ModelConfig:
    start_date: str = "2025-01-01"
    months: int = 120  # 10 years monthly
    pricing_cost_basis_month: int = 24  # cost basis month to compute cost-plus base prices (0-indexed)
    price_inflation_annual: float = 0.015
    cost_inflation_annual: float = 0.015
    tax_rate: float = 0.25

    # Working capital assumptions
    days_receivables: float = 20.0
    days_inventory: float = 15.0
    days_payables: float = 30.0
    other_current_assets_pct_revenue: float = 0.05
    other_current_liabilities_pct_direct_costs: float = 0.05

    # Valuation
    wacc_annual: float = 0.122
    exit_month: Optional[int] = None  # if None -> last month
    exit_ev_ebitda_multiple: float = 8.0

    # Cash
    initial_cash: float = 0.0

    # Liquidity / revolver / covenant support
    revolver_limit: float = 750_000.0
    revolver_interest_annual: float = 0.085
    revolver_target_cash: float = 250_000.0
    min_dscr: float = 1.20
    min_interest_coverage: float = 2.00
    max_leverage_ratio: float = 4.50

    # Capacity shortfall support
    temporary_labor_premium_pct: float = 0.20


@dataclass(frozen=True)
class DividendPolicy:
    enabled: bool = True
    model: Literal["cash_sweep", "share_of_profits"] = "cash_sweep"
    start_month: int = 60  # e.g., "Year 5" in a 0-indexed 10-year model
    minimum_cash_position: float = 1_500_000.0
    payout_ratio: float = 0.25  # only used for share_of_profits


@dataclass(frozen=True)
class DebtFacility:
    name: str
    principal: float
    annual_interest_rate: float
    draw_month: int = 0
    grace_months: int = 0
    term_months: int = 60
    repayment_type: RepaymentType = "linear"

    # Only used for repayment_type == "specified"
    specified_principal_payments: Optional[Dict[int, float]] = None  # month -> principal payment


@dataclass(frozen=True)
class CapexItem:
    name: str
    amount: float
    capex_month: int = 0
    depreciation_years: float = 0.0  # 0 => non-depreciable (e.g., land)


@dataclass(frozen=True)
class CostPoolInput:
    name: str
    cost_type: Literal["direct", "indirect"] = "indirect"
    behavior: Literal["variable", "fixed", "step_fixed", "blended"] = "blended"
    allocation_driver: Literal["units", "liters", "revenue", "channel_units", "channel_revenue", "active_sku", "complexity"] = "units"
    scope: Literal["global", "family", "channel", "sku"] = "global"
    channel: Optional[str] = None
    unit_variable_cost: float = 0.0
    fixed_monthly_cost: float = 0.0
    step_threshold: float = 0.0
    step_increment: float = 0.0


@dataclass(frozen=True)
class OtherIncomeItem:
    other_income_name: str
    amount: float
    active: bool = True
    category: Optional[str] = None
    notes: Optional[str] = None


@dataclass
class ModelInputs:
    """
    Required inputs:
    - skus: DataFrame with at minimum:
        sku_id (unique), name, markup_pct, relative_opex_weight (optional)
      direct_cost_per_unit is derived from direct cost pools unless explicit override is provided.
    - channels: DataFrame with at minimum:
        channel (unique), price_factor
    - sales_plan: DataFrame with columns:
        date (datetime), sku_id, channel, units

    Optional inputs:
    - other_income_items: list of OtherIncomeItem with monthly amounts
    - other_income_monthly: (deprecated compatibility path) Series indexed by date or scalar float
    - cost_pools: list of CostPoolInput
    - capex_items: list of CapexItem
    - debt_facilities: list of DebtFacility
    - equity_injections: dict month -> amount (positive cash-in)
    """
    skus: pd.DataFrame
    channels: pd.DataFrame
    sales_plan: pd.DataFrame
    sales_plan_frequency: Literal["monthly", "quarterly", "yearly"] = "monthly"
    sales_plan_propagation_mode: SalesPlanPropagationMode = "manual"
    sales_plan_propagation_growth_annual: float = 0.0

    other_income_items: Optional[List[OtherIncomeItem]] = None
    other_income_monthly: float | pd.Series = 0.0
    cost_pools: Optional[List[CostPoolInput]] = None

    direct_labor_schedule: Optional[pd.DataFrame] = None
    indirect_labor_schedule: Optional[pd.DataFrame] = None
    inventory_schedule: Optional[pd.DataFrame] = None
    receivables_schedule: Optional[pd.DataFrame] = None
    payables_schedule: Optional[pd.DataFrame] = None
    sku_operations: Optional[pd.DataFrame] = None
    brewhouse_schedule: Optional[pd.DataFrame] = None
    cellar_schedule: Optional[pd.DataFrame] = None
    packaging_schedule: Optional[pd.DataFrame] = None

    capex_items: Optional[List[CapexItem]] = None
    debt_facilities: Optional[List[DebtFacility]] = None
    equity_injections: Optional[Dict[int, float]] = None


@dataclass
class ModelRunResult:
    monthly: pd.DataFrame
    annual: pd.DataFrame
    prices: pd.DataFrame
    debt_schedules: Dict[str, pd.DataFrame]
    valuation: Dict[str, float]
    opex_allocation_views: Dict[str, pd.DataFrame]
    supporting_schedules: Dict[str, pd.DataFrame] = field(default_factory=dict)


def propagate_sales_plan_monthly(
    monthly_sales: pd.DataFrame,
    idx: pd.DatetimeIndex,
    mode: SalesPlanPropagationMode = "manual",
    annual_growth_pct: float = 0.0,
) -> pd.DataFrame:
    if monthly_sales is None or monthly_sales.empty:
        return pd.DataFrame(columns=["date", "sku_id", "channel", "units"])

    out = monthly_sales.copy()
    out["date"] = pd.to_datetime(out.get("date"), errors="coerce")
    out["units"] = pd.to_numeric(out.get("units", 0.0), errors="coerce").fillna(0.0)
    out = out.dropna(subset=["date"])
    out = out[out["date"].isin(idx)].copy()
    if out.empty or mode == "manual":
        return out[["date", "sku_id", "channel", "units"]].sort_values(["date", "sku_id", "channel"]).reset_index(drop=True)

    template_year = int(out["date"].dt.year.min())
    template = out[out["date"].dt.year == template_year].copy()
    if template.empty:
        return out[["date", "sku_id", "channel", "units"]].sort_values(["date", "sku_id", "channel"]).reset_index(drop=True)

    template["month_num"] = template["date"].dt.month
    template = (
        template.groupby(["sku_id", "channel", "month_num"], dropna=False, as_index=False)["units"]
        .sum()
        .sort_values(["sku_id", "channel", "month_num"])
        .reset_index(drop=True)
    )
    template_lookup = template.set_index(["sku_id", "channel", "month_num"])["units"].to_dict()
    combos = template[["sku_id", "channel"]].drop_duplicates().to_dict("records")

    growth_rate = float(annual_growth_pct)
    propagated_rows: List[Dict[str, object]] = []
    for date in idx:
        year_offset = int(date.year) - template_year
        if mode == "grow_first_year":
            growth_factor = (1.0 + growth_rate) ** max(year_offset, 0)
        else:
            growth_factor = 1.0
        for combo in combos:
            base_units = float(template_lookup.get((combo["sku_id"], combo["channel"], int(date.month)), 0.0))
            propagated_rows.append(
                {
                    "date": date,
                    "sku_id": combo["sku_id"],
                    "channel": combo["channel"],
                    "units": base_units * growth_factor,
                }
            )
    return pd.DataFrame(propagated_rows, columns=["date", "sku_id", "channel", "units"])


# =============================
# Core engine
# =============================
class MicrobreweryFinancialModel:
    def __init__(
        self,
        config: ModelConfig,
        dividend_policy: DividendPolicy,
        inputs: ModelInputs,
    ) -> None:
        self.cfg = config
        self.div = dividend_policy
        self.inputs = inputs

        self._validate_inputs()

    # ---------- validation ----------
    def _validate_inputs(self) -> None:
        skus = self.inputs.skus.copy()
        channels = self.inputs.channels.copy()
        sales = self.inputs.sales_plan.copy()

        required_sku_cols = {"sku_id", "name", "markup_pct"}
        missing = required_sku_cols - set(skus.columns)
        if missing:
            raise ValueError(f"skus is missing columns: {sorted(missing)}")

        if "direct_cost_per_unit" not in skus.columns:
            skus["direct_cost_per_unit"] = 0.0
        if "direct_cost_per_unit_override" not in skus.columns:
            skus["direct_cost_per_unit_override"] = np.nan
        if "relative_opex_weight" not in skus.columns:
            skus["relative_opex_weight"] = 1.0
        else:
            skus["relative_opex_weight"] = skus["relative_opex_weight"].replace("", np.nan).fillna(1.0)

        required_channel_cols = {"channel", "price_factor"}
        missing = required_channel_cols - set(channels.columns)
        if missing:
            raise ValueError(f"channels is missing columns: {sorted(missing)}")

        required_sales_cols = {"date", "sku_id", "channel", "units"}
        missing = required_sales_cols - set(sales.columns)
        if missing:
            raise ValueError(f"sales_plan is missing columns: {sorted(missing)}")

        # Normalize dtypes
        sales["date"] = pd.to_datetime(sales["date"])
        # Make sku_id type consistent
        sales["sku_id"] = sales["sku_id"].astype(skus["sku_id"].dtype, copy=False)
        sales["units"] = pd.to_numeric(sales["units"], errors="coerce")

        # Enforce numeric columns to avoid object-dtype arithmetic TypeErrors.
        sku_numeric_cols = ["direct_cost_per_unit", "direct_cost_per_unit_override", "markup_pct", "relative_opex_weight"]
        for col in sku_numeric_cols:
            skus[col] = pd.to_numeric(skus[col], errors="coerce")
        skus["direct_cost_per_unit_override"] = skus["direct_cost_per_unit_override"].where(
            skus["direct_cost_per_unit_override"].notna(),
            np.nan,
        )
        channels["price_factor"] = pd.to_numeric(channels["price_factor"], errors="coerce")

        required_numeric = ["direct_cost_per_unit", "markup_pct", "relative_opex_weight"]
        if skus[required_numeric].isna().any().any():
            bad_rows = skus.loc[skus[required_numeric].isna().any(axis=1), ["sku_id", "name"]].to_dict("records")
            raise ValueError(
                "skus has non-numeric values in direct_cost_per_unit/direct_cost_per_unit_override/markup_pct/relative_opex_weight "
                f"for rows: {bad_rows}"
            )
        if channels["price_factor"].isna().any():
            bad_channels = channels.loc[channels["price_factor"].isna(), "channel"].tolist()
            raise ValueError(f"channels has non-numeric price_factor for channels: {bad_channels}")
        if sales["units"].isna().any():
            bad_sales = sales.loc[sales["units"].isna(), ["date", "sku_id", "channel"]].to_dict("records")
            raise ValueError(f"sales_plan has non-numeric units for rows: {bad_sales}")
        if self.inputs.sales_plan_frequency not in {"monthly", "quarterly", "yearly"}:
            raise ValueError("sales_plan_frequency must be one of: monthly, quarterly, yearly")
        if self.inputs.sales_plan_propagation_mode not in {"manual", "repeat_first_year", "grow_first_year"}:
            raise ValueError("sales_plan_propagation_mode must be one of: manual, repeat_first_year, grow_first_year")
        growth_pct = pd.to_numeric(self.inputs.sales_plan_propagation_growth_annual, errors="coerce")
        if pd.isna(growth_pct):
            raise ValueError("sales_plan_propagation_growth_annual must be numeric")
        self.inputs.sales_plan_propagation_growth_annual = float(growth_pct)

        # Keep normalized copies
        self.inputs.skus = skus
        self.inputs.channels = channels
        self.inputs.sales_plan = sales

        if self.inputs.capex_items is None:
            self.inputs.capex_items = []
        if self.inputs.debt_facilities is None:
            self.inputs.debt_facilities = []
        if self.inputs.equity_injections is None:
            self.inputs.equity_injections = {}
        if self.inputs.cost_pools is None:
            self.inputs.cost_pools = []
        if self.inputs.other_income_items is None:
            self.inputs.other_income_items = []
        if self.inputs.direct_labor_schedule is None:
            self.inputs.direct_labor_schedule = pd.DataFrame()
        if self.inputs.indirect_labor_schedule is None:
            self.inputs.indirect_labor_schedule = pd.DataFrame()
        if self.inputs.inventory_schedule is None:
            self.inputs.inventory_schedule = pd.DataFrame()
        if self.inputs.receivables_schedule is None:
            self.inputs.receivables_schedule = pd.DataFrame()
        if self.inputs.payables_schedule is None:
            self.inputs.payables_schedule = pd.DataFrame()
        if self.inputs.sku_operations is None:
            self.inputs.sku_operations = pd.DataFrame()
        if self.inputs.brewhouse_schedule is None:
            self.inputs.brewhouse_schedule = pd.DataFrame()
        if self.inputs.cellar_schedule is None:
            self.inputs.cellar_schedule = pd.DataFrame()
        if self.inputs.packaging_schedule is None:
            self.inputs.packaging_schedule = pd.DataFrame()

    def _other_income_series(self, idx: pd.DatetimeIndex) -> pd.Series:
        if self.inputs.other_income_items:
            total = 0.0
            for item in self.inputs.other_income_items:
                if bool(item.active):
                    total += float(item.amount)
            return pd.Series(total, index=idx, name="other_income")
        return self._as_monthly_series(self.inputs.other_income_monthly, idx, "other_income")

    # ---------- timeline ----------
    def _timeline(self) -> pd.DatetimeIndex:
        start = pd.to_datetime(self.cfg.start_date)
        return pd.date_range(start=start, periods=int(self.cfg.months), freq="MS")  # month start

    def _inflation_index(self, annual_rate: float, idx: pd.DatetimeIndex) -> pd.Series:
        m = annual_to_monthly_rate(annual_rate)
        factors = (1.0 + m) ** np.arange(len(idx))
        return pd.Series(factors, index=idx, name="inflation_index")

    # ---------- scalars/series helpers ----------
    def _as_monthly_series(self, x: float | pd.Series, idx: pd.DatetimeIndex, name: str) -> pd.Series:
        if isinstance(x, pd.Series):
            s = x.copy()
            s.index = pd.to_datetime(s.index)
            s = s.reindex(idx).ffill().fillna(0.0)
            s.name = name
            return s
        return pd.Series(float(x), index=idx, name=name)

    def _model_year_count(self) -> int:
        return max(int(np.ceil(float(self.cfg.months) / 12.0)), 1)

    def _month_year_numbers(self, idx: pd.DatetimeIndex) -> np.ndarray:
        return (np.arange(len(idx)) // 12) + 1

    def _year_value(self, row: pd.Series, year_num: int, default: float = 0.0) -> float:
        value = row.get(f"Year {year_num}", default)
        if pd.isna(value):
            return float(default)
        return float(pd.to_numeric(value, errors="coerce"))

    def _schedule_numeric(self, row: pd.Series, key: str, default: float = 0.0) -> float:
        value = row.get(key, default)
        if pd.isna(value):
            return float(default)
        return float(pd.to_numeric(value, errors="coerce"))

    def _sku_units_monthly(self, idx: pd.DatetimeIndex, units_wide: pd.DataFrame) -> pd.DataFrame:
        sku_ids = list(self.inputs.skus["sku_id"])
        if units_wide.empty:
            return pd.DataFrame(0.0, index=idx, columns=sku_ids)
        sku_units = units_wide.T.groupby(level=0).sum().T.reindex(idx).fillna(0.0)
        return sku_units.reindex(columns=sku_ids, fill_value=0.0)

    def _sku_liters_monthly(self, idx: pd.DatetimeIndex, units_wide: pd.DataFrame) -> pd.DataFrame:
        sku_units = self._sku_units_monthly(idx, units_wide)
        liters = pd.DataFrame(0.0, index=idx, columns=sku_units.columns)
        sku_lookup = self.inputs.skus.set_index("sku_id")
        for sku_id in liters.columns:
            liters_per = self._liters_per_unit_from_name(str(sku_lookup.loc[sku_id, "name"]))
            liters[sku_id] = sku_units[sku_id] * liters_per
        return liters

    def _labor_schedule_monthly(
        self,
        idx: pd.DatetimeIndex,
        units_wide: pd.DataFrame,
        labor_df: pd.DataFrame,
    ) -> tuple[pd.DataFrame, pd.DataFrame]:
        sku_ids = list(self.inputs.skus["sku_id"])
        empty_alloc = pd.DataFrame(0.0, index=idx, columns=sku_ids)
        if labor_df is None or labor_df.empty:
            empty_summary = pd.DataFrame(
                {
                    "labor_cost": pd.Series(0.0, index=idx),
                    "required_liters": pd.Series(0.0, index=idx),
                    "capacity_liters": pd.Series(0.0, index=idx),
                    "capacity_shortfall_liters": pd.Series(0.0, index=idx),
                    "temporary_labor_cost": pd.Series(0.0, index=idx),
                }
            )
            return empty_alloc, empty_summary

        sku_units = self._sku_units_monthly(idx, units_wide)
        sku_liters = self._sku_liters_monthly(idx, units_wide)
        year_nums = self._month_year_numbers(idx)

        alloc = pd.DataFrame(0.0, index=idx, columns=sku_ids)
        labor_cost = pd.Series(0.0, index=idx, name="labor_cost")
        capacity_liters = pd.Series(0.0, index=idx, name="capacity_liters")
        required_liters = sku_liters.sum(axis=1).rename("required_liters")

        for _, row in labor_df.iterrows():
            if pd.isna(row.get("role")):
                continue
            monthly_cost_year1 = self._schedule_numeric(row, "monthly_cost_per_fte", 0.0)
            annual_raise_pct = self._schedule_numeric(row, "annual_raise_pct", 0.0)
            benefits_pct = self._schedule_numeric(row, "benefits_pct", 0.0)
            payroll_tax_pct = self._schedule_numeric(row, "payroll_tax_pct", 0.0)
            overtime_pct = self._schedule_numeric(row, "overtime_pct", 0.0)
            capacity_per_fte = self._schedule_numeric(row, "capacity_liters_per_fte_month", 0.0)
            scope = str(row.get("scope", "global") or "global").strip().lower()
            target_sku_id = pd.to_numeric(row.get("target_sku_id"), errors="coerce")
            allocation_driver = str(row.get("allocation_driver", "liters") or "liters").strip().lower()

            headcounts = np.array([self._year_value(row, int(y), 0.0) for y in year_nums], dtype=float)
            salary_factors = (1.0 + annual_raise_pct) ** np.maximum(year_nums - 1, 0)
            loaded_cost = monthly_cost_year1 * salary_factors * (
                1.0 + benefits_pct + payroll_tax_pct + overtime_pct
            )
            role_cost = pd.Series(headcounts * loaded_cost, index=idx)
            labor_cost = labor_cost.add(role_cost, fill_value=0.0)
            capacity_liters = capacity_liters.add(headcounts * capacity_per_fte, fill_value=0.0)

            basis = sku_liters if allocation_driver == "liters" else sku_units
            basis = basis.copy()
            if scope == "sku" and not pd.isna(target_sku_id):
                target = int(target_sku_id)
                basis = pd.DataFrame(0.0, index=idx, columns=sku_ids)
                if target in basis.columns:
                    basis[target] = (
                        sku_liters[target] if allocation_driver == "liters" else sku_units[target]
                    )
            elif allocation_driver == "fixed":
                basis = (sku_units > 0).astype(float)
            denom = basis.sum(axis=1).replace(0.0, np.nan)
            shares = basis.div(denom, axis=0).fillna(0.0)
            alloc = alloc.add(shares.mul(role_cost, axis=0), fill_value=0.0)

        capacity_shortfall = (required_liters - capacity_liters).clip(lower=0.0)
        shortfall_ratio = np.divide(
            capacity_shortfall.to_numpy(),
            required_liters.to_numpy(),
            out=np.zeros(len(idx), dtype=float),
            where=required_liters.to_numpy() > 0.0,
        )
        temporary_labor_cost = labor_cost * shortfall_ratio * float(self.cfg.temporary_labor_premium_pct)
        labor_cost = labor_cost.add(temporary_labor_cost, fill_value=0.0)
        if labor_cost.sum() > 0.0:
            denom = alloc.sum(axis=1).replace(0.0, np.nan)
            shares = alloc.div(denom, axis=0).fillna(0.0)
            alloc = alloc.add(shares.mul(temporary_labor_cost, axis=0), fill_value=0.0)

        summary = pd.DataFrame(
            {
                "labor_cost": labor_cost,
                "required_liters": required_liters,
                "capacity_liters": capacity_liters,
                "capacity_shortfall_liters": capacity_shortfall,
                "temporary_labor_cost": temporary_labor_cost,
            },
            index=idx,
        )
        return alloc, summary

    def _revenue_schedule_adjustments(
        self,
        idx: pd.DatetimeIndex,
        revenue_wide: pd.DataFrame,
    ) -> tuple[pd.Series, pd.DataFrame, pd.DataFrame]:
        if revenue_wide.empty:
            empty = pd.Series(0.0, index=idx)
            return empty.rename("net_revenue"), pd.DataFrame(index=idx), pd.DataFrame(index=idx)

        channel_revenue = revenue_wide.T.groupby(level=1).sum().T.reindex(idx).fillna(0.0)
        schedule = self.inputs.receivables_schedule
        if schedule is None or schedule.empty:
            net_revenue = channel_revenue.sum(axis=1).rename("net_revenue")
            breakdown = pd.DataFrame(
                {
                    "gross_revenue": net_revenue,
                    "trade_spend": pd.Series(0.0, index=idx),
                    "returns_allowances": pd.Series(0.0, index=idx),
                    "bad_debt_expense": pd.Series(0.0, index=idx),
                },
                index=idx,
            )
            ar_details = pd.DataFrame({"receivables": net_revenue * (self.cfg.days_receivables / 365.0)}, index=idx)
            return net_revenue, breakdown, ar_details

        trade = pd.Series(0.0, index=idx)
        returns = pd.Series(0.0, index=idx)
        bad_debt = pd.Series(0.0, index=idx)
        receivables = pd.Series(0.0, index=idx)
        detail_cols: Dict[str, pd.Series] = {}
        for _, row in schedule.iterrows():
            channel = str(row.get("channel", "")).strip()
            if not channel or channel not in channel_revenue.columns:
                continue
            gross = channel_revenue[channel]
            trade_pct = self._schedule_numeric(row, "trade_spend_pct", 0.0)
            returns_pct = self._schedule_numeric(row, "returns_pct", 0.0)
            bad_pct = self._schedule_numeric(row, "bad_debt_pct", 0.0)
            days = pd.Series(
                [self._year_value(row, int(y), self.cfg.days_receivables) for y in self._month_year_numbers(idx)],
                index=idx,
            )
            channel_trade = gross * trade_pct
            channel_returns = gross * returns_pct
            channel_net = gross - channel_trade - channel_returns
            trade = trade.add(channel_trade, fill_value=0.0)
            returns = returns.add(channel_returns, fill_value=0.0)
            bad_debt = bad_debt.add(channel_net * bad_pct, fill_value=0.0)
            channel_ar = channel_net * (days / 365.0)
            receivables = receivables.add(channel_ar, fill_value=0.0)
            detail_cols[f"receivables_{channel}"] = channel_ar

        net_revenue = (channel_revenue.sum(axis=1) - trade - returns).rename("net_revenue")
        breakdown = pd.DataFrame(
            {
                "gross_revenue": channel_revenue.sum(axis=1),
                "trade_spend": trade,
                "returns_allowances": returns,
                "bad_debt_expense": bad_debt,
            },
            index=idx,
        )
        ar_details = pd.DataFrame(detail_cols, index=idx)
        ar_details["receivables"] = receivables
        return net_revenue, breakdown, ar_details

    def _inventory_schedule_components(
        self,
        idx: pd.DatetimeIndex,
        direct_costs_base: pd.Series,
    ) -> tuple[pd.DataFrame, pd.Series, pd.Series]:
        schedule = self.inputs.inventory_schedule
        if schedule is None or schedule.empty:
            inventory = direct_costs_base * (self.cfg.days_inventory / 365.0)
            details = pd.DataFrame({"inventory": inventory}, index=idx)
            return details, pd.Series(0.0, index=idx, name="inventory_writeoff"), pd.Series(
                0.0, index=idx, name="inventory_reserve"
            )

        gross_inventory = pd.Series(0.0, index=idx)
        reserve = pd.Series(0.0, index=idx)
        writeoff = pd.Series(0.0, index=idx)
        detail_cols: Dict[str, pd.Series] = {}
        for _, row in schedule.iterrows():
            stage = str(row.get("stage", "")).strip() or "inventory"
            share = self._schedule_numeric(row, "cost_share_pct", 0.0)
            reserve_pct = self._schedule_numeric(row, "reserve_pct", 0.0)
            writeoff_pct = self._schedule_numeric(row, "writeoff_pct", 0.0)
            days = pd.Series(
                [self._year_value(row, int(y), self.cfg.days_inventory) for y in self._month_year_numbers(idx)],
                index=idx,
            )
            component_base = direct_costs_base * share
            component_gross = component_base * (days / 365.0)
            component_reserve = component_gross * reserve_pct
            component_writeoff = component_base * writeoff_pct
            gross_inventory = gross_inventory.add(component_gross, fill_value=0.0)
            reserve = reserve.add(component_reserve, fill_value=0.0)
            writeoff = writeoff.add(component_writeoff, fill_value=0.0)
            detail_cols[f"inventory_{stage}"] = component_gross - component_reserve

        details = pd.DataFrame(detail_cols, index=idx)
        details["inventory"] = gross_inventory - reserve
        return details, writeoff.rename("inventory_writeoff"), reserve.rename("inventory_reserve")

    def _payables_schedule_components(
        self,
        idx: pd.DatetimeIndex,
        direct_costs_base: pd.Series,
    ) -> tuple[pd.DataFrame, pd.Series]:
        schedule = self.inputs.payables_schedule
        if schedule is None or schedule.empty:
            payables = direct_costs_base * (self.cfg.days_payables / 365.0)
            return pd.DataFrame({"payables": payables}, index=idx), pd.Series(
                0.0, index=idx, name="payable_discount_benefit"
            )

        payables = pd.Series(0.0, index=idx)
        benefit = pd.Series(0.0, index=idx)
        detail_cols: Dict[str, pd.Series] = {}
        for _, row in schedule.iterrows():
            category = str(row.get("supplier_category", "")).strip() or "payables"
            share = self._schedule_numeric(row, "cost_share_pct", 0.0)
            early_discount = self._schedule_numeric(row, "early_pay_discount_pct", 0.0)
            capture_pct = self._schedule_numeric(row, "discount_capture_pct", 0.0)
            days = pd.Series(
                [self._year_value(row, int(y), self.cfg.days_payables) for y in self._month_year_numbers(idx)],
                index=idx,
            )
            component_base = direct_costs_base * share
            component_ap = component_base * (days / 365.0)
            component_benefit = component_base * early_discount * capture_pct
            payables = payables.add(component_ap, fill_value=0.0)
            benefit = benefit.add(component_benefit, fill_value=0.0)
            detail_cols[f"payables_{category}"] = component_ap

        details = pd.DataFrame(detail_cols, index=idx)
        details["payables"] = payables
        return details, benefit.rename("payable_discount_benefit")

    def _current_liability_split(
        self,
        idx: pd.DatetimeIndex,
        debt_principal: pd.Series,
        revolver_balance: pd.Series,
    ) -> tuple[pd.Series, pd.Series]:
        current_portion = debt_principal.rolling(window=12, min_periods=1).sum().clip(lower=0.0)
        current_debt = current_portion.add(revolver_balance, fill_value=0.0).rename("current_debt")
        long_term_debt = (revolver_balance * 0.0).add(0.0, fill_value=0.0)
        return current_debt, long_term_debt

    # ---------- sales ----------
    def _expand_sales_plan_to_monthly(self, idx: pd.DatetimeIndex) -> pd.DataFrame:
        sales = self.inputs.sales_plan.copy()
        sales["date"] = pd.to_datetime(sales["date"], errors="coerce")
        sales = sales.dropna(subset=["date"])
        freq = self.inputs.sales_plan_frequency
        if freq == "monthly":
            monthly = sales
        else:
            rows: List[Dict[str, object]] = []
            for _, r in sales.iterrows():
                date = pd.to_datetime(r["date"], errors="coerce")
                if pd.isna(date):
                    continue
                if freq == "quarterly":
                    start = date.to_period("Q").start_time
                    months = pd.date_range(start=start, periods=3, freq="MS")
                else:
                    start = date.to_period("Y").start_time
                    months = pd.date_range(start=start, periods=12, freq="MS")
                months = [m for m in months if m in set(idx)]
                if not months:
                    continue
                portion = float(r["units"]) / len(months)
                for m in months:
                    rows.append({"date": m, "sku_id": r["sku_id"], "channel": r["channel"], "units": portion})
            monthly = pd.DataFrame(rows, columns=["date", "sku_id", "channel", "units"])

        return propagate_sales_plan_monthly(
            monthly,
            idx,
            mode=self.inputs.sales_plan_propagation_mode,
            annual_growth_pct=self.inputs.sales_plan_propagation_growth_annual,
        )

    def _units_matrix(self, idx: pd.DatetimeIndex) -> pd.DataFrame:
        """
        Returns monthly units sold as a wide DataFrame with MultiIndex columns:
            (sku_id, channel)
        Missing combinations are filled with 0.
        """
        sales = self._expand_sales_plan_to_monthly(idx)
        sales = sales[sales["date"].isin(idx)].copy()
        if sales.empty:
            return pd.DataFrame(
                index=idx,
                columns=pd.MultiIndex.from_tuples([], names=["sku_id", "channel"]),
            )

        wide = (
            sales.pivot_table(
                index="date",
                columns=["sku_id", "channel"],
                values="units",
                aggfunc="sum",
                fill_value=0.0,
            )
            .reindex(idx)
            .fillna(0.0)
        )
        return wide.sort_index(axis=1)

    def _scale_channel_units_to_sku_totals(
        self,
        demand_units_wide: pd.DataFrame,
        actual_units_by_sku: pd.DataFrame,
    ) -> pd.DataFrame:
        if demand_units_wide.empty:
            return demand_units_wide.copy()
        out = demand_units_wide.copy().astype(float)
        demand_sku = out.T.groupby(level=0).sum().T.reindex(out.index).fillna(0.0)
        actual_sku = actual_units_by_sku.reindex(index=out.index, columns=demand_sku.columns, fill_value=0.0)
        ratios = actual_sku.div(demand_sku.replace(0.0, np.nan)).fillna(0.0)
        for sku_id in demand_sku.columns:
            if sku_id not in out.columns.get_level_values(0):
                continue
            sku_slice = out.xs(sku_id, axis=1, level=0, drop_level=False)
            out.loc[:, sku_slice.columns] = sku_slice.mul(ratios[sku_id], axis=0).to_numpy()
        return out.sort_index(axis=1)

    def _estimated_revenue_wide_from_units(self, idx: pd.DatetimeIndex, units_wide: pd.DataFrame) -> pd.DataFrame:
        if units_wide.empty:
            return pd.DataFrame(index=idx)
        skus = self.inputs.skus.set_index("sku_id")
        channels = self.inputs.channels.set_index("channel")
        cols: List[Tuple[int, str]] = []
        data: List[np.ndarray] = []
        for sku_id, row in skus.iterrows():
            base_direct_cost = float(row.get("direct_cost_per_unit_override", np.nan))
            if not np.isfinite(base_direct_cost) or base_direct_cost <= 0:
                base_direct_cost = float(row.get("direct_cost_per_unit", 0.0))
            if not np.isfinite(base_direct_cost) or base_direct_cost <= 0:
                base_direct_cost = 1.0
            base_price = base_direct_cost * (1.0 + float(row["markup_pct"]))
            for ch, ch_row in channels.iterrows():
                cols.append((sku_id, ch))
                data.append(np.full(len(idx), base_price * float(ch_row["price_factor"])))
        price_proxy = pd.DataFrame(
            np.array(data).T,
            index=idx,
            columns=pd.MultiIndex.from_tuples(cols, names=["sku_id", "channel"]),
        ).sort_index(axis=1)
        return units_wide.mul(price_proxy, fill_value=0.0)

    def _pool_driver_series(
        self,
        idx: pd.DatetimeIndex,
        units_wide: pd.DataFrame,
        revenue_wide: pd.DataFrame,
        pool: CostPoolInput,
    ) -> pd.Series:
        sku_levels = units_wide.columns.get_level_values(0) if isinstance(units_wide.columns, pd.MultiIndex) else pd.Index([])
        channel_levels = units_wide.columns.get_level_values(1) if isinstance(units_wide.columns, pd.MultiIndex) else pd.Index([])
        revenue_channel_levels = revenue_wide.columns.get_level_values(1) if isinstance(revenue_wide.columns, pd.MultiIndex) else pd.Index([])

        if pool.allocation_driver == "units":
            return units_wide.sum(axis=1).reindex(idx).fillna(0.0) if not units_wide.empty else pd.Series(0.0, index=idx)
        if pool.allocation_driver == "liters":
            liters = pd.Series(0.0, index=idx)
            if not units_wide.empty:
                for sku_id in self.inputs.skus["sku_id"].tolist():
                    if sku_id not in sku_levels:
                        continue
                    liters_per = self._liters_per_unit_from_name(str(self.inputs.skus.set_index("sku_id").loc[sku_id, "name"]))
                    sku_units = units_wide.xs(sku_id, axis=1, level=0, drop_level=False).sum(axis=1)
                    liters = liters + sku_units * liters_per
            return liters
        if pool.allocation_driver == "revenue":
            return revenue_wide.sum(axis=1).reindex(idx).fillna(0.0) if not revenue_wide.empty else pd.Series(0.0, index=idx)
        if pool.allocation_driver == "channel_units":
            if units_wide.empty:
                return pd.Series(0.0, index=idx)
            if pool.channel and pool.channel in channel_levels:
                return units_wide.xs(pool.channel, axis=1, level=1, drop_level=False).sum(axis=1)
            return units_wide.sum(axis=1)
        if pool.allocation_driver == "channel_revenue":
            if revenue_wide.empty:
                return pd.Series(0.0, index=idx)
            if pool.channel and pool.channel in revenue_channel_levels:
                return revenue_wide.xs(pool.channel, axis=1, level=1, drop_level=False).sum(axis=1)
            return revenue_wide.sum(axis=1)
        if pool.allocation_driver == "active_sku":
            if units_wide.empty:
                return pd.Series(0.0, index=idx)
            # pandas>=3 removed axis=1 on DataFrame.groupby; group columns through transpose.
            sku_units = units_wide.T.groupby(level=0).sum().T
            active = (sku_units > 0).sum(axis=1)
            return active.astype(float)
        if pool.allocation_driver == "complexity":
            if units_wide.empty:
                return pd.Series(0.0, index=idx)
            sku_weights = self.inputs.skus.set_index("sku_id")["relative_opex_weight"].to_dict()
            totals = pd.Series(0.0, index=idx)
            for sku_id, w in sku_weights.items():
                if sku_id not in sku_levels:
                    continue
                sku_units = units_wide.xs(sku_id, axis=1, level=0, drop_level=False).sum(axis=1)
                totals = totals + (sku_units > 0).astype(float) * float(w)
            return totals
        return pd.Series(0.0, index=idx)

    def _cost_pool_monthly_by_type(
        self,
        idx: pd.DatetimeIndex,
        units_wide: pd.DataFrame,
        revenue_wide: pd.DataFrame,
        cost_type: Literal["direct", "indirect"],
    ) -> pd.DataFrame:
        excluded_names: set[str] = set()
        if cost_type == "direct" and self.inputs.direct_labor_schedule is not None and not self.inputs.direct_labor_schedule.empty:
            excluded_names.add("production direct labor")
        if cost_type == "indirect" and self.inputs.indirect_labor_schedule is not None and not self.inputs.indirect_labor_schedule.empty:
            excluded_names.add("indirect labor")
        pools = [
            p
            for p in (self.inputs.cost_pools or [])
            if p.cost_type == cost_type and str(p.name).strip().lower() not in excluded_names
        ]
        data: Dict[str, pd.Series] = {}
        for pool in pools:
            driver = self._pool_driver_series(idx, units_wide, revenue_wide, pool)
            variable_component = float(pool.unit_variable_cost) * driver
            fixed_component = pd.Series(float(pool.fixed_monthly_cost), index=idx)
            if pool.behavior == "variable":
                total = variable_component
            elif pool.behavior == "fixed":
                total = fixed_component
            elif pool.behavior == "step_fixed":
                steps = np.floor(driver / max(float(pool.step_threshold), 1.0))
                total = fixed_component + steps * float(pool.step_increment)
            else:
                steps = np.floor(driver / max(float(pool.step_threshold), 1.0)) if float(pool.step_threshold) > 0 else 0.0
                total = fixed_component + variable_component + (steps * float(pool.step_increment))
            data[pool.name] = total.astype(float)
        return pd.DataFrame(data, index=idx)

    def _indirect_cost_pool_monthly(self, idx: pd.DatetimeIndex, units_wide: pd.DataFrame, revenue_wide: pd.DataFrame) -> pd.DataFrame:
        return self._cost_pool_monthly_by_type(idx, units_wide, revenue_wide, "indirect")

    def _direct_cost_pool_monthly(self, idx: pd.DatetimeIndex, units_wide: pd.DataFrame, revenue_wide: pd.DataFrame) -> pd.DataFrame:
        return self._cost_pool_monthly_by_type(idx, units_wide, revenue_wide, "direct")

    def _allocate_monthly_pool_to_sku(
        self,
        idx: pd.DatetimeIndex,
        units_wide: pd.DataFrame,
        pool_monthly: pd.DataFrame,
    ) -> pd.DataFrame:
        sku_index = self.inputs.skus["sku_id"]
        if pool_monthly.empty or units_wide.empty:
            return pd.DataFrame(0.0, index=idx, columns=sku_index)
        sku_units = units_wide.T.groupby(level=0).sum().T.reindex(idx).fillna(0.0)
        sku_units = sku_units.reindex(columns=sku_index, fill_value=0.0)
        weights = self.inputs.skus.set_index("sku_id")["relative_opex_weight"].reindex(sku_index).fillna(1.0)
        weighted = sku_units.mul(weights, axis=1)
        denom = weighted.sum(axis=1).replace(0.0, np.nan)
        shares = weighted.div(denom, axis=0).fillna(0.0)
        total_pool = pool_monthly.sum(axis=1).reindex(idx).fillna(0.0)
        return shares.mul(total_pool, axis=0)

    def _derived_direct_cost_per_unit_by_sku(
        self,
        idx: pd.DatetimeIndex,
        units_wide: pd.DataFrame,
        include_scheduled_labor: bool = True,
    ) -> pd.Series:
        skus = self.inputs.skus.set_index("sku_id")
        revenue_proxy = self._estimated_revenue_wide_from_units(idx, units_wide)
        direct_pool_monthly = self._direct_cost_pool_monthly(idx, units_wide, revenue_proxy)
        allocated = self._allocate_monthly_pool_to_sku(idx, units_wide, direct_pool_monthly)
        if include_scheduled_labor:
            labor_alloc, _ = self._labor_schedule_monthly(idx, units_wide, self.inputs.direct_labor_schedule)
            allocated = allocated.add(labor_alloc, fill_value=0.0)

        basis_m = int(np.clip(self.cfg.pricing_cost_basis_month, 0, len(idx) - 1))
        basis_date = idx[basis_m]
        if units_wide.empty:
            units_sku = pd.Series(0.0, index=skus.index)
        else:
            units_sku = units_wide.loc[basis_date].groupby(level=0).sum().reindex(skus.index).fillna(0.0)
        per_unit = allocated.loc[basis_date].reindex(skus.index).fillna(0.0) / units_sku.replace(0.0, np.nan)
        per_unit = per_unit.fillna(0.0)
        overrides = pd.to_numeric(skus.get("direct_cost_per_unit_override"), errors="coerce")
        per_unit = per_unit.where(~overrides.notna(), overrides)
        per_unit = per_unit.fillna(0.0).rename("direct_cost_per_unit")
        if include_scheduled_labor:
            self.inputs.skus = self.inputs.skus.set_index("sku_id")
            self.inputs.skus["direct_cost_per_unit"] = per_unit
            self.inputs.skus = self.inputs.skus.reset_index()
        return per_unit

    # ---------- pricing ----------
    @staticmethod
    def _liters_per_unit_from_name(sku_name: str) -> float:
        text = str(sku_name).lower()
        m_ml = re.search(r"(\\d+(?:\\.\\d+)?)\\s*ml", text)
        if m_ml:
            return float(m_ml.group(1)) / 1000.0
        m_l = re.search(r"(\\d+(?:\\.\\d+)?)\\s*l", text)
        if m_l:
            return float(m_l.group(1))
        if "keg" in text:
            return 20.0
        return 0.5

    def _allocate_basis_opex_by_pool(
        self,
        basis_date: pd.Timestamp,
        units_wide: pd.DataFrame,
        channels: pd.DataFrame,
        skus: pd.DataFrame,
        opex_basis: float,
    ) -> pd.Series:
        units_row = units_wide.loc[basis_date] if not units_wide.empty else pd.Series(dtype=float)
        channel_price_factor = channels.set_index("channel")["price_factor"].to_dict()

        contexts: List[SKUCostContext] = []
        for sku_id, row in skus.iterrows():
            channel_units = {}
            for ch in channels["channel"].tolist():
                channel_units[ch] = float(units_row.get((sku_id, ch), 0.0))
            total_units = float(sum(channel_units.values()))
            liters_per_unit = self._liters_per_unit_from_name(str(row.get("name", "")))
            estimated_base_unit_price = float(row["direct_cost_per_unit"]) * (1.0 + float(row["markup_pct"]))
            channel_revenue = {
                ch: channel_units[ch] * estimated_base_unit_price * float(channel_price_factor.get(ch, 1.0))
                for ch in channel_units
            }
            contexts.append(
                SKUCostContext(
                    sku_id=int(sku_id),
                    sku_name=str(row.get("name", f"SKU {sku_id}")),
                    product_family=str(row.get("product_family", "Core")),
                    package_type=str(row.get("package_type", "Standard")),
                    package_size=str(row.get("package_size", "Standard")),
                    active=total_units > 0.0,
                    units_sold_by_year={"Basis": total_units},
                    liters_sold_by_year={"Basis": total_units * liters_per_unit},
                    revenue_by_year={"Basis": float(sum(channel_revenue.values()))},
                    revenue_by_channel_by_year={"Basis": channel_revenue},
                    units_by_channel_by_year={"Basis": channel_units},
                    complexity_score=float(row.get("relative_opex_weight", 1.0)),
                    batch_count_by_year={"Basis": (total_units * liters_per_unit) / 500.0},
                    order_count_by_year={"Basis": total_units / 1000.0},
                    shipment_count_by_year={"Basis": total_units / 1200.0},
                )
            )

        pools = build_default_opex_cost_pools({"Basis": float(max(opex_basis, 0.0))})
        report = allocate_opex_by_drivers(["Basis"], contexts, pools)
        allocated = {a.sku_id: a.total_allocated_opex for a in report.allocations if a.year == "Basis"}
        return pd.Series(allocated, index=skus.index).fillna(0.0)

    def _base_cost_plus_price_by_sku(self, idx: pd.DatetimeIndex, units_wide: pd.DataFrame) -> pd.Series:
        """
        Compute a base (month-0) cost-plus price for each SKU using a cost basis month:

            base_price_sku =
                (direct_cost_per_unit_basis + allocated_opex_per_unit_basis) * (1 + markup_pct)

        Returns Series indexed by sku_id.
        """
        skus = self.inputs.skus.set_index("sku_id")
        cost_index = self._inflation_index(self.cfg.cost_inflation_annual, idx)

        basis_m = int(np.clip(self.cfg.pricing_cost_basis_month, 0, len(idx) - 1))
        basis_date = idx[basis_m]

        # total units per SKU at cost basis month
        if units_wide.empty:
            units_sku = pd.Series(0.0, index=skus.index)
        else:
            units_sku = units_wide.loc[basis_date].groupby(level=0).sum()
        units_sku = units_sku.reindex(skus.index).fillna(0.0)

        # OPEX cash in basis month from indirect cost pools (no monthly lump assumption).
        revenue_proxy = self._estimated_revenue_wide_from_units(idx, units_wide)
        pool_monthly = self._indirect_cost_pool_monthly(idx, units_wide, revenue_proxy)
        opex_basis = float(pool_monthly.sum(axis=1).reindex(idx).fillna(0.0).loc[basis_date] * cost_index.loc[basis_date])

        # Allocate OPEX via cost-pool allocator (no lump-style weighted-smear).
        alloc = self._allocate_basis_opex_by_pool(
            basis_date=basis_date,
            units_wide=units_wide,
            channels=self.inputs.channels,
            skus=skus,
            opex_basis=opex_basis,
        )
        opex_per_unit = alloc / units_sku.replace(0.0, np.nan)
        opex_per_unit = opex_per_unit.fillna(0.0)

        derived_direct_cost_per_unit = self._derived_direct_cost_per_unit_by_sku(idx, units_wide)
        direct_cost_basis = derived_direct_cost_per_unit * cost_index.loc[basis_date]
        total_cost_basis = direct_cost_basis + opex_per_unit
        base_price = total_cost_basis * (1.0 + skus["markup_pct"])

        # If a SKU has no volume in basis month, fall back to direct-cost-plus only
        fallback = direct_cost_basis * (1.0 + skus["markup_pct"])
        base_price = base_price.where(base_price > 0.0, fallback)

        return base_price.rename("base_price_per_unit")

    def _prices_matrix(self, idx: pd.DatetimeIndex, units_wide: pd.DataFrame) -> pd.DataFrame:
        """
        Returns a wide DataFrame of prices with MultiIndex columns (sku_id, channel).
        """
        channels = self.inputs.channels.set_index("channel")
        price_index = self._inflation_index(self.cfg.price_inflation_annual, idx)
        base_price = self._base_cost_plus_price_by_sku(idx, units_wide)

        cols: List[Tuple[int, str]] = []
        data: List[np.ndarray] = []
        for sku_id, base_p in base_price.items():
            for channel, row in channels.iterrows():
                cols.append((sku_id, channel))
                series = base_p * price_index.values * float(row["price_factor"])
                data.append(series)

        prices = pd.DataFrame(
            np.array(data).T,
            index=idx,
            columns=pd.MultiIndex.from_tuples(cols, names=["sku_id", "channel"]),
        )
        return prices.sort_index(axis=1)

    # ---------- CAPEX & depreciation ----------
    def _capex_series(self, idx: pd.DatetimeIndex) -> pd.Series:
        capex = pd.Series(0.0, index=idx, name="capex")
        for item in self.inputs.capex_items:
            m = int(item.capex_month)
            if 0 <= m < len(idx):
                capex.iloc[m] += float(item.amount)
        return capex

    def _depreciation_series(self, idx: pd.DatetimeIndex) -> pd.Series:
        dep = pd.Series(0.0, index=idx, name="depreciation")
        for item in self.inputs.capex_items:
            years = float(item.depreciation_years)
            if years <= 0.0:
                continue
            m0 = int(item.capex_month)
            n = int(round(years * 12))
            if n <= 0:
                continue
            monthly_dep = float(item.amount) / n
            for k in range(n):
                m = m0 + k
                if 0 <= m < len(idx):
                    dep.iloc[m] += monthly_dep
        return dep

    def _net_fixed_assets(self, capex: pd.Series, dep: pd.Series) -> pd.Series:
        gross = capex.cumsum()
        accum_dep = dep.cumsum()
        return (gross - accum_dep).rename("net_fixed_assets")

    # ---------- Working capital ----------
    def _nwc(self, idx: pd.DatetimeIndex, revenue: pd.Series, direct_costs: pd.Series) -> Tuple[pd.DataFrame, pd.Series]:
        """
        Computes NWC components and change in NWC.
        """
        days_year = 365.0
        ar = revenue * (self.cfg.days_receivables / days_year)
        inv = direct_costs * (self.cfg.days_inventory / days_year)
        ap = direct_costs * (self.cfg.days_payables / days_year)

        oca = revenue * float(self.cfg.other_current_assets_pct_revenue)
        ocl = direct_costs * float(self.cfg.other_current_liabilities_pct_direct_costs)

        nwc = ar + inv + oca - ap - ocl
        chg = nwc.diff().fillna(nwc.iloc[0]).rename("change_in_nwc")

        comp = pd.DataFrame(
            {
                "receivables": ar,
                "inventory": inv,
                "other_current_assets": oca,
                "payables": ap,
                "other_current_liabilities": ocl,
                "net_working_capital": nwc,
            },
            index=idx,
        )
        return comp, chg

    # ---------- Debt ----------
    def _debt_schedule_one(self, idx: pd.DatetimeIndex, fac: DebtFacility) -> pd.DataFrame:
        """
        Build a month-by-month debt schedule for a single facility.
        Columns:
            beginning_balance, draw, interest, principal_payment, ending_balance
        """
        n = len(idx)
        m_draw = int(fac.draw_month)
        grace = int(fac.grace_months)
        term = int(fac.term_months)
        r_m = annual_to_monthly_rate(fac.annual_interest_rate)

        beg = np.zeros(n)
        draw = np.zeros(n)
        interest = np.zeros(n)
        principal = np.zeros(n)
        end = np.zeros(n)

        for t in range(n):
            beg[t] = end[t - 1] if t > 0 else 0.0

            if t == m_draw:
                draw[t] = float(fac.principal)
                beg[t] += draw[t]  # treat draw as immediately outstanding for interest calc

            # Interest on beginning balance (post-draw if draw this month)
            interest[t] = beg[t] * r_m

            repay_start = m_draw + grace
            repay_end = m_draw + term  # exclusive

            if fac.repayment_type == "specified":
                p = 0.0
                if fac.specified_principal_payments and t in fac.specified_principal_payments:
                    p = float(fac.specified_principal_payments[t])
                principal[t] = min(p, beg[t])

            elif t < repay_start or t >= repay_end:
                principal[t] = 0.0

            else:
                outstanding = beg[t]

                if fac.repayment_type == "linear":
                    amort_months = max(repay_end - repay_start, 1)
                    pmt_principal = float(fac.principal) / amort_months
                    principal[t] = min(pmt_principal, outstanding)

                elif fac.repayment_type == "annuity":
                    amort_months = max(repay_end - repay_start, 1)
                    P = float(fac.principal)
                    if abs(r_m) < 1e-12:
                        total_pmt = P / amort_months
                    else:
                        total_pmt = P * r_m / (1.0 - (1.0 + r_m) ** (-amort_months))
                    principal[t] = min(max(total_pmt - interest[t], 0.0), outstanding)

                elif fac.repayment_type == "interest_only_then_linear":
                    amort_months = max(repay_end - repay_start, 1)
                    pmt_principal = float(fac.principal) / amort_months
                    principal[t] = min(pmt_principal, outstanding)

                else:
                    principal[t] = 0.0

            end[t] = max(beg[t] - principal[t], 0.0)

        return pd.DataFrame(
            {
                "beginning_balance": beg,
                "draw": draw,
                "interest": interest,
                "principal_payment": principal,
                "ending_balance": end,
            },
            index=idx,
        )

    def _debt_schedules(self, idx: pd.DatetimeIndex) -> Dict[str, pd.DataFrame]:
        schedules: Dict[str, pd.DataFrame] = {}
        for fac in self.inputs.debt_facilities:
            schedules[fac.name] = self._debt_schedule_one(idx, fac)
        return schedules

    def _annual_opex_allocation_views(
        self,
        idx: pd.DatetimeIndex,
        units_wide: pd.DataFrame,
        revenue_wide: pd.DataFrame,
        opex_series: pd.Series,
    ) -> Dict[str, pd.DataFrame]:
        skus = self.inputs.skus.set_index("sku_id")
        channels = self.inputs.channels["channel"].tolist()
        year_labels = sorted(idx.year.unique())
        year_keys = {y: f"Year {i + 1}" for i, y in enumerate(year_labels)}

        annual_opex = {}
        for y in year_labels:
            annual_opex[year_keys[y]] = float(opex_series[idx.year == y].sum())

        contexts: List[SKUCostContext] = []
        for sku_id, row in skus.iterrows():
            units_by_year = {}
            liters_by_year = {}
            revenue_by_year = {}
            revenue_ch_by_year: Dict[str, Dict[str, float]] = {}
            units_ch_by_year: Dict[str, Dict[str, float]] = {}
            batch_by_year = {}
            order_by_year = {}
            ship_by_year = {}
            liters_per_unit = self._liters_per_unit_from_name(str(row.get("name", "")))
            for y in year_labels:
                key = year_keys[y]
                mask = idx.year == y
                units_year = units_wide.loc[mask] if not units_wide.empty else pd.DataFrame(index=idx[mask])
                rev_year = revenue_wide.loc[mask] if not revenue_wide.empty else pd.DataFrame(index=idx[mask])
                ch_units = {ch: float(units_year.get((sku_id, ch), pd.Series(dtype=float)).sum()) for ch in channels}
                ch_revenue = {ch: float(rev_year.get((sku_id, ch), pd.Series(dtype=float)).sum()) for ch in channels}
                total_units = float(sum(ch_units.values()))
                total_revenue = float(sum(ch_revenue.values()))
                units_by_year[key] = total_units
                liters_by_year[key] = total_units * liters_per_unit
                revenue_by_year[key] = total_revenue
                units_ch_by_year[key] = ch_units
                revenue_ch_by_year[key] = ch_revenue
                batch_by_year[key] = (total_units * liters_per_unit) / 500.0
                order_by_year[key] = total_units / 1000.0
                ship_by_year[key] = total_units / 1200.0

            contexts.append(
                SKUCostContext(
                    sku_id=int(sku_id),
                    sku_name=str(row.get("name", f"SKU {sku_id}")),
                    product_family=str(row.get("product_family", "Core")),
                    package_type=str(row.get("package_type", "Standard")),
                    package_size=str(row.get("package_size", "Standard")),
                    active=any(v > 0.0 for v in units_by_year.values()),
                    units_sold_by_year=units_by_year,
                    liters_sold_by_year=liters_by_year,
                    revenue_by_year=revenue_by_year,
                    revenue_by_channel_by_year=revenue_ch_by_year,
                    units_by_channel_by_year=units_ch_by_year,
                    complexity_score=float(row.get("relative_opex_weight", 1.0)),
                    batch_count_by_year=batch_by_year,
                    order_count_by_year=order_by_year,
                    shipment_count_by_year=ship_by_year,
                )
            )

        pools: List[OpexCostPool] = build_default_opex_cost_pools(annual_opex)
        report = allocate_opex_by_drivers(list(annual_opex.keys()), contexts, pools)
        pool_view = pd.DataFrame({s.year: s.by_pool_totals for s in report.summaries}).T
        driver_view = pd.DataFrame({s.year: s.by_driver_type_totals for s in report.summaries}).T
        rec_view = pd.DataFrame(
            {
                s.year: {
                    "total_pool_opex": s.total_pool_opex,
                    "total_allocated_opex": s.total_allocated_opex,
                    "reconciliation_gap": s.reconciliation_gap,
                }
                for s in report.summaries
            }
        ).T
        product_view = pd.DataFrame(
            [
                {
                    "year": a.year,
                    "sku_id": a.sku_id,
                    "sku_name": a.sku_name,
                    "total_allocated_opex": a.total_allocated_opex,
                    "opex_per_unit": a.opex_per_unit,
                    "opex_per_liter": a.opex_per_liter,
                    "opex_per_case": a.opex_per_case,
                }
                for a in report.allocations
            ]
        )
        return {
            "pool_view": pool_view,
            "driver_view": driver_view,
            "product_view": product_view,
            "reconciliation_view": rec_view,
        }

    # ---------- Run model ----------
    def run(self) -> ModelRunResult:
        idx = self._timeline()

        # Inflation indices
        cost_idx = self._inflation_index(self.cfg.cost_inflation_annual, idx)

        # Units, operations, prices, revenue
        demand_units_wide = self._units_matrix(idx)
        liters_per_unit = self.inputs.skus.set_index("sku_id")["name"].map(self._liters_per_unit_from_name)
        operations_plan = plan_brewery_operations(
            idx=idx,
            demand_units_wide=demand_units_wide,
            skus=self.inputs.skus,
            liters_per_unit=liters_per_unit,
            sku_operations=self.inputs.sku_operations,
            brewhouse_schedule=self.inputs.brewhouse_schedule,
            cellar_schedule=self.inputs.cellar_schedule,
            packaging_schedule=self.inputs.packaging_schedule,
        )
        units_wide = self._scale_channel_units_to_sku_totals(
            demand_units_wide,
            operations_plan.shipment_units_by_sku,
        )
        production_units_wide = self._scale_channel_units_to_sku_totals(
            demand_units_wide,
            operations_plan.production_units_by_sku,
        )
        prices_wide = self._prices_matrix(idx, units_wide)
        if isinstance(units_wide.columns, pd.MultiIndex):
            units_wide.columns = units_wide.columns.set_names(["sku_id", "channel"])
        if isinstance(production_units_wide.columns, pd.MultiIndex):
            production_units_wide.columns = production_units_wide.columns.set_names(["sku_id", "channel"])
        if isinstance(prices_wide.columns, pd.MultiIndex):
            prices_wide.columns = prices_wide.columns.set_names(["sku_id", "channel"])

        if units_wide.empty:
            revenue_wide = pd.DataFrame(index=idx, columns=units_wide.columns)
        else:
            revenue_wide = units_wide.mul(prices_wide, fill_value=0.0)
        net_revenue, revenue_breakdown, receivables_details = self._revenue_schedule_adjustments(idx, revenue_wide)
        gross_revenue = revenue_breakdown.get("gross_revenue", revenue_wide.sum(axis=1)).rename("gross_revenue")

        other_income = self._other_income_series(idx)
        other_income = (other_income * cost_idx).rename("other_income")  # default: inflate with costs
        total_revenue = (net_revenue + other_income).rename("total_revenue")

        # Direct costs
        skus = self.inputs.skus.set_index("sku_id")
        production_basis_wide = production_units_wide if not production_units_wide.empty else units_wide
        derived_direct_cost_per_unit = self._derived_direct_cost_per_unit_by_sku(
            idx,
            production_basis_wide,
            include_scheduled_labor=False,
        )
        if production_basis_wide.empty:
            direct_material_costs = pd.Series(0.0, index=idx, name="direct_material_costs")
        else:
            cost_cols = []
            cost_data = []
            for sku_id, row in skus.iterrows():
                for channel in self.inputs.channels["channel"].tolist():
                    cost_cols.append((sku_id, channel))
                    series = float(derived_direct_cost_per_unit.get(sku_id, 0.0)) * cost_idx.values
                    cost_data.append(series)
            costs_wide = pd.DataFrame(
                np.array(cost_data).T,
                index=idx,
                columns=pd.MultiIndex.from_tuples(cost_cols, names=["sku_id", "channel"]),
            ).sort_index(axis=1)
            direct_costs_wide = production_basis_wide.mul(costs_wide, fill_value=0.0)
            direct_material_costs = direct_costs_wide.sum(axis=1).rename("direct_material_costs")

        direct_labor_alloc, direct_labor_summary = self._labor_schedule_monthly(
            idx,
            production_basis_wide,
            self.inputs.direct_labor_schedule,
        )
        direct_labor_cost = direct_labor_summary["labor_cost"].rename("direct_labor_cost")
        inventory_details, inventory_writeoff, inventory_reserve = self._inventory_schedule_components(
            idx,
            direct_material_costs,
        )
        payables_details, payable_discount_benefit = self._payables_schedule_components(
            idx,
            direct_material_costs,
        )
        direct_costs = (
            direct_material_costs
            + direct_labor_cost
            + inventory_writeoff
            - payable_discount_benefit
        ).rename("direct_costs")
        gross_profit = (net_revenue - direct_costs).rename("gross_profit")

        # Indirect operating costs from explicit cost pools plus indirect labor and bad-debt expense.
        pool_monthly = self._indirect_cost_pool_monthly(idx, units_wide, revenue_wide)
        _, indirect_labor_summary = self._labor_schedule_monthly(
            idx,
            units_wide,
            self.inputs.indirect_labor_schedule,
        )
        indirect_labor_cost = indirect_labor_summary["labor_cost"].rename("indirect_labor_cost")
        bad_debt_expense = revenue_breakdown.get(
            "bad_debt_expense",
            pd.Series(0.0, index=idx),
        ).rename("bad_debt_expense")
        opex = (
            (pool_monthly.sum(axis=1) * cost_idx)
            + indirect_labor_cost
            + bad_debt_expense
        ).rename("opex")
        ebitda = (total_revenue - direct_costs - opex).rename("ebitda")

        # CAPEX + depreciation + net fixed assets
        capex = self._capex_series(idx)
        dep = self._depreciation_series(idx)
        net_fixed_assets = self._net_fixed_assets(capex, dep)
        ebit = (ebitda - dep).rename("ebit")

        # Working capital
        receivables = receivables_details.get(
            "receivables",
            pd.Series(0.0, index=idx),
        ).rename("receivables")
        inventory = inventory_details.get(
            "inventory",
            pd.Series(0.0, index=idx),
        ).rename("inventory")
        payables = payables_details.get(
            "payables",
            pd.Series(0.0, index=idx),
        ).rename("payables")
        other_current_assets = (net_revenue * float(self.cfg.other_current_assets_pct_revenue)).rename(
            "other_current_assets"
        )
        other_current_liabilities = (
            direct_material_costs * float(self.cfg.other_current_liabilities_pct_direct_costs)
        ).rename("other_current_liabilities")
        inventory_stage_cols = [
            c for c in inventory_details.columns
            if c.startswith("inventory_") and c not in {"inventory", "inventory_reserve"}
        ]
        net_working_capital = (
            receivables + inventory + other_current_assets - payables - other_current_liabilities
        ).rename("net_working_capital")
        change_nwc = net_working_capital.diff().fillna(net_working_capital.iloc[0]).rename("change_in_nwc")
        nwc_comp = pd.DataFrame(
            {
                "receivables": receivables,
                "inventory": inventory,
                "inventory_reserve": inventory_reserve,
                "other_current_assets": other_current_assets,
                "payables": payables,
                "other_current_liabilities": other_current_liabilities,
                "net_working_capital": net_working_capital,
            },
            index=idx,
        ).join(
            inventory_details.reindex(columns=inventory_stage_cols, fill_value=0.0)
            if inventory_stage_cols
            else pd.DataFrame(index=idx)
        )

        # Debt schedules
        debt_schedules = self._debt_schedules(idx)
        if debt_schedules:
            debt_interest = sum(df["interest"] for df in debt_schedules.values()).rename("interest_expense")
            debt_draw = sum(df["draw"] for df in debt_schedules.values()).rename("debt_draw")
            debt_principal = sum(df["principal_payment"] for df in debt_schedules.values()).rename("debt_principal_payment")
            debt_balance = sum(df["ending_balance"] for df in debt_schedules.values()).rename("debt_ending_balance")
        else:
            debt_interest = pd.Series(0.0, index=idx, name="interest_expense")
            debt_draw = pd.Series(0.0, index=idx, name="debt_draw")
            debt_principal = pd.Series(0.0, index=idx, name="debt_principal_payment")
            debt_balance = pd.Series(0.0, index=idx, name="debt_ending_balance")

        # Equity injections
        equity_inj = pd.Series(0.0, index=idx, name="equity_injection")
        for m, amt in (self.inputs.equity_injections or {}).items():
            m = int(m)
            if 0 <= m < len(idx):
                equity_inj.iloc[m] += float(amt)

        # Sequential taxes, revolver, cash, and dividends
        revolver_rate_m = annual_to_monthly_rate(self.cfg.revolver_interest_annual)
        pre_tax_income = pd.Series(0.0, index=idx, name="pre_tax_income")
        taxes = pd.Series(0.0, index=idx, name="taxes")
        net_income = pd.Series(0.0, index=idx, name="net_income")
        revolver_interest = pd.Series(0.0, index=idx, name="revolver_interest_expense")
        revolver_draw = pd.Series(0.0, index=idx, name="revolver_draw")
        revolver_principal = pd.Series(0.0, index=idx, name="revolver_principal_payment")
        revolver_balance = pd.Series(0.0, index=idx, name="revolver_ending_balance")
        cfo = pd.Series(0.0, index=idx, name="cash_flow_from_operations")
        cfi = (-capex).rename("cash_flow_from_investing")
        cff = pd.Series(0.0, index=idx, name="cash_flow_from_financing")
        dividends = pd.Series(0.0, index=idx, name="dividends")
        cash = pd.Series(0.0, index=idx, name="cash")
        cash_prev = float(self.cfg.initial_cash)
        revolver_prev = 0.0

        for t, _date in enumerate(idx):
            revolver_interest_t = revolver_prev * revolver_rate_m
            revolver_interest.iloc[t] = revolver_interest_t
            pre_tax_t = float(ebit.iloc[t] - debt_interest.iloc[t] - revolver_interest_t)
            tax_t = max(pre_tax_t, 0.0) * float(self.cfg.tax_rate)
            net_income_t = pre_tax_t - tax_t
            pre_tax_income.iloc[t] = pre_tax_t
            taxes.iloc[t] = tax_t
            net_income.iloc[t] = net_income_t

            cfo_t = net_income_t + float(dep.iloc[t]) - float(change_nwc.iloc[t])
            cfo.iloc[t] = cfo_t
            cash_pre = cash_prev + float(cfo_t + cfi.iloc[t] + equity_inj.iloc[t] + debt_draw.iloc[t] - debt_principal.iloc[t])

            draw_t = 0.0
            repay_t = 0.0
            if cash_pre < float(self.cfg.revolver_target_cash):
                draw_t = min(
                    float(self.cfg.revolver_target_cash) - cash_pre,
                    max(float(self.cfg.revolver_limit) - revolver_prev, 0.0),
                )
            elif cash_pre > float(self.cfg.revolver_target_cash) and revolver_prev > 0.0:
                repay_t = min(cash_pre - float(self.cfg.revolver_target_cash), revolver_prev)
            revolver_post = revolver_prev + draw_t - repay_t
            cash_post_revolver = cash_pre + draw_t - repay_t

            div_t = 0.0
            if self.div.enabled and t >= int(self.div.start_month):
                if self.div.model == "cash_sweep":
                    target_floor = max(
                        float(self.div.minimum_cash_position),
                        float(self.cfg.revolver_target_cash),
                    )
                    div_t = max(cash_post_revolver - target_floor, 0.0)
                else:
                    div_t = float(self.div.payout_ratio) * max(net_income_t, 0.0)

            cash_end = cash_post_revolver - div_t
            dividends.iloc[t] = div_t
            revolver_draw.iloc[t] = draw_t
            revolver_principal.iloc[t] = repay_t
            revolver_balance.iloc[t] = revolver_post
            cff.iloc[t] = (
                float(equity_inj.iloc[t])
                + float(debt_draw.iloc[t])
                + draw_t
                - float(debt_principal.iloc[t])
                - repay_t
                - div_t
            )
            cash.iloc[t] = cash_end
            cash_prev = cash_end
            revolver_prev = revolver_post

        # FCFF and valuation
        nopat = (ebit * (1.0 - float(self.cfg.tax_rate))).rename("nopat")
        fcff = (nopat + dep - capex - change_nwc).rename("fcff")

        exit_m = self.cfg.exit_month if self.cfg.exit_month is not None else (len(idx) - 1)
        exit_m = int(np.clip(exit_m, 0, len(idx) - 1))
        terminal_value = max(float(ebitda.iloc[exit_m]), 0.0) * float(self.cfg.exit_ev_ebitda_multiple)
        wacc_m = annual_to_monthly_rate(self.cfg.wacc_annual)
        discount = (1.0 + wacc_m) ** np.arange(len(idx))
        enterprise_value = float((fcff.values / discount).sum() + terminal_value / discount[exit_m])
        debt_exit = float((debt_balance + revolver_balance).iloc[exit_m])
        cash_exit = float(cash.iloc[exit_m])
        equity_value_exit = enterprise_value - debt_exit + cash_exit

        # Investor IRR and MOIC (equity injections are cash-in to company => negative to investor)
        equity_cashflows = (-equity_inj).copy()
        equity_cashflows += dividends
        equity_cashflows.iloc[exit_m] += float(equity_value_exit)
        irr_m = irr(equity_cashflows.values, guess=0.02)
        irr_annual = (1.0 + irr_m) ** 12 - 1.0 if np.isfinite(irr_m) else np.nan
        invested = float(equity_inj.sum())
        returned = float(dividends.sum() + max(equity_value_exit, 0.0))
        moic = safe_div(returned, invested, default=np.nan)
        valuation = {
            "wacc_annual": float(self.cfg.wacc_annual),
            "wacc_monthly": float(wacc_m),
            "exit_month_index": float(exit_m),
            "terminal_value": float(terminal_value),
            "enterprise_value_dcf": float(enterprise_value),
            "equity_value_exit": float(equity_value_exit),
            "equity_irr_monthly": float(irr_m) if np.isfinite(irr_m) else np.nan,
            "equity_irr_annual": float(irr_annual) if np.isfinite(irr_annual) else np.nan,
            "equity_moic": float(moic) if np.isfinite(moic) else np.nan,
        }

        total_interest = (debt_interest + revolver_interest).rename("interest_expense_total")
        total_debt_balance = (debt_balance + revolver_balance).rename("total_debt_ending_balance")
        debt_service = (debt_principal + revolver_principal + total_interest).rename("debt_service")
        annualized_ebitda = ebitda.rolling(window=12, min_periods=1).sum()
        dscr = pd.Series(np.where(debt_service > 0, ebitda / debt_service, np.nan), index=idx, name="dscr")
        interest_coverage = pd.Series(
            np.where(total_interest > 0, ebitda / total_interest, np.nan),
            index=idx,
            name="interest_coverage",
        )
        leverage_ratio = pd.Series(
            np.where(annualized_ebitda > 0, total_debt_balance / annualized_ebitda, np.nan),
            index=idx,
            name="leverage_ratio",
        )
        contributed_capital = equity_inj.cumsum().rename("contributed_capital")
        retained_earnings = (net_income - dividends).cumsum().rename("retained_earnings")
        current_assets = (cash + receivables + inventory + other_current_assets).rename("current_assets")
        current_liabilities = (payables + other_current_liabilities + revolver_balance).rename("current_liabilities")
        total_assets = (current_assets + net_fixed_assets).rename("total_assets")
        total_liabilities = (current_liabilities + debt_balance).rename("total_liabilities")
        equity = (total_assets - total_liabilities).rename("equity")
        balance_sheet_gap = (equity - (contributed_capital + retained_earnings)).rename("balance_sheet_gap")
        net_change_in_cash = cash.diff().fillna(cash.iloc[0] - float(self.cfg.initial_cash)).rename("net_change_in_cash")

        # Statements
        monthly = pd.DataFrame(
            {
                "revenue": net_revenue,
                "gross_revenue": gross_revenue,
                "net_revenue": net_revenue,
                "trade_spend": revenue_breakdown.get("trade_spend", pd.Series(0.0, index=idx)),
                "returns_allowances": revenue_breakdown.get("returns_allowances", pd.Series(0.0, index=idx)),
                "other_income": other_income,
                "total_revenue": total_revenue,
                "direct_material_costs": direct_material_costs,
                "direct_labor_cost": direct_labor_cost,
                "inventory_writeoff": inventory_writeoff,
                "payable_discount_benefit": payable_discount_benefit,
                "direct_costs": direct_costs,
                "gross_profit": gross_profit,
                "indirect_labor_cost": indirect_labor_cost,
                "bad_debt_expense": bad_debt_expense,
                "opex": opex,
                "ebitda": ebitda,
                "depreciation": dep,
                "ebit": ebit,
                "term_debt_interest_expense": debt_interest,
                "revolver_interest_expense": revolver_interest,
                "interest_expense": total_interest,
                "interest_expense_total": total_interest,
                "pre_tax_income": pre_tax_income,
                "taxes": taxes,
                "net_income": net_income,
                "capex": capex,
                "change_in_nwc": change_nwc,
                "debt_draw": debt_draw,
                "debt_principal_payment": debt_principal,
                "revolver_draw": revolver_draw,
                "revolver_principal_payment": revolver_principal,
                "equity_injection": equity_inj,
                "dividends": dividends,
                "cash_flow_from_operations": cfo,
                "cash_flow_from_investing": cfi,
                "cash_flow_from_financing": cff,
                "net_change_in_cash": net_change_in_cash,
                "cash": cash,
                "debt_ending_balance": debt_balance,
                "revolver_ending_balance": revolver_balance,
                "total_debt_ending_balance": total_debt_balance,
                "net_fixed_assets": net_fixed_assets,
                "current_assets": current_assets,
                "total_assets": total_assets,
                "current_liabilities": current_liabilities,
                "total_liabilities": total_liabilities,
                "contributed_capital": contributed_capital,
                "retained_earnings": retained_earnings,
                "equity": equity,
                "balance_sheet_gap": balance_sheet_gap,
                "required_liters": direct_labor_summary["required_liters"],
                "capacity_liters": direct_labor_summary["capacity_liters"],
                "capacity_shortfall_liters": direct_labor_summary["capacity_shortfall_liters"],
                "temporary_labor_cost": direct_labor_summary["temporary_labor_cost"],
                "shipment_demand_liters": operations_plan.monthly_summary.get("demand_liters", pd.Series(0.0, index=idx)),
                "production_target_packaged_liters": operations_plan.monthly_summary.get("required_packaged_liters", pd.Series(0.0, index=idx)),
                "actual_packaged_liters": operations_plan.monthly_summary.get("actual_packaged_liters", pd.Series(0.0, index=idx)),
                "actual_shipped_liters": operations_plan.monthly_summary.get("actual_shipped_liters", pd.Series(0.0, index=idx)),
                "ops_ending_fg_liters": operations_plan.monthly_summary.get("ending_fg_liters", pd.Series(0.0, index=idx)),
                "unmet_demand_liters": operations_plan.monthly_summary.get("unmet_demand_liters", pd.Series(0.0, index=idx)),
                "brewhouse_capacity_liters": operations_plan.monthly_summary.get("brewhouse_capacity_liters", pd.Series(0.0, index=idx)),
                "cellar_capacity_liters": operations_plan.monthly_summary.get("cellar_capacity_liters", pd.Series(0.0, index=idx)),
                "packaging_capacity_liters": operations_plan.monthly_summary.get("packaging_capacity_liters", pd.Series(0.0, index=idx)),
                "operations_capacity_scale": operations_plan.monthly_summary.get("capacity_scale", pd.Series(1.0, index=idx)),
                "fg_target_gap_liters": operations_plan.monthly_summary.get("fg_target_gap_liters", pd.Series(0.0, index=idx)),
                "debt_service": debt_service,
                "dscr": dscr,
                "interest_coverage": interest_coverage,
                "leverage_ratio": leverage_ratio,
                "dscr_breach_flag": (dscr < float(self.cfg.min_dscr)).astype(int),
                "interest_coverage_breach_flag": (interest_coverage < float(self.cfg.min_interest_coverage)).astype(int),
                "leverage_breach_flag": (leverage_ratio > float(self.cfg.max_leverage_ratio)).astype(int),
                "fcff": fcff,
            },
            index=idx,
        ).join(nwc_comp)

        flow_cols = [
            "revenue",
            "gross_revenue",
            "net_revenue",
            "trade_spend",
            "returns_allowances",
            "other_income",
            "total_revenue",
            "direct_material_costs",
            "direct_labor_cost",
            "inventory_writeoff",
            "payable_discount_benefit",
            "direct_costs",
            "gross_profit",
            "indirect_labor_cost",
            "bad_debt_expense",
            "opex",
            "ebitda",
            "depreciation",
            "ebit",
            "term_debt_interest_expense",
            "interest_expense",
            "revolver_interest_expense",
            "interest_expense_total",
            "pre_tax_income",
            "taxes",
            "net_income",
            "capex",
            "change_in_nwc",
            "debt_draw",
            "debt_principal_payment",
            "revolver_draw",
            "revolver_principal_payment",
            "equity_injection",
            "dividends",
            "cash_flow_from_operations",
            "cash_flow_from_investing",
            "cash_flow_from_financing",
            "net_change_in_cash",
            "debt_service",
            "fcff",
            "shipment_demand_liters",
            "production_target_packaged_liters",
            "actual_packaged_liters",
            "actual_shipped_liters",
            "unmet_demand_liters",
        ]
        stock_cols = [
            "receivables",
            "inventory",
            "inventory_reserve",
            *inventory_stage_cols,
            "other_current_assets",
            "payables",
            "other_current_liabilities",
            "net_working_capital",
            "cash",
            "debt_ending_balance",
            "revolver_ending_balance",
            "total_debt_ending_balance",
            "net_fixed_assets",
            "current_assets",
            "total_assets",
            "current_liabilities",
            "total_liabilities",
            "contributed_capital",
            "retained_earnings",
            "equity",
            "balance_sheet_gap",
            "required_liters",
            "capacity_liters",
            "capacity_shortfall_liters",
            "ops_ending_fg_liters",
            "brewhouse_capacity_liters",
            "cellar_capacity_liters",
            "packaging_capacity_liters",
            "operations_capacity_scale",
            "fg_target_gap_liters",
            "dscr",
            "interest_coverage",
            "leverage_ratio",
            "dscr_breach_flag",
            "interest_coverage_breach_flag",
            "leverage_breach_flag",
        ]
        annual = monthly[flow_cols].resample("Y").sum(numeric_only=True).rename_axis("year_end")
        annual = annual.join(monthly[stock_cols].resample("Y").last())

        opex_allocation_views = self._annual_opex_allocation_views(
            idx=idx,
            units_wide=units_wide,
            revenue_wide=revenue_wide,
            opex_series=opex,
        )

        supporting_schedules = {
            "receivables_detail": receivables_details,
            "inventory_detail": inventory_details,
            "payables_detail": payables_details,
            "direct_labor_detail": direct_labor_summary,
            "indirect_labor_detail": indirect_labor_summary,
            "operations_summary": operations_plan.monthly_summary,
            "operations_by_sku": operations_plan.sku_schedule,
            "operations_resources": operations_plan.resource_schedule,
        }

        return ModelRunResult(
            monthly=monthly,
            annual=annual,
            prices=prices_wide,
            debt_schedules=debt_schedules,
            valuation=valuation,
            opex_allocation_views=opex_allocation_views,
            supporting_schedules=supporting_schedules,
        )


# =============================
# Convenience helper
# =============================
def phase_growth_series(
    idx: pd.DatetimeIndex,
    start_month: int,
    start_units: float,
    monthly_growth: float,
    stop_month: Optional[int] = None,
    cap_units: Optional[float] = None,
) -> pd.Series:
    """
    Simple helper to generate a monthly series that starts at start_month and grows by monthly_growth.
    - If stop_month is None, it grows until the end of the index.
    - cap_units (optional) applies a ceiling.
    """
    n = len(idx)
    s = np.zeros(n, dtype=float)
    stop = stop_month if stop_month is not None else n
    stop = int(np.clip(stop, 0, n))

    u = float(start_units)
    for t in range(int(start_month), stop):
        s[t] = u
        u = u * (1.0 + float(monthly_growth))
        if cap_units is not None:
            u = min(u, float(cap_units))
    return pd.Series(s, index=idx)


def write_comprehensive_excel_report(result: ModelRunResult, writer: pd.ExcelWriter) -> None:
    """Write a comprehensive, presentation-ready Excel workbook for model outputs."""
    # 1) Core statements
    result.monthly.to_excel(writer, sheet_name="01_Monthly_Financials")
    result.annual.to_excel(writer, sheet_name="02_Annual_Financials")
    result.prices.to_excel(writer, sheet_name="03_Pricing_Matrix")

    # 2) Key results dashboard
    valuation_df = pd.DataFrame(result.valuation, index=["value"]).T
    valuation_df.index.name = "metric"
    valuation_df.to_excel(writer, sheet_name="00_Key_Results")
    annual_key_cols = [
        c
        for c in [
            "gross_revenue",
            "net_revenue",
            "total_revenue",
            "direct_costs",
            "gross_profit",
            "opex",
            "ebitda",
            "net_income",
            "cash",
            "total_debt_ending_balance",
            "fcff",
        ]
        if c in result.annual.columns
    ]
    result.annual[annual_key_cols].to_excel(writer, sheet_name="00_Key_Results", startrow=valuation_df.shape[0] + 3)

    # 2b) Annual statements (requested views)
    perf_cols = [
        "gross_revenue",
        "trade_spend",
        "returns_allowances",
        "net_revenue",
        "total_revenue",
        "other_income",
        "direct_material_costs",
        "direct_labor_cost",
        "inventory_writeoff",
        "payable_discount_benefit",
        "direct_costs",
        "gross_profit",
        "indirect_labor_cost",
        "bad_debt_expense",
        "opex",
        "ebitda",
        "depreciation",
        "ebit",
        "interest_expense",
        "pre_tax_income",
        "taxes",
        "net_income",
    ]
    perf_existing = [c for c in perf_cols if c in result.annual.columns]
    result.annual[perf_existing].to_excel(writer, sheet_name="Annual_Performance_IS")

    position_cols = [
        "cash",
        "receivables",
        "inventory",
        "inventory_reserve",
        "other_current_assets",
        "current_assets",
        "net_fixed_assets",
        "total_assets",
        "payables",
        "other_current_liabilities",
        "current_liabilities",
        "debt_ending_balance",
        "revolver_ending_balance",
        "total_debt_ending_balance",
        "total_liabilities",
        "contributed_capital",
        "retained_earnings",
        "equity",
        "balance_sheet_gap",
    ]
    position_existing = [c for c in position_cols if c in result.annual.columns]
    result.annual[position_existing].to_excel(writer, sheet_name="Annual_Position_BS")

    cashflow_cols = [
        "cash_flow_from_operations",
        "change_in_nwc",
        "capex",
        "cash_flow_from_investing",
        "debt_draw",
        "debt_principal_payment",
        "revolver_draw",
        "revolver_principal_payment",
        "equity_injection",
        "dividends",
        "cash_flow_from_financing",
        "net_change_in_cash",
        "debt_service",
        "fcff",
    ]
    cashflow_existing = [c for c in cashflow_cols if c in result.annual.columns]
    result.annual[cashflow_existing].to_excel(writer, sheet_name="Annual_Cash_Flow")

    # 3) Schedules
    wc_cols = [
        c
        for c in [
            "receivables",
            "inventory",
            "inventory_reserve",
            "other_current_assets",
            "payables",
            "other_current_liabilities",
            "net_working_capital",
            "change_in_nwc",
        ]
        if c in result.monthly.columns
    ]
    result.monthly[wc_cols].to_excel(writer, sheet_name="04_Working_Capital")

    capex_cols = [c for c in ["capex", "depreciation", "net_fixed_assets"] if c in result.monthly.columns]
    result.monthly[capex_cols].to_excel(writer, sheet_name="05_CAPEX_Depreciation")

    financing_cols = [
        c
        for c in [
            "debt_draw",
            "debt_principal_payment",
            "revolver_draw",
            "revolver_principal_payment",
            "term_debt_interest_expense",
            "revolver_interest_expense",
            "interest_expense",
            "debt_service",
            "equity_injection",
            "dividends",
            "cash",
            "debt_ending_balance",
            "revolver_ending_balance",
            "total_debt_ending_balance",
            "dscr",
            "interest_coverage",
            "leverage_ratio",
        ]
        if c in result.monthly.columns
    ]
    result.monthly[financing_cols].to_excel(writer, sheet_name="06_Financing_Cash")

    # 4) Debt facility schedules
    for name, df in result.debt_schedules.items():
        df.to_excel(writer, sheet_name=f"07_Debt_{name[:22]}")

    # 5) Allocation views
    views = result.opex_allocation_views or {}
    if "pool_view" in views:
        views["pool_view"].to_excel(writer, sheet_name="08_OPEX_By_Pool")
    if "driver_view" in views:
        views["driver_view"].to_excel(writer, sheet_name="09_OPEX_By_Driver")
    if "product_view" in views:
        views["product_view"].to_excel(writer, sheet_name="10_OPEX_By_Product", index=False)
    if "reconciliation_view" in views:
        views["reconciliation_view"].to_excel(writer, sheet_name="11_OPEX_Reconciliation")

    # Consolidated driver-based OPEX views on a single sheet for quick review.
    opex_sheet = "Driver_OPEX_Views"
    row_ptr = 0
    if "pool_view" in views:
        views["pool_view"].to_excel(writer, sheet_name=opex_sheet, startrow=row_ptr)
        row_ptr += len(views["pool_view"]) + 4
    if "driver_view" in views:
        views["driver_view"].to_excel(writer, sheet_name=opex_sheet, startrow=row_ptr)
        row_ptr += len(views["driver_view"]) + 4
    if "product_view" in views:
        views["product_view"].to_excel(writer, sheet_name=opex_sheet, startrow=row_ptr, index=False)
        row_ptr += len(views["product_view"]) + 4
    if "reconciliation_view" in views:
        views["reconciliation_view"].to_excel(writer, sheet_name=opex_sheet, startrow=row_ptr)

    for sheet_name, df in (result.supporting_schedules or {}).items():
        safe_sheet = f"13_{sheet_name[:28]}"
        df.to_excel(writer, sheet_name=safe_sheet)

    # 6) Charts data + embedded charts
    chart_cols = [
        c
        for c in ["total_revenue", "ebitda", "net_income", "cash", "total_debt_ending_balance", "fcff"]
        if c in result.annual.columns
    ]
    chart_df = result.annual[chart_cols].copy()
    chart_df.to_excel(writer, sheet_name="12_Charts_Data")
    chart_df.to_excel(writer, sheet_name="Graphs_and_Plots")
    cash_debt_cols = [c for c in ["cash", "total_debt_ending_balance"] if c in result.annual.columns]
    result.annual[cash_debt_cols].to_excel(writer, sheet_name="Cash_vs_Debt_EndBal")

    # 6b) Key analytics sheet
    latest_month = result.monthly.iloc[-1] if not result.monthly.empty else pd.Series(dtype=float)
    latest_annual = result.annual.iloc[-1] if not result.annual.empty else pd.Series(dtype=float)
    key_analytics = pd.DataFrame(
        [
            {"metric": "latest_annual_revenue", "value": float(latest_annual.get("total_revenue", np.nan))},
            {"metric": "latest_annual_ebitda", "value": float(latest_annual.get("ebitda", np.nan))},
            {"metric": "latest_annual_net_income", "value": float(latest_annual.get("net_income", np.nan))},
            {
                "metric": "latest_gross_margin_pct",
                "value": float(latest_annual.get("gross_profit", np.nan)) / max(float(latest_annual.get("total_revenue", np.nan)), 1e-9),
            },
            {
                "metric": "latest_ebitda_margin_pct",
                "value": float(latest_annual.get("ebitda", np.nan)) / max(float(latest_annual.get("total_revenue", np.nan)), 1e-9),
            },
            {
                "metric": "cash_to_debt_last_month",
                "value": float(latest_month.get("cash", np.nan)) / max(float(latest_month.get("total_debt_ending_balance", np.nan)), 1e-9),
            },
            {"metric": "latest_fcff", "value": float(latest_annual.get("fcff", np.nan))},
        ]
    )
    key_analytics.to_excel(writer, sheet_name="Key_Analytics", index=False)

    try:
        from openpyxl.chart import BarChart, LineChart, Reference
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        wb = writer.book
        accent = "1F4E78"
        dark = "0E2A47"
        light = "EAF0F6"
        border_color = "C5CFDA"
        thin = Side(border_style="thin", color=border_color)

        def _style_sheet(ws, title: str) -> None:
            max_col = max(ws.max_column, 1)
            ws.insert_rows(1)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
            banner = ws.cell(row=1, column=1, value=title)
            banner.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
            banner.fill = PatternFill(fill_type="solid", fgColor=dark)
            banner.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[1].height = 24

            for cell in ws[2]:
                cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                cell.fill = PatternFill(fill_type="solid", fgColor=accent)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            ws.row_dimensions[2].height = 22

            for r in range(3, ws.max_row + 1):
                fill = PatternFill(fill_type="solid", fgColor=light if r % 2 == 0 else "FFFFFF")
                for c in range(1, ws.max_column + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.fill = fill
                    cell.font = Font(name="Calibri", size=10, color="1C1C1C")
                    cell.alignment = Alignment(horizontal="right" if c > 1 else "left", vertical="center")
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for col_idx in range(1, ws.max_column + 1):
                letter = get_column_letter(col_idx)
                sample_values = [str(ws.cell(row=r, column=col_idx).value or "") for r in range(1, min(ws.max_row, 80) + 1)]
                width = min(max(len(max(sample_values, key=len)) + 2, 12), 40)
                ws.column_dimensions[letter].width = width

            ws.freeze_panes = "A3"
            ws.sheet_view.showGridLines = False
            ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"

        chart_next_row: dict[str, int] = {}

        def _add_chart_below(ws, title: str, data_cols: list[str], chart_type: str = "line") -> None:
            headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
            header_map = {str(h): idx + 1 for idx, h in enumerate(headers) if h is not None}
            cols = [header_map[c] for c in data_cols if c in header_map]
            if len(cols) < 1 or ws.max_row < 4:
                return
            cat_col = 1
            chart = LineChart() if chart_type == "line" else BarChart()
            chart.title = title
            chart.style = 10
            chart.height = 7
            chart.width = 14
            chart.y_axis.title = "Value"
            chart.x_axis.title = "Period"
            data_ref = Reference(ws, min_col=min(cols), max_col=max(cols), min_row=2, max_row=ws.max_row)
            cat_ref = Reference(ws, min_col=cat_col, min_row=3, max_row=ws.max_row)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cat_ref)
            base_anchor = ws.max_row + 3
            chart_anchor_row = max(base_anchor, chart_next_row.get(ws.title, base_anchor))
            ws.cell(row=chart_anchor_row - 1, column=1, value=f"{title} — Visual")
            ws.cell(row=chart_anchor_row - 1, column=1).font = Font(name="Calibri", size=11, bold=True, color=dark)
            ws.add_chart(chart, f"A{chart_anchor_row}")
            chart_next_row[ws.title] = chart_anchor_row + 16

        # Style workbook tabs with clear hierarchy.
        sheet_titles = {
            "00_Key_Results": "Executive Summary & Key Results",
            "01_Monthly_Financials": "Monthly Financial Statements",
            "02_Annual_Financials": "Annual Financial Statements",
            "03_Pricing_Matrix": "Pricing Matrix",
            "04_Working_Capital": "Working Capital Schedule",
            "05_CAPEX_Depreciation": "CAPEX & Depreciation Schedule",
            "06_Financing_Cash": "Financing & Cash Schedule",
            "08_OPEX_By_Pool": "OPEX Allocation by Cost Pool",
            "09_OPEX_By_Driver": "OPEX Allocation by Driver",
            "10_OPEX_By_Product": "OPEX Allocation by Product",
            "11_OPEX_Reconciliation": "OPEX Reconciliation",
            "12_Charts_Data": "Charts Data",
            "Annual_Performance_IS": "Annual Performance (Income Statement)",
            "Annual_Position_BS": "Annual Position (Balance Sheet)",
            "Annual_Cash_Flow": "Annual Cash Flow",
            "Driver_OPEX_Views": "Driver-based OPEX Review",
            "Graphs_and_Plots": "Graphs and Plots",
            "Cash_vs_Debt_EndBal": "Cash vs Debt Ending Balance",
            "Key_Analytics": "Key Analytics Dashboard",
        }
        for sheet_name in wb.sheetnames:
            if sheet_name.startswith("13_") and sheet_name not in sheet_titles:
                sheet_titles[sheet_name] = sheet_name.replace("13_", "").replace("_", " ").title()
        for name, title in sheet_titles.items():
            if name in wb.sheetnames:
                _style_sheet(wb[name], title)

        for debt_sheet in [s for s in wb.sheetnames if s.startswith("07_Debt_")]:
            _style_sheet(wb[debt_sheet], debt_sheet.replace("07_Debt_", "Debt Facility: "))

        # Place visuals directly below major output tables.
        _add_chart_below(wb["01_Monthly_Financials"], "Monthly Revenue / EBITDA / Net Income", ["total_revenue", "ebitda", "net_income"], "line")
        _add_chart_below(wb["02_Annual_Financials"], "Annual Revenue / EBITDA / Net Income", ["total_revenue", "ebitda", "net_income"], "line")
        _add_chart_below(wb["04_Working_Capital"], "Working Capital Bridge", ["net_working_capital", "change_in_nwc"], "line")
        _add_chart_below(wb["05_CAPEX_Depreciation"], "CAPEX vs Depreciation", ["capex", "depreciation"], "bar")
        _add_chart_below(wb["06_Financing_Cash"], "Cash vs Debt", ["cash", "total_debt_ending_balance"], "line")
        if "Annual_Performance_IS" in wb.sheetnames:
            _add_chart_below(wb["Annual_Performance_IS"], "Income Statement Trend", ["total_revenue", "gross_profit", "ebitda", "net_income"], "line")
        if "Annual_Position_BS" in wb.sheetnames:
            _add_chart_below(wb["Annual_Position_BS"], "Balance Sheet Composition", ["cash", "total_assets", "total_liabilities", "equity"], "bar")
        if "Annual_Cash_Flow" in wb.sheetnames:
            _add_chart_below(wb["Annual_Cash_Flow"], "Cash Flow Components", ["cash_flow_from_operations", "cash_flow_from_investing", "cash_flow_from_financing", "fcff"], "bar")
        if "08_OPEX_By_Pool" in wb.sheetnames:
            _add_chart_below(wb["08_OPEX_By_Pool"], "OPEX by Pool", ["allocated_opex"], "bar")
        if "09_OPEX_By_Driver" in wb.sheetnames:
            _add_chart_below(wb["09_OPEX_By_Driver"], "OPEX by Driver", ["allocated_opex"], "bar")
        if "11_OPEX_Reconciliation" in wb.sheetnames:
            _add_chart_below(wb["11_OPEX_Reconciliation"], "Reconciliation Gap by Year", ["reconciliation_gap"], "bar")
        if "Cash_vs_Debt_EndBal" in wb.sheetnames:
            _add_chart_below(wb["Cash_vs_Debt_EndBal"], "Cash vs Debt Ending Balance", ["cash", "total_debt_ending_balance"], "line")
        if "Key_Analytics" in wb.sheetnames:
            _add_chart_below(wb["Key_Analytics"], "Key Analytics Snapshot", ["value"], "bar")
        if "Graphs_and_Plots" in wb.sheetnames:
            _add_chart_below(wb["Graphs_and_Plots"], "Revenue / EBITDA / Net Income", ["total_revenue", "ebitda", "net_income"], "line")
            _add_chart_below(wb["Graphs_and_Plots"], "Cash / Debt / FCFF", ["cash", "total_debt_ending_balance", "fcff"], "line")

    except Exception:
        # If style/chart libs are unavailable, workbook remains complete with data.
        pass


# =============================
# Example runner
# =============================
def main() -> None:
    # ----------------------------
    # Model horizon (10 years monthly)
    # ----------------------------
    cfg = ModelConfig(
        start_date="2025-01-01",
        months=120,
        pricing_cost_basis_month=24,  # compute base cost-plus prices using month 25 as a cost basis
        price_inflation_annual=0.015,
        cost_inflation_annual=0.015,
        tax_rate=0.25,
        wacc_annual=0.122,
        exit_ev_ebitda_multiple=8.0,
        initial_cash=0.0,
    )

    div = DividendPolicy(
        enabled=True,
        model="cash_sweep",
        start_month=60,              # "Year 5" in a 0-indexed monthly model
        minimum_cash_position=1_500_000.0,
        payout_ratio=0.25,
    )

    idx = pd.date_range(cfg.start_date, periods=cfg.months, freq="MS")

    # ----------------------------
    # SKUs (replace with your own SKU list)
    # direct_cost_per_unit: USD/unit
    # markup_pct: e.g., 0.60 => 60% markup on cost-plus base
    # ----------------------------
    skus = pd.DataFrame(
        [
            {"sku_id": 1, "name": "Pale Ale 330ml", "direct_cost_per_unit": 0.0, "markup_pct": 0.65, "relative_opex_weight": 1.0},
            {"sku_id": 2, "name": "Pilsner 500ml", "direct_cost_per_unit": 0.0, "markup_pct": 0.60, "relative_opex_weight": 1.1},
        ]
    )

    # ----------------------------
    # Distribution channels and price factors
    # (factor multiplies the SKU base price)
    # ----------------------------
    channels = pd.DataFrame(
        [
            {"channel": "Wholesale", "price_factor": 1.40},
            {"channel": "Retail", "price_factor": 2.00},
            {"channel": "E-Commerce", "price_factor": 1.75},
            {"channel": "On-Premise", "price_factor": 1.00},
        ]
    )

    # ----------------------------
    # Sales plan (units per month, per SKU, per channel)
    # Replace this with your own monthly forecast.
    # ----------------------------
    u_sku1 = phase_growth_series(idx, start_month=3, start_units=8_000, monthly_growth=0.04, stop_month=None, cap_units=25_000)
    u_sku2 = phase_growth_series(idx, start_month=3, start_units=6_000, monthly_growth=0.04, stop_month=None, cap_units=20_000)

    channel_mix = {"Wholesale": 0.45, "Retail": 0.35, "E-Commerce": 0.15, "On-Premise": 0.05}

    rows = []
    for date in idx:
        for sku_id, series in [(1, u_sku1), (2, u_sku2)]:
            total_units = float(series.loc[date])
            for channel, share in channel_mix.items():
                rows.append({"date": date, "sku_id": sku_id, "channel": channel, "units": total_units * share})
    sales_plan = pd.DataFrame(rows)

    # ----------------------------
    # Cost pools and other income (monthly)
    # ----------------------------
    cost_pools = [
        CostPoolInput(name="Malt & Grain", cost_type="direct", behavior="variable", allocation_driver="liters", unit_variable_cost=0.22),
        CostPoolInput(name="Hops & Yeast", cost_type="direct", behavior="variable", allocation_driver="liters", unit_variable_cost=0.09),
        CostPoolInput(name="Packaging Materials", cost_type="direct", behavior="variable", allocation_driver="units", unit_variable_cost=0.14),
        CostPoolInput(name="Production Direct Labor", cost_type="direct", behavior="step_fixed", allocation_driver="liters", fixed_monthly_cost=6_000.0, step_threshold=180_000.0, step_increment=850.0),
        CostPoolInput(name="Brew QA Consumables", cost_type="direct", behavior="variable", allocation_driver="liters", unit_variable_cost=0.015),
        CostPoolInput(name="Indirect Labor", cost_type="indirect", behavior="step_fixed", allocation_driver="liters", fixed_monthly_cost=22_000.0, step_threshold=250_000.0, step_increment=2_000.0),
        CostPoolInput(name="Utilities", cost_type="indirect", behavior="variable", allocation_driver="liters", unit_variable_cost=0.035),
        CostPoolInput(name="Supplies", cost_type="indirect", behavior="variable", allocation_driver="units", unit_variable_cost=0.015),
        CostPoolInput(name="Marketing & Advertising", cost_type="indirect", behavior="blended", allocation_driver="channel_revenue", fixed_monthly_cost=8_500.0, unit_variable_cost=0.003, channel=None),
        CostPoolInput(name="Events & Promotion", cost_type="indirect", behavior="variable", allocation_driver="channel_units", unit_variable_cost=0.01, channel="On-Premise"),
        CostPoolInput(name="Insurance", cost_type="indirect", behavior="fixed", allocation_driver="revenue", fixed_monthly_cost=3_000.0),
        CostPoolInput(name="Permits & License", cost_type="indirect", behavior="blended", allocation_driver="active_sku", fixed_monthly_cost=1_250.0, unit_variable_cost=350.0),
        CostPoolInput(name="Local Fees", cost_type="indirect", behavior="fixed", allocation_driver="units", fixed_monthly_cost=900.0),
        CostPoolInput(name="Transport", cost_type="indirect", behavior="variable", allocation_driver="channel_units", unit_variable_cost=0.018),
        CostPoolInput(name="Administrative Expense", cost_type="indirect", behavior="blended", allocation_driver="active_sku", fixed_monthly_cost=9_500.0, unit_variable_cost=250.0),
        CostPoolInput(name="Quality Control", cost_type="indirect", behavior="blended", allocation_driver="liters", fixed_monthly_cost=2_500.0, unit_variable_cost=0.01),
        CostPoolInput(name="Certificates", cost_type="indirect", behavior="variable", allocation_driver="active_sku", unit_variable_cost=180.0),
        CostPoolInput(name="Professional Services", cost_type="indirect", behavior="fixed", allocation_driver="units", fixed_monthly_cost=2_200.0),
        CostPoolInput(name="Other Expense", cost_type="indirect", behavior="fixed", allocation_driver="units", fixed_monthly_cost=1_400.0),
        CostPoolInput(name="Contingencies", cost_type="indirect", behavior="fixed", allocation_driver="units", fixed_monthly_cost=1_800.0),
    ]

    other_income_items = [
        OtherIncomeItem(other_income_name="Sponsorships", amount=15_000.0, active=True, category="Commercial"),
        OtherIncomeItem(other_income_name="Other Income 2", amount=0.0, active=False),
        OtherIncomeItem(other_income_name="Other Income 3", amount=0.0, active=False),
        OtherIncomeItem(other_income_name="Other Income 4", amount=0.0, active=False),
    ]

    # ----------------------------
    # CAPEX schedule (simplified example)
    # ----------------------------
    capex_items = [
        CapexItem(name="Land (non-depreciable)", amount=875_000, capex_month=0, depreciation_years=0),
        CapexItem(name="Building", amount=1_750_000, capex_month=0, depreciation_years=25),
        CapexItem(name="Brewhouse equipment", amount=1_250_000, capex_month=1, depreciation_years=10),
        CapexItem(name="Expansion equipment", amount=900_000, capex_month=60, depreciation_years=10),
    ]

    # ----------------------------
    # Debt facilities (example terms)
    # ----------------------------
    debt_facilities = [
        DebtFacility(name="Mortgage", principal=750_000, annual_interest_rate=0.03, draw_month=0, grace_months=6, term_months=120, repayment_type="linear"),
        DebtFacility(name="Loan A", principal=450_000, annual_interest_rate=0.025, draw_month=5, grace_months=4, term_months=60, repayment_type="annuity"),
        DebtFacility(name="Loan B", principal=200_000, annual_interest_rate=0.015, draw_month=58, grace_months=0, term_months=36, repayment_type="linear"),
    ]

    equity_injections = {0: 5_500_000.0, 12: 1_000_000.0}

    inputs = ModelInputs(
        skus=skus,
        channels=channels,
        sales_plan=sales_plan,
        other_income_items=other_income_items,
        cost_pools=cost_pools,
        capex_items=capex_items,
        debt_facilities=debt_facilities,
        equity_injections=equity_injections,
    )

    model = MicrobreweryFinancialModel(cfg, div, inputs)
    result = model.run()

    print("\n=== Valuation summary ===")
    for k, v in result.valuation.items():
        if isinstance(v, float):
            print(f"{k:>28}: {v:,.6f}")
        else:
            print(f"{k:>28}: {v}")

    print("\n=== Annual Income Statement (selected lines) ===")
    annual = result.annual[["total_revenue", "direct_costs", "gross_profit", "opex", "ebitda", "net_income"]].copy()
    print(annual.tail(5).to_string())

    # Export outputs (optional)
    out_xlsx = "brewery_model_output.xlsx"
    try:
        with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
            write_comprehensive_excel_report(result, writer)
        print(f"\nWrote: {out_xlsx}")
    except Exception as e:
        print(f"\nExcel export skipped (openpyxl missing or other error): {e}")


if __name__ == "__main__":
    main()
