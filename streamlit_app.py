"""Streamlit UI for the microbrewery financial model.

This app wraps the existing `MicrobreweryFinancialModel` with the sample
assumptions from ``brewery_financial_model_all_in_one.py`` and lets users tweak
key valuation parameters before running the projection. The UI surfaces the
valuation summary and selected statement tables, plus an Excel download of the
full outputs.
"""

from __future__ import annotations

from copy import deepcopy
from dataclasses import dataclass, fields, is_dataclass, replace
import hashlib
import io
import json
import time
import urllib.parse
import urllib.request
from typing import Any, Tuple

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

st.set_page_config(layout="wide", initial_sidebar_state="collapsed")

from finmodel.decision_ai import DecisionSessionMemory, QuestionType, build_structured_answer, classify_question
from brewery_financial_model_all_in_one import (
    CapexItem,
    CostPoolInput,
    DebtFacility,
    DividendPolicy,
    ModelRunResult,
    MicrobreweryFinancialModel,
    ModelConfig,
    ModelInputs,
    OtherIncomeItem,
    phase_growth_series,
    propagate_sales_plan_monthly,
    write_comprehensive_excel_report,
)


_CURRENT_RESULT_KEY = "microbrewery_current_result_bundle"
_DRAFT_SIGNATURE_VERSION = "microbrewery-draft-v1"


@dataclass
class StoredModelRun:
    signature: str
    cfg: ModelConfig
    div: DividendPolicy
    inputs: ModelInputs
    result: ModelRunResult
    computed_at: float
    duration_seconds: float


def _inject_app_theme() -> None:
    st.markdown(
        """
        <style>
        :root {
            --shell-ink: #0f172a;
            --shell-muted: #475569;
            --shell-line: rgba(15, 23, 42, 0.10);
            --shell-panel: rgba(255, 255, 255, 0.86);
            --shell-brand: #8f3d22;
            --shell-brand-soft: #fff0e8;
            --shell-accent: #174c43;
        }
        .stApp {
            background:
                radial-gradient(circle at top left, rgba(248, 192, 138, 0.28), transparent 34%),
                radial-gradient(circle at top right, rgba(102, 180, 155, 0.18), transparent 30%),
                linear-gradient(180deg, #f7efe8 0%, #f3f5f9 52%, #eef3f0 100%);
        }
        .block-container {
            padding-top: 1.35rem;
            padding-bottom: 3rem;
            max-width: 1380px;
        }
        .designer-hero {
            position: relative;
            overflow: hidden;
            margin: 0 0 1.2rem 0;
            padding: 1.7rem 1.8rem;
            border: 1px solid rgba(143, 61, 34, 0.14);
            border-radius: 28px;
            background:
                linear-gradient(135deg, rgba(255, 240, 232, 0.95), rgba(255, 255, 255, 0.94)),
                linear-gradient(135deg, rgba(143, 61, 34, 0.06), rgba(23, 76, 67, 0.08));
            box-shadow: 0 24px 48px rgba(15, 23, 42, 0.08);
        }
        .designer-hero::after {
            content: "";
            position: absolute;
            inset: auto -4rem -4rem auto;
            width: 15rem;
            height: 15rem;
            border-radius: 999px;
            background: radial-gradient(circle, rgba(143, 61, 34, 0.14), transparent 72%);
        }
        .designer-kicker {
            margin: 0 0 0.45rem 0;
            font-size: 0.78rem;
            letter-spacing: 0.16em;
            text-transform: uppercase;
            color: var(--shell-brand);
            font-weight: 700;
        }
        .designer-title {
            margin: 0;
            font-size: clamp(2rem, 2.8vw, 3.15rem);
            line-height: 1.02;
            color: var(--shell-ink);
            font-weight: 800;
        }
        .designer-copy {
            max-width: 52rem;
            margin: 0.65rem 0 0 0;
            color: var(--shell-muted);
            font-size: 1rem;
            line-height: 1.6;
        }
        .designer-badges {
            display: flex;
            flex-wrap: wrap;
            gap: 0.55rem;
            margin-top: 1rem;
        }
        .designer-badge {
            padding: 0.42rem 0.78rem;
            border-radius: 999px;
            border: 1px solid rgba(15, 23, 42, 0.08);
            background: rgba(255, 255, 255, 0.92);
            color: var(--shell-accent);
            font-size: 0.82rem;
            font-weight: 700;
        }
        div[data-baseweb="tab-list"] {
            gap: 0.55rem;
            margin-bottom: 1rem;
        }
        div[data-baseweb="tab-list"] button {
            min-height: 3rem;
            border-radius: 999px;
            border: 1px solid rgba(15, 23, 42, 0.08);
            background: rgba(255, 255, 255, 0.7);
            color: var(--shell-muted);
            padding: 0.25rem 1rem;
        }
        div[data-baseweb="tab-list"] button[aria-selected="true"] {
            background: linear-gradient(135deg, #8f3d22, #174c43);
            color: white;
            border-color: transparent;
            box-shadow: 0 12px 24px rgba(23, 76, 67, 0.18);
        }
        div[data-testid="stMetric"],
        div[data-testid="stDataFrame"],
        div[data-testid="stExpander"] {
            border-radius: 20px;
        }
        div[data-testid="stMetric"] {
            border: 1px solid var(--shell-line);
            background: var(--shell-panel);
            padding: 0.6rem 0.7rem;
            box-shadow: 0 12px 28px rgba(15, 23, 42, 0.05);
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def _render_model_hero() -> None:
    badges = "".join(
        f'<span class="designer-badge">{label}</span>'
        for label in (
            "Investor-ready workbook",
            "Operating model",
            "Scenario-led valuation",
            "Driver-based OPEX",
        )
    )
    st.markdown(
        f"""
        <section class="designer-hero">
            <p class="designer-kicker">Craft brewery planning suite</p>
            <h1 class="designer-title">Microbrewery Financial Model</h1>
            <p class="designer-copy">
                Review production, channels, pricing, debt, and valuation in one polished workspace,
                then export a lender-friendly workbook instead of a raw data dump.
            </p>
            <div class="designer-badges">{badges}</div>
        </section>
        """,
        unsafe_allow_html=True,
    )


def _format_dashboard_value(value: object) -> object:
    if isinstance(value, (int, float, np.floating)) and np.isfinite(value):
        abs_value = abs(float(value))
        if abs_value >= 1_000_000:
            return f"${value / 1_000_000:,.2f}M"
        if abs_value >= 1_000:
            return f"${value / 1_000:,.1f}K"
        return f"${value:,.0f}"
    return value


def _style_workbook_sheet(ws, *, accent: str, accent_soft: str, is_overview: bool = False) -> None:
    ws.sheet_view.showGridLines = False
    if is_overview:
        ws.freeze_panes = "A6"
    elif ws.max_row > 1:
        ws.freeze_panes = "A2"
        header_fill = PatternFill("solid", fgColor=accent)
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.auto_filter.ref = ws.dimensions
        for row in range(2, min(ws.max_row, 120) + 1):
            if row % 2 == 0:
                for cell in ws[row]:
                    cell.fill = PatternFill("solid", fgColor=accent_soft)
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        for row_idx in range(1, min(ws.max_row, 80) + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_length + 2, 14), 34)


def _extract_workbook_summary(result) -> list[tuple[str, object]]:
    summary: list[tuple[str, object]] = []
    annual = result.annual.copy() if isinstance(result.annual, pd.DataFrame) else pd.DataFrame()
    if not annual.empty:
        latest = annual.iloc[-1]
        for label, key in (
            ("Latest Revenue", "total_revenue"),
            ("Latest EBITDA", "ebitda"),
            ("Latest Net Income", "net_income"),
            ("Latest FCFF", "fcff"),
        ):
            if key in latest.index and pd.notna(latest[key]):
                summary.append((label, _format_dashboard_value(float(latest[key]))))
    for label, key in (
        ("NPV", "npv"),
        ("IRR", "irr"),
        ("Terminal Value", "terminal_value"),
    ):
        value = result.valuation.get(key)
        if value is None:
            continue
        if "irr" in key:
            summary.append((label, f"{float(value):.1%}"))
        else:
            summary.append((label, _format_dashboard_value(float(value))))
    summary.append(("Sheets Included", len(result.debt_schedules) + 4))
    return summary[:8]


def _build_professional_excel_output(result) -> bytes:
    raw_buffer = io.BytesIO()
    with pd.ExcelWriter(raw_buffer, engine="openpyxl") as writer:
        write_comprehensive_excel_report(result, writer)
    raw_buffer.seek(0)

    workbook = load_workbook(raw_buffer)
    accent = "8F3D22"
    accent_soft = "FCE9DF"
    ink = "0F172A"

    if "Overview" in workbook.sheetnames:
        del workbook["Overview"]
    overview = workbook.create_sheet("Overview", 0)
    overview["A1"] = "Microbrewery Financial Model"
    overview["A1"].font = Font(size=20, bold=True, color=ink)
    overview["A2"] = "Executive workbook with valuation snapshot, detailed statements, and operating schedules."
    overview["A2"].font = Font(size=11, color="475569")
    overview["A4"] = "Executive Snapshot"
    overview["A4"].font = Font(size=12, bold=True, color=accent)
    overview["A5"] = "Metric"
    overview["B5"] = "Value"
    for cell in overview[5]:
        cell.fill = PatternFill("solid", fgColor=accent)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_idx, (label, value) in enumerate(_extract_workbook_summary(result), start=6):
        overview.cell(row=row_idx, column=1, value=label)
        overview.cell(row=row_idx, column=2, value=value)
    overview["D4"] = "Design Notes"
    overview["D4"].font = Font(size=12, bold=True, color=accent)
    notes = [
        "Use the assumptions tab to change timing, pricing, debt, or payout policy.",
        "Results sheets provide monthly and annual statements for investor review.",
        "The workbook is formatted for management, lender, and board distribution.",
    ]
    for row_idx, note in enumerate(notes, start=5):
        overview.cell(row=row_idx, column=4, value=f"• {note}")
    overview.column_dimensions["A"].width = 26
    overview.column_dimensions["B"].width = 18
    overview.column_dimensions["D"].width = 58

    for sheet in workbook.worksheets:
        _style_workbook_sheet(
            sheet,
            accent=accent,
            accent_soft=accent_soft,
            is_overview=sheet.title == "Overview",
        )

    final_buffer = io.BytesIO()
    workbook.save(final_buffer)
    return final_buffer.getvalue()

def _normalize_for_signature(value: Any) -> Any:
    if is_dataclass(value):
        return {
            field.name: _normalize_for_signature(getattr(value, field.name))
            for field in fields(value)
        }
    if isinstance(value, pd.DataFrame):
        return {
            "columns": [_normalize_for_signature(col) for col in list(value.columns)],
            "rows": [
                [_normalize_for_signature(cell) for cell in row]
                for row in value.itertuples(index=False, name=None)
            ],
        }
    if isinstance(value, pd.Series):
        return {
            "name": _normalize_for_signature(value.name),
            "index": [_normalize_for_signature(item) for item in list(value.index)],
            "values": [_normalize_for_signature(item) for item in value.tolist()],
        }
    if isinstance(value, pd.Index):
        return [_normalize_for_signature(item) for item in list(value)]
    if isinstance(value, dict):
        normalized_items = [
            {
                "key": _normalize_for_signature(key),
                "value": _normalize_for_signature(val),
            }
            for key, val in value.items()
        ]
        return sorted(
            normalized_items,
            key=lambda item: json.dumps(item["key"], sort_keys=True),
        )
    if isinstance(value, (list, tuple)):
        return [_normalize_for_signature(item) for item in value]
    if isinstance(value, (pd.Timestamp, np.datetime64)):
        return pd.Timestamp(value).isoformat()
    if isinstance(value, np.generic):
        return _normalize_for_signature(value.item())
    if value is None or value is pd.NaT:
        return None
    if isinstance(value, float):
        return None if not np.isfinite(value) else value
    if isinstance(value, (bool, int, str)):
        return value
    if pd.isna(value):
        return None
    return str(value)


def _signature_for_payload(version: str, payload: object) -> str:
    encoded = json.dumps(
        {
            "version": version,
            "payload": _normalize_for_signature(payload),
        },
        sort_keys=True,
        separators=(",", ":"),
    )
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()


def _build_draft_signature(cfg: ModelConfig, div: DividendPolicy, inputs: ModelInputs) -> str:
    return _signature_for_payload(
        _DRAFT_SIGNATURE_VERSION,
        {"cfg": cfg, "div": div, "inputs": inputs},
    )


def _current_result_bundle() -> StoredModelRun | None:
    bundle = st.session_state.get(_CURRENT_RESULT_KEY)
    return bundle if isinstance(bundle, StoredModelRun) else None


def _activate_result_bundle(bundle: StoredModelRun) -> None:
    st.session_state[_CURRENT_RESULT_KEY] = bundle


def _draft_is_dirty(draft_signature: str, bundle: StoredModelRun | None) -> bool:
    return bundle is not None and bundle.signature != draft_signature


def _render_stale_banner(message: str) -> None:
    st.warning(message)


def _driver_based_opex_views_section(result: ModelRunResult) -> None:
    """
    Display the three OPEX output views from the driver-based allocation engine:
      A) OPEX by pool
      B) OPEX by driver type
      C) OPEX by product
    """
    st.subheader("Driver-based OPEX Allocation Views")
    st.caption(
        "This section renders the new pool-driven OPEX engine outputs (management, model-control, and pricing views)."
    )

    views = result.opex_allocation_views
    pool_view = views.get("pool_view", pd.DataFrame())
    driver_view = views.get("driver_view", pd.DataFrame())
    product_df = views.get("product_view", pd.DataFrame())
    rec_df = views.get("reconciliation_view", pd.DataFrame())

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("#### A. OPEX by pool")
        st.dataframe(pool_view, use_container_width=True)
    with c2:
        st.markdown("#### B. OPEX by driver type")
        st.dataframe(driver_view, use_container_width=True)

    st.markdown("#### C. OPEX by product")
    if not product_df.empty:
        product_df = product_df.sort_values(["sku_id", "year"])
    st.dataframe(product_df, use_container_width=True)

    st.markdown("#### Reconciliation check")
    if not rec_df.empty:
        rec_df["within_tolerance"] = rec_df["reconciliation_gap"].abs() <= 1e-6
    st.dataframe(rec_df, use_container_width=True)


def _build_sample_assumptions(
    cfg: ModelConfig,
    div: DividendPolicy,
) -> ModelInputs:
    idx = pd.date_range(cfg.start_date, periods=cfg.months, freq="MS")
    year_cols = _year_columns_for_months(cfg.months)

    skus = pd.DataFrame(
        [
            {
                "sku_id": 1,
                "name": "Pale Ale 330ml",
                "direct_cost_per_unit": 0.0,
                "markup_pct": 0.65,
                "relative_opex_weight": 1.0,
            },
            {
                "sku_id": 2,
                "name": "Pilsner 500ml",
                "direct_cost_per_unit": 0.0,
                "markup_pct": 0.60,
                "relative_opex_weight": 1.1,
            },
            {
                "sku_id": 3,
                "name": "Hazy IPA 440ml",
                "direct_cost_per_unit": 0.0,
                "markup_pct": 0.72,
                "relative_opex_weight": 1.25,
            },
        ]
    )

    channels = pd.DataFrame(
        [
            {"channel": "Wholesale", "price_factor": 1.40},
            {"channel": "Retail", "price_factor": 2.00},
            {"channel": "E-Commerce", "price_factor": 1.75},
            {"channel": "On-Premise", "price_factor": 1.00},
            {"channel": "Export", "price_factor": 1.55},
        ]
    )

    u_sku1 = phase_growth_series(
        idx, start_month=3, start_units=8_000, monthly_growth=0.04, stop_month=None, cap_units=25_000
    )
    u_sku2 = phase_growth_series(
        idx, start_month=3, start_units=6_000, monthly_growth=0.04, stop_month=None, cap_units=20_000
    )
    u_sku3 = phase_growth_series(
        idx, start_month=6, start_units=3_000, monthly_growth=0.05, stop_month=None, cap_units=18_000
    )

    channel_mix = {"Wholesale": 0.40, "Retail": 0.30, "E-Commerce": 0.15, "On-Premise": 0.10, "Export": 0.05}
    rows = []
    for date in idx:
        for sku_id, series in [(1, u_sku1), (2, u_sku2), (3, u_sku3)]:
            total_units = float(series.loc[date])
            for channel, share in channel_mix.items():
                rows.append({"date": date, "sku_id": sku_id, "channel": channel, "units": total_units * share})
    sales_plan = pd.DataFrame(rows)

    cost_pools = [
        CostPoolInput(name="Malt & Grain", cost_type="direct", behavior="variable", allocation_driver="liters", unit_variable_cost=0.22),
        CostPoolInput(name="Hops & Yeast", cost_type="direct", behavior="variable", allocation_driver="liters", unit_variable_cost=0.09),
        CostPoolInput(name="Packaging Materials", cost_type="direct", behavior="variable", allocation_driver="units", unit_variable_cost=0.14),
        CostPoolInput(name="Brew QA Consumables", cost_type="direct", behavior="variable", allocation_driver="liters", unit_variable_cost=0.015),
        CostPoolInput(name="Utilities", cost_type="indirect", behavior="variable", allocation_driver="liters", unit_variable_cost=0.035),
        CostPoolInput(name="Supplies", cost_type="indirect", behavior="variable", allocation_driver="units", unit_variable_cost=0.015),
        CostPoolInput(name="Marketing & Advertising", cost_type="indirect", behavior="blended", allocation_driver="channel_revenue", fixed_monthly_cost=8_500.0, unit_variable_cost=0.003),
        CostPoolInput(name="Events & Promotion", cost_type="indirect", behavior="variable", allocation_driver="channel_units", unit_variable_cost=0.01, channel="On-Premise"),
        CostPoolInput(name="Insurance", cost_type="indirect", behavior="fixed", allocation_driver="revenue", fixed_monthly_cost=3_000.0),
        CostPoolInput(name="Permits & License", cost_type="indirect", behavior="blended", allocation_driver="active_sku", fixed_monthly_cost=1_250.0, unit_variable_cost=350.0),
        CostPoolInput(name="Local Fees", cost_type="indirect", behavior="fixed", allocation_driver="units", fixed_monthly_cost=900.0),
        CostPoolInput(name="Transport", cost_type="indirect", behavior="variable", allocation_driver="channel_units", unit_variable_cost=0.018),
        CostPoolInput(name="Administrative Expense", cost_type="indirect", behavior="blended", allocation_driver="active_sku", fixed_monthly_cost=9_500.0, unit_variable_cost=250.0),
        CostPoolInput(name="Quality Control", cost_type="indirect", behavior="blended", allocation_driver="liters", fixed_monthly_cost=2_500.0, unit_variable_cost=0.01),
        CostPoolInput(name="Certificates", cost_type="indirect", behavior="variable", allocation_driver="active_sku", unit_variable_cost=180.0),
        CostPoolInput(name="Professional Services", cost_type="indirect", behavior="fixed", allocation_driver="units", fixed_monthly_cost=2_200.0),
        CostPoolInput(name="Other Expense", cost_type="indirect", behavior="fixed", allocation_driver="units", fixed_monthly_cost=1_400.0),
        CostPoolInput(name="Contingencies", cost_type="indirect", behavior="fixed", allocation_driver="units", fixed_monthly_cost=1_800.0),
    ]
    other_income_items = [
        OtherIncomeItem(other_income_name="Sponsorships", amount=15_000.0, active=True, category="Commercial"),
        OtherIncomeItem(other_income_name="Other Income 2", amount=0.0, active=False),
        OtherIncomeItem(other_income_name="Other Income 3", amount=0.0, active=False),
        OtherIncomeItem(other_income_name="Other Income 4", amount=0.0, active=False),
    ]

    capex_items = [
        CapexItem(name="Land (non-depreciable)", amount=875_000, capex_month=0, depreciation_years=0),
        CapexItem(name="Building", amount=1_750_000, capex_month=0, depreciation_years=25),
        CapexItem(name="Brewhouse equipment", amount=1_250_000, capex_month=1, depreciation_years=10),
        CapexItem(name="Expansion equipment", amount=900_000, capex_month=60, depreciation_years=10),
        CapexItem(name="Packaging line upgrade", amount=650_000, capex_month=36, depreciation_years=8),
    ]

    debt_facilities = [
        DebtFacility(
            name="Mortgage",
            principal=750_000,
            annual_interest_rate=0.03,
            draw_month=0,
            grace_months=6,
            term_months=120,
            repayment_type="linear",
        ),
        DebtFacility(
            name="Loan A",
            principal=450_000,
            annual_interest_rate=0.025,
            draw_month=5,
            grace_months=4,
            term_months=60,
            repayment_type="annuity",
        ),
        DebtFacility(
            name="Loan B",
            principal=200_000,
            annual_interest_rate=0.015,
            draw_month=58,
            grace_months=0,
            term_months=36,
            repayment_type="linear",
        ),
    ]

    equity_injections = {0: 5_500_000.0, 12: 1_000_000.0, 36: 750_000.0}

    def _year_row(values: list[float]) -> dict[str, float]:
        padded = list(values) + ([values[-1]] * max(len(year_cols) - len(values), 0))
        return {col: float(padded[i]) for i, col in enumerate(year_cols)}

    direct_labor_schedule = pd.DataFrame(
        [
            {
                "role": "Brewhouse Operators",
                "allocation_driver": "liters",
                "scope": "global",
                "target_sku_id": np.nan,
                "monthly_cost_per_fte": 3000.0,
                "annual_raise_pct": 0.03,
                "benefits_pct": 0.18,
                "payroll_tax_pct": 0.08,
                "overtime_pct": 0.05,
                "capacity_liters_per_fte_month": 30_000.0,
                **_year_row([5, 5, 5, 5, 5, 6, 6, 6, 6, 6]),
            },
            {
                "role": "Packaging Technicians",
                "allocation_driver": "liters",
                "scope": "global",
                "target_sku_id": np.nan,
                "monthly_cost_per_fte": 2600.0,
                "annual_raise_pct": 0.03,
                "benefits_pct": 0.16,
                "payroll_tax_pct": 0.08,
                "overtime_pct": 0.04,
                "capacity_liters_per_fte_month": 22_000.0,
                **_year_row([3, 3, 3, 4, 4, 4, 4, 4, 5, 5]),
            },
            {
                "role": "Quality & Cellar",
                "allocation_driver": "liters",
                "scope": "global",
                "target_sku_id": np.nan,
                "monthly_cost_per_fte": 2800.0,
                "annual_raise_pct": 0.03,
                "benefits_pct": 0.18,
                "payroll_tax_pct": 0.08,
                "overtime_pct": 0.03,
                "capacity_liters_per_fte_month": 18_000.0,
                **_year_row([1, 1, 1, 1, 1, 1, 2, 2, 2, 2]),
            },
        ]
    )

    sku_operations = pd.DataFrame(
        [
            {
                "sku_id": 1,
                "opening_fg_liters": 1_500.0,
                "brew_batch_liters": 2_500.0,
                "brewhouse_yield_pct": 0.97,
                "cellar_yield_pct": 0.98,
                "packaging_yield_pct": 0.995,
                "fermentation_days": 14.0,
                "conditioning_days": 7.0,
                "target_fg_days": 10.0,
            },
            {
                "sku_id": 2,
                "opening_fg_liters": 1_200.0,
                "brew_batch_liters": 2_500.0,
                "brewhouse_yield_pct": 0.97,
                "cellar_yield_pct": 0.98,
                "packaging_yield_pct": 0.995,
                "fermentation_days": 12.0,
                "conditioning_days": 7.0,
                "target_fg_days": 9.0,
            },
            {
                "sku_id": 3,
                "opening_fg_liters": 800.0,
                "brew_batch_liters": 2_500.0,
                "brewhouse_yield_pct": 0.96,
                "cellar_yield_pct": 0.975,
                "packaging_yield_pct": 0.992,
                "fermentation_days": 16.0,
                "conditioning_days": 8.0,
                "target_fg_days": 12.0,
            },
        ]
    )

    brewhouse_schedule = pd.DataFrame(
        [
            {
                "resource_name": "Main Brewhouse",
                "liters_per_batch": 2_500.0,
                "batches_per_day": 3.0,
                "brew_days_per_month": 22.0,
                "utilization_pct": 0.83,
                "downtime_pct": 0.06,
                "changeover_loss_pct": 0.04,
                **_year_row([1, 1, 1, 1, 1, 2, 2, 2, 2, 2]),
            }
        ]
    )

    cellar_schedule = pd.DataFrame(
        [
            {
                "resource_name": "Unitank Cellar",
                "tank_count": 8.0,
                "liters_per_tank": 5_000.0,
                "utilization_pct": 0.88,
                "downtime_pct": 0.05,
                **_year_row([1, 1, 1, 1, 1, 1.5, 1.5, 1.5, 1.5, 1.5]),
            }
        ]
    )

    packaging_schedule = pd.DataFrame(
        [
            {
                "resource_name": "Bottling & Can Line",
                "liters_per_hour": 1_350.0,
                "hours_per_day": 9.0,
                "run_days_per_month": 20.0,
                "utilization_pct": 0.85,
                "downtime_pct": 0.07,
                "changeover_loss_pct": 0.03,
                **_year_row([1, 1, 1, 1, 1, 1, 1.25, 1.25, 1.25, 1.25]),
            }
        ]
    )

    indirect_labor_schedule = pd.DataFrame(
        [
            {
                "role": "Sales & Distribution",
                "allocation_driver": "fixed",
                "scope": "global",
                "target_sku_id": np.nan,
                "monthly_cost_per_fte": 3200.0,
                "annual_raise_pct": 0.03,
                "benefits_pct": 0.18,
                "payroll_tax_pct": 0.08,
                "overtime_pct": 0.02,
                "capacity_liters_per_fte_month": 0.0,
                **_year_row([2, 2, 2, 2, 3, 3, 3, 3, 3, 4]),
            },
            {
                "role": "Finance & Admin",
                "allocation_driver": "fixed",
                "scope": "global",
                "target_sku_id": np.nan,
                "monthly_cost_per_fte": 3500.0,
                "annual_raise_pct": 0.03,
                "benefits_pct": 0.18,
                "payroll_tax_pct": 0.08,
                "overtime_pct": 0.01,
                "capacity_liters_per_fte_month": 0.0,
                **_year_row([2, 2, 2, 2, 2, 2, 3, 3, 3, 3]),
            },
        ]
    )

    inventory_schedule = pd.DataFrame(
        [
            {
                "stage": "Raw Materials",
                "cost_share_pct": 0.45,
                "writeoff_pct": 0.002,
                "reserve_pct": 0.010,
                **_year_row([22, 22, 22, 23, 23, 23, 24, 24, 24, 24]),
            },
            {
                "stage": "WIP",
                "cost_share_pct": 0.15,
                "writeoff_pct": 0.001,
                "reserve_pct": 0.005,
                **_year_row([5, 5, 5, 5, 6, 6, 6, 6, 6, 6]),
            },
            {
                "stage": "Finished Goods",
                "cost_share_pct": 0.40,
                "writeoff_pct": 0.003,
                "reserve_pct": 0.015,
                **_year_row([14, 14, 14, 15, 15, 15, 15, 15, 16, 16]),
            },
        ]
    )

    receivables_schedule = pd.DataFrame(
        [
            {"channel": "Wholesale", "trade_spend_pct": 0.045, "returns_pct": 0.010, "bad_debt_pct": 0.004, **_year_row([32, 32, 31, 31, 30, 30, 30, 29, 29, 28])},
            {"channel": "Retail", "trade_spend_pct": 0.020, "returns_pct": 0.004, "bad_debt_pct": 0.001, **_year_row([12, 12, 12, 12, 11, 11, 11, 11, 10, 10])},
            {"channel": "E-Commerce", "trade_spend_pct": 0.015, "returns_pct": 0.015, "bad_debt_pct": 0.002, **_year_row([6, 6, 6, 6, 5, 5, 5, 5, 5, 5])},
            {"channel": "On-Premise", "trade_spend_pct": 0.010, "returns_pct": 0.003, "bad_debt_pct": 0.001, **_year_row([10, 10, 10, 9, 9, 9, 9, 9, 8, 8])},
            {"channel": "Export", "trade_spend_pct": 0.030, "returns_pct": 0.006, "bad_debt_pct": 0.003, **_year_row([45, 45, 44, 44, 43, 42, 42, 41, 41, 40])},
        ]
    )

    payables_schedule = pd.DataFrame(
        [
            {"supplier_category": "Malt & Grain", "cost_share_pct": 0.30, "early_pay_discount_pct": 0.004, "discount_capture_pct": 0.25, **_year_row([35, 35, 35, 34, 34, 34, 33, 33, 33, 32])},
            {"supplier_category": "Hops & Yeast", "cost_share_pct": 0.12, "early_pay_discount_pct": 0.003, "discount_capture_pct": 0.20, **_year_row([28, 28, 28, 28, 27, 27, 27, 27, 26, 26])},
            {"supplier_category": "Packaging", "cost_share_pct": 0.20, "early_pay_discount_pct": 0.005, "discount_capture_pct": 0.30, **_year_row([30, 30, 30, 29, 29, 29, 28, 28, 28, 28])},
            {"supplier_category": "Utilities & Services", "cost_share_pct": 0.10, "early_pay_discount_pct": 0.001, "discount_capture_pct": 0.10, **_year_row([20, 20, 20, 20, 19, 19, 19, 19, 18, 18])},
        ]
    )

    return ModelInputs(
        skus=skus,
        channels=channels,
        sales_plan=sales_plan,
        cost_pools=cost_pools,
        other_income_items=other_income_items,
        direct_labor_schedule=direct_labor_schedule,
        indirect_labor_schedule=indirect_labor_schedule,
        inventory_schedule=inventory_schedule,
        receivables_schedule=receivables_schedule,
        payables_schedule=payables_schedule,
        sku_operations=sku_operations,
        brewhouse_schedule=brewhouse_schedule,
        cellar_schedule=cellar_schedule,
        packaging_schedule=packaging_schedule,
        capex_items=capex_items,
        debt_facilities=debt_facilities,
        equity_injections=equity_injections,
    )


def _run_model(cfg: ModelConfig, div: DividendPolicy) -> Tuple[ModelInputs, MicrobreweryFinancialModel]:
    inputs = _build_sample_assumptions(cfg, div)
    model = MicrobreweryFinancialModel(cfg, div, inputs)
    return inputs, model


def _apply_yearly_increment(df: pd.DataFrame, increment_pct: float) -> pd.DataFrame:
    out = df.copy()
    year_cols = []
    for c in out.columns:
        if isinstance(c, str) and c.lower().startswith("year "):
            try:
                y = int(c.split()[-1])
                year_cols.append((c, y))
            except ValueError:
                continue
    if year_cols:
        for c, y in sorted(year_cols, key=lambda x: x[1]):
            factor = (1.0 + increment_pct) ** max(y - 1, 0)
            series = pd.to_numeric(out[c], errors="coerce")
            out[c] = series.where(series.isna(), series * factor)
    else:
        numeric_cols = out.select_dtypes(include=["number"]).columns
        for c in numeric_cols:
            out[c] = out[c] * (1.0 + increment_pct)
    return out


def _assumption_editor(title: str, key: str, df: pd.DataFrame) -> pd.DataFrame:
    st.markdown(f"### {title}")
    data_key = f"assump_data_{key}"
    saved_key = f"assump_saved_{key}"
    work_key = f"assump_work_{key}"
    edit_key = f"assump_edit_{key}"
    inc_key = f"assump_inc_{key}"
    if data_key not in st.session_state:
        st.session_state[data_key] = df.copy()
    if saved_key not in st.session_state:
        st.session_state[saved_key] = st.session_state[data_key].copy()
    if work_key not in st.session_state:
        st.session_state[work_key] = st.session_state[saved_key].copy()
    if edit_key not in st.session_state:
        st.session_state[edit_key] = False
    if inc_key not in st.session_state:
        st.session_state[inc_key] = 0.0

    c1, c2, c3, c4, c5 = st.columns([1, 1, 1, 1, 2])
    if c1.button("Edit", key=f"btn_edit_{key}"):
        st.session_state[edit_key] = not st.session_state[edit_key]
        if st.session_state[edit_key]:
            st.session_state[work_key] = st.session_state[saved_key].copy()
    if c2.button("Add row", key=f"btn_addrow_{key}") and st.session_state[edit_key]:
        blank = {col: None for col in st.session_state[work_key].columns}
        st.session_state[work_key] = pd.concat(
            [st.session_state[work_key], pd.DataFrame([blank])],
            ignore_index=True,
        )
    if c3.button("Delete row", key=f"btn_delrow_{key}") and st.session_state[edit_key]:
        wdf = st.session_state[work_key]
        if not wdf.empty:
            sel = st.session_state.get(f"assump_row_select_{key}", 0)
            if sel in wdf.index:
                st.session_state[work_key] = wdf.drop(index=sel).reset_index(drop=True)
    if c4.button("Reset changes", key=f"btn_reset_{key}"):
        st.session_state[work_key] = st.session_state[saved_key].copy()
        st.session_state[data_key] = st.session_state[saved_key].copy()
    inc_col1, inc_col2 = st.columns([1, 1])
    increment_pct = inc_col1.number_input(
        "Yearly Increment Percentage",
        min_value=-1.0,
        max_value=2.0,
        value=float(st.session_state[inc_key]),
        step=0.01,
        key=f"input_inc_{key}",
        format="%.4f",
    )
    st.session_state[inc_key] = increment_pct
    if inc_col2.button("Propagate Increment", key=f"btn_prop_{key}"):
        target_key = work_key if st.session_state[edit_key] else data_key
        st.session_state[target_key] = _apply_yearly_increment(st.session_state[target_key], increment_pct)

    if st.session_state[edit_key]:
        wdf = st.session_state[work_key]
        if not wdf.empty:
            row_options = list(wdf.index)
            selected_row = st.selectbox(
                "Row selection",
                options=row_options,
                key=f"assump_row_select_{key}",
                format_func=lambda row_idx: _row_selection_label(wdf, row_idx),
            )
            st.markdown("**Editable row form fields**")
            updated_row = {}
            for col in wdf.columns:
                val = wdf.loc[selected_row, col]
                col_series = wdf[col]
                if pd.api.types.is_bool_dtype(col_series):
                    updated_row[col] = st.checkbox(
                        f"{col}",
                        value=_coerce_bool(val),
                        key=f"field_{key}_{selected_row}_{col}",
                    )
                elif pd.api.types.is_numeric_dtype(col_series):
                    default_val = 0.0 if pd.isna(val) else float(val)
                    updated_row[col] = st.number_input(
                        f"{col}",
                        value=default_val,
                        key=f"field_{key}_{selected_row}_{col}",
                    )
                else:
                    updated_row[col] = st.text_input(
                        f"{col}",
                        value="" if pd.isna(val) else str(val),
                        key=f"field_{key}_{selected_row}_{col}",
                    )

            if st.button("Apply row changes", key=f"btn_apply_row_{key}"):
                for col, val in updated_row.items():
                    casted = _cast_value_for_dtype(st.session_state[work_key][col], val)
                    st.session_state[work_key].at[selected_row, col] = casted
        else:
            st.info("No rows available. Use 'Add row' to create one.")

        b2, b3, b4 = st.columns(3)
        if b2.button("Save changes", key=f"btn_save_{key}"):
            st.session_state[saved_key] = st.session_state[work_key].copy()
            st.session_state[data_key] = st.session_state[saved_key].copy()
        if b3.button("Discard edits", key=f"btn_discard_{key}"):
            st.session_state[work_key] = st.session_state[saved_key].copy()
        if b4.button("Close editor", key=f"btn_close_{key}"):
            st.session_state[edit_key] = False
            st.session_state[data_key] = st.session_state[saved_key].copy()

        st.dataframe(st.session_state[work_key], use_container_width=True)
    else:
        st.dataframe(st.session_state[data_key], use_container_width=True)
    return st.session_state[data_key]


def _non_empty_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    return df.dropna(how="all").copy()


def _coerce_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "y"}


def _cast_value_for_dtype(column: pd.Series, value: object) -> object:
    if pd.api.types.is_bool_dtype(column):
        return _coerce_bool(value)
    if pd.api.types.is_integer_dtype(column):
        if value is None or (isinstance(value, float) and np.isnan(value)):
            return 0
        return int(float(value))
    if pd.api.types.is_float_dtype(column):
        if value is None or (isinstance(value, float) and np.isnan(value)):
            return 0.0
        return float(value)
    if pd.api.types.is_datetime64_any_dtype(column):
        if value is None or (isinstance(value, float) and np.isnan(value)):
            return pd.NaT
        return pd.to_datetime(value, errors="coerce")
    return value


def _format_row_label_value(value: object, col_name: str) -> str:
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, pd.Timestamp):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (float, np.floating)):
        if not np.isfinite(value):
            return ""
        text = str(int(value)) if float(value).is_integer() else f"{float(value):g}"
    else:
        text = str(value).strip()
    if not text:
        return ""
    if col_name == "sku_id":
        return f"SKU {text}"
    return text


def _row_selection_label(df: pd.DataFrame, row_idx: object, prefix: str = "Row") -> str:
    if df is None or not isinstance(df, pd.DataFrame) or row_idx not in df.index:
        return f"{prefix} {row_idx}"

    row = df.loc[row_idx]
    if isinstance(row, pd.DataFrame):
        row = row.iloc[0]

    preferred_cols = [
        "name",
        "role",
        "stage",
        "channel",
        "supplier_category",
        "other_income_name",
        "case",
        "scenario",
        "variable",
        "month",
        "quarter",
        "year",
        "date",
        "sku_name",
        "sku_id",
        "product_name",
        "product",
        "metric",
    ]
    parts: list[str] = []
    seen: set[str] = set()

    def add_part(col_name: object) -> None:
        if col_name not in df.columns:
            return
        if isinstance(col_name, str) and col_name.startswith("Year "):
            return
        value = _format_row_label_value(row.get(col_name), str(col_name))
        if not value:
            return
        normalized = value.lower()
        if normalized in seen:
            return
        seen.add(normalized)
        parts.append(value)

    for col_name in preferred_cols:
        add_part(col_name)
        if len(parts) >= 3:
            break

    if not parts:
        for col_name in df.columns:
            if len(parts) >= 3:
                break
            if isinstance(col_name, str) and col_name.startswith("Year "):
                continue
            if col_name != "sku_id" and pd.api.types.is_numeric_dtype(df[col_name]):
                continue
            add_part(col_name)

    label = " | ".join(parts) if parts else f"{prefix} {row_idx}"
    if isinstance(row_idx, (int, np.integer)):
        return f"{label} [Row {int(row_idx) + 1}]"
    return f"{label} [{prefix} {row_idx}]"


def _year_columns_for_months(months: int) -> list[str]:
    years = max(int(np.ceil(float(months) / 12.0)), 1)
    return [f"Year {i}" for i in range(1, years + 1)]


def _align_year_schedule_df(df: pd.DataFrame, template_df: pd.DataFrame) -> pd.DataFrame:
    if df is None or not isinstance(df, pd.DataFrame):
        return template_df.copy()
    out = df.copy()
    template_cols = list(template_df.columns)
    year_cols = [c for c in template_cols if isinstance(c, str) and c.startswith("Year ")]
    existing_year_cols = [c for c in out.columns if isinstance(c, str) and c.startswith("Year ")]
    last_year_series = None
    if existing_year_cols:
        last_year_series = pd.to_numeric(out[existing_year_cols[-1]], errors="coerce").fillna(0.0)
    for col in year_cols:
        if col not in out.columns:
            out[col] = last_year_series if last_year_series is not None else 0.0
    extra_year_cols = [c for c in out.columns if isinstance(c, str) and c.startswith("Year ") and c not in year_cols]
    if extra_year_cols:
        out = out.drop(columns=extra_year_cols, errors="ignore")
    missing_non_year = [c for c in template_cols if c not in out.columns]
    for col in missing_non_year:
        out[col] = template_df[col] if col in template_df.columns else np.nan
    ordered = [c for c in template_cols if c in out.columns]
    extras = [c for c in out.columns if c not in ordered]
    return out[ordered + extras]


def _sync_assumption_table_state(key: str, default_df: pd.DataFrame) -> None:
    data_key = f"assump_data_{key}"
    saved_key = f"assump_saved_{key}"
    work_key = f"assump_work_{key}"
    edit_key = f"assump_edit_{key}"
    inc_key = f"assump_inc_{key}"

    for state_key in (data_key, saved_key, work_key):
        current = st.session_state.get(state_key)
        if isinstance(current, pd.DataFrame):
            st.session_state[state_key] = _align_year_schedule_df(current, default_df)
        else:
            st.session_state[state_key] = default_df.copy()
    if edit_key not in st.session_state:
        st.session_state[edit_key] = False
    if inc_key not in st.session_state:
        st.session_state[inc_key] = 0.0


def _expand_sales_plan_to_monthly(sales_df: pd.DataFrame, frequency: str) -> pd.DataFrame:
    if sales_df is None or sales_df.empty:
        return pd.DataFrame(columns=["date", "sku_id", "channel", "units"])
    base = sales_df.copy()
    if "date" not in base.columns:
        return pd.DataFrame(columns=["date", "sku_id", "channel", "units"])
    base["date"] = pd.to_datetime(base["date"], errors="coerce")
    base["units"] = pd.to_numeric(base.get("units", 0.0), errors="coerce").fillna(0.0)
    base = base.dropna(subset=["date"])
    if frequency == "monthly":
        return base[["date", "sku_id", "channel", "units"]].copy()

    rows = []
    for _, r in base.iterrows():
        dt = pd.to_datetime(r["date"])
        if frequency == "quarterly":
            start = dt.to_period("Q").start_time
            month_points = pd.date_range(start=start, periods=3, freq="MS")
        else:
            start = dt.to_period("Y").start_time
            month_points = pd.date_range(start=start, periods=12, freq="MS")
        portion = float(r["units"]) / max(len(month_points), 1)
        for month_dt in month_points:
            rows.append(
                {
                    "date": month_dt,
                    "sku_id": r.get("sku_id"),
                    "channel": r.get("channel"),
                    "units": portion,
                }
            )
    return pd.DataFrame(rows, columns=["date", "sku_id", "channel", "units"])


def _compress_monthly_sales_plan(monthly_df: pd.DataFrame, frequency: str) -> pd.DataFrame:
    if monthly_df is None or monthly_df.empty:
        return pd.DataFrame(columns=["date", "sku_id", "channel", "units"])
    out = monthly_df.copy()
    out["date"] = pd.to_datetime(out["date"], errors="coerce")
    out["units"] = pd.to_numeric(out.get("units", 0.0), errors="coerce").fillna(0.0)
    out = out.dropna(subset=["date"])
    if frequency == "monthly":
        out["date"] = out["date"].dt.to_period("M").dt.start_time
    elif frequency == "quarterly":
        out["date"] = out["date"].dt.to_period("Q").dt.start_time
    else:
        out["date"] = out["date"].dt.to_period("Y").dt.start_time
    out = (
        out.groupby(["date", "sku_id", "channel"], dropna=False, as_index=False)["units"]
        .sum()
        .sort_values(["date", "sku_id", "channel"])
        .reset_index(drop=True)
    )
    return out


def _fill_monthly_sales_plan_to_timeline(
    monthly_df: pd.DataFrame,
    idx: pd.DatetimeIndex,
    fallback_monthly_df: pd.DataFrame,
) -> pd.DataFrame:
    monthly = monthly_df.copy()
    monthly["date"] = pd.to_datetime(monthly.get("date"), errors="coerce")
    monthly["units"] = pd.to_numeric(monthly.get("units", 0.0), errors="coerce").fillna(0.0)
    monthly = monthly.dropna(subset=["date"])
    monthly = monthly[monthly["date"].isin(idx)].copy()

    fallback = fallback_monthly_df.copy()
    fallback["date"] = pd.to_datetime(fallback.get("date"), errors="coerce")
    fallback = fallback.dropna(subset=["date"])
    fallback = fallback[fallback["date"].isin(idx)].copy()

    if monthly.empty:
        monthly = fallback.copy()

    combos = monthly[["sku_id", "channel"]].drop_duplicates() if not monthly.empty else fallback[["sku_id", "channel"]].drop_duplicates()
    if combos.empty:
        return pd.DataFrame(columns=["date", "sku_id", "channel", "units"])

    skeleton_rows = []
    for _, combo in combos.iterrows():
        for dt in idx:
            skeleton_rows.append({"date": dt, "sku_id": combo["sku_id"], "channel": combo["channel"], "units": 0.0})
    skeleton = pd.DataFrame(skeleton_rows)

    observed = (
        monthly.groupby(["date", "sku_id", "channel"], as_index=False)["units"]
        .sum()
    )
    merged = skeleton.merge(observed, on=["date", "sku_id", "channel"], how="left", suffixes=("_base", "_obs"))
    merged["units"] = merged["units_obs"].fillna(merged["units_base"])
    return merged[["date", "sku_id", "channel", "units"]].sort_values(["date", "sku_id", "channel"]).reset_index(drop=True)


def _sales_plan_template_idx(idx: pd.DatetimeIndex) -> pd.DatetimeIndex:
    return idx[: min(len(idx), 12)]


def _rebase_monthly_sales_plan_to_idx(monthly_df: pd.DataFrame, target_idx: pd.DatetimeIndex) -> pd.DataFrame:
    if monthly_df is None or monthly_df.empty or len(target_idx) == 0:
        return pd.DataFrame(columns=["date", "sku_id", "channel", "units"])
    out = monthly_df.copy()
    out["date"] = pd.to_datetime(out.get("date"), errors="coerce")
    out["units"] = pd.to_numeric(out.get("units", 0.0), errors="coerce").fillna(0.0)
    out = out.dropna(subset=["date"])
    if out.empty:
        return pd.DataFrame(columns=["date", "sku_id", "channel", "units"])
    source_dates = list(pd.Index(sorted(out["date"].drop_duplicates()))[: len(target_idx)])
    if not source_dates:
        return pd.DataFrame(columns=["date", "sku_id", "channel", "units"])
    remap = {src: dst for src, dst in zip(source_dates, list(target_idx))}
    out = out[out["date"].isin(remap)].copy()
    out["date"] = out["date"].map(remap)
    return (
        out.groupby(["date", "sku_id", "channel"], dropna=False, as_index=False)["units"]
        .sum()
        .sort_values(["date", "sku_id", "channel"])
        .reset_index(drop=True)
    )


def _sales_plan_editor_monthly_view(
    sales_df: pd.DataFrame,
    frequency: str,
    idx: pd.DatetimeIndex,
    propagation_mode: str,
) -> pd.DataFrame:
    monthly = _expand_sales_plan_to_monthly(sales_df, frequency)
    if propagation_mode == "manual":
        return monthly
    template_idx = _sales_plan_template_idx(idx)
    monthly = _rebase_monthly_sales_plan_to_idx(monthly, template_idx)
    monthly = _fill_monthly_sales_plan_to_timeline(monthly, template_idx, monthly)
    return monthly


def _sales_plan_editor_view(sales_df: pd.DataFrame, frequency: str) -> pd.DataFrame:
    if sales_df is None or sales_df.empty:
        label = {"monthly": "month", "quarterly": "quarter", "yearly": "year"}[frequency]
        return pd.DataFrame(columns=[label, "sku_id", "channel", "units"])
    view = sales_df.copy()
    view["date"] = pd.to_datetime(view["date"], errors="coerce")
    if frequency == "monthly":
        view.insert(0, "month", view["date"].dt.to_period("M").astype(str))
        view = view.drop(columns=["date"])
    elif frequency == "quarterly":
        view.insert(0, "quarter", view["date"].dt.to_period("Q").astype(str).str.replace("Q", "-Q", regex=False))
        view = view.drop(columns=["date"])
    else:
        view.insert(0, "year", view["date"].dt.year.astype("Int64").astype(str))
        view = view.drop(columns=["date"])
    return view


def _sales_plan_editor_view_to_base(view_df: pd.DataFrame, frequency: str) -> pd.DataFrame:
    if view_df is None or view_df.empty:
        return pd.DataFrame(columns=["date", "sku_id", "channel", "units"])
    out = view_df.copy()
    safe_col = lambda name, default="": out[name] if name in out.columns else pd.Series([default] * len(out), index=out.index)
    if frequency == "monthly":
        out["date"] = pd.to_datetime(safe_col("month"), errors="coerce").dt.to_period("M").dt.start_time
        out = out.drop(columns=["month"], errors="ignore")
    elif frequency == "quarterly":
        quarter_raw = safe_col("quarter").astype(str).str.upper().str.replace(" ", "", regex=False).str.replace("-Q", "Q", regex=False)
        extracted = quarter_raw.str.extract(r"(?P<year>\d{4})Q(?P<q>[1-4])")
        year = pd.to_numeric(extracted["year"], errors="coerce")
        q = pd.to_numeric(extracted["q"], errors="coerce")
        month = ((q - 1) * 3 + 1).astype("Int64")
        out["date"] = pd.to_datetime(
            year.astype("Int64").astype(str) + "-" + month.astype(str).str.zfill(2) + "-01",
            errors="coerce",
        )
        out = out.drop(columns=["quarter"], errors="ignore")
    else:
        year_numeric = pd.to_numeric(safe_col("year"), errors="coerce")
        out["date"] = pd.to_datetime(year_numeric.astype("Int64").astype(str) + "-01-01", errors="coerce")
        out = out.drop(columns=["year"], errors="ignore")
    out["units"] = pd.to_numeric(out.get("units", 0.0), errors="coerce")
    return out[["date", "sku_id", "channel", "units"]]


def _dynamic_table_editor(title: str, key: str, default_df: pd.DataFrame, label: str = "row") -> pd.DataFrame:
    data_key = f"dyn_data_{key}"
    saved_key = f"dyn_saved_{key}"
    work_key = f"dyn_work_{key}"
    edit_key = f"dyn_edit_{key}"
    select_key = f"dyn_select_{key}"
    if data_key not in st.session_state:
        st.session_state[data_key] = default_df.copy()
    if saved_key not in st.session_state:
        st.session_state[saved_key] = st.session_state[data_key].copy()
    if work_key not in st.session_state:
        st.session_state[work_key] = st.session_state[saved_key].copy()
    if edit_key not in st.session_state:
        st.session_state[edit_key] = False

    st.markdown(f"#### {title}")
    c1, c2, c3, c4 = st.columns([1, 1, 1, 1])
    if c1.button("Edit", key=f"dyn_edit_btn_{key}"):
        st.session_state[edit_key] = not st.session_state[edit_key]
        if st.session_state[edit_key]:
            st.session_state[work_key] = st.session_state[saved_key].copy()
    if c2.button(f"Add {label}", key=f"dyn_add_btn_{key}") and st.session_state[edit_key]:
        blank = {col: None for col in st.session_state[work_key].columns}
        st.session_state[work_key] = pd.concat([st.session_state[work_key], pd.DataFrame([blank])], ignore_index=True)
    if c3.button(f"Delete {label}", key=f"dyn_del_btn_{key}") and st.session_state[edit_key]:
        wdf = st.session_state[work_key]
        if not wdf.empty:
            sel = st.session_state.get(select_key, 0)
            if sel in wdf.index:
                st.session_state[work_key] = wdf.drop(index=sel).reset_index(drop=True)
    if c4.button("Reset changes", key=f"dyn_reset_btn_{key}"):
        st.session_state[work_key] = st.session_state[saved_key].copy()
        st.session_state[data_key] = st.session_state[saved_key].copy()

    if st.session_state[edit_key]:
        wdf = st.session_state[work_key]
        if not wdf.empty:
            selected = st.selectbox(
                f"{label.title()} selection",
                options=list(wdf.index),
                key=select_key,
                format_func=lambda row_idx: _row_selection_label(wdf, row_idx, prefix=label.title()),
            )
            updates = {}
            for col in wdf.columns:
                v = wdf.loc[selected, col]
                if pd.api.types.is_numeric_dtype(wdf[col]):
                    updates[col] = st.number_input(f"{col}", value=0.0 if pd.isna(v) else float(v), key=f"dyn_field_{key}_{selected}_{col}")
                else:
                    updates[col] = st.text_input(f"{col}", value="" if pd.isna(v) else str(v), key=f"dyn_field_{key}_{selected}_{col}")
            b1, b2, b3, b4 = st.columns(4)
            if b1.button(f"Apply {label} changes", key=f"dyn_apply_btn_{key}"):
                for col, val in updates.items():
                    st.session_state[work_key].loc[selected, col] = val
            if b2.button("Save changes", key=f"dyn_save_btn_{key}"):
                st.session_state[saved_key] = st.session_state[work_key].copy()
                st.session_state[data_key] = st.session_state[saved_key].copy()
            if b3.button("Discard edits", key=f"dyn_discard_btn_{key}"):
                st.session_state[work_key] = st.session_state[saved_key].copy()
            if b4.button("Close editor", key=f"dyn_close_btn_{key}"):
                st.session_state[edit_key] = False
                st.session_state[data_key] = st.session_state[saved_key].copy()
        st.dataframe(st.session_state[work_key], use_container_width=True)
    else:
        st.dataframe(st.session_state[data_key], use_container_width=True)
    return st.session_state[data_key]


def _build_inputs_from_state(base_inputs: ModelInputs) -> ModelInputs:
    skus = _non_empty_rows(st.session_state.get("assump_data_skus", base_inputs.skus)).copy()
    channels = _non_empty_rows(st.session_state.get("assump_data_channels", base_inputs.channels)).copy()
    sales_plan = _non_empty_rows(st.session_state.get("assump_data_sales_plan_base", base_inputs.sales_plan)).copy()
    sales_plan_frequency = st.session_state.get("assump_sales_plan_frequency", base_inputs.sales_plan_frequency)
    sales_plan_propagation_mode = st.session_state.get(
        "assump_sales_plan_propagation_mode",
        getattr(base_inputs, "sales_plan_propagation_mode", "manual"),
    )
    sales_plan_propagation_growth_annual = st.session_state.get(
        "assump_sales_plan_propagation_growth_annual",
        getattr(base_inputs, "sales_plan_propagation_growth_annual", 0.0),
    )

    cp_src = st.session_state.get("assump_data_cost_pools")
    if isinstance(cp_src, pd.DataFrame):
        cp_src = _non_empty_rows(cp_src)
        cost_pools = [
            CostPoolInput(
                name=str(r["name"]),
                cost_type=str(r.get("cost_type", "indirect")),
                behavior=str(r.get("behavior", "blended")),
                allocation_driver=str(r.get("allocation_driver", "units")),
                scope=str(r.get("scope", "global")),
                channel=None if pd.isna(r.get("channel")) else str(r.get("channel")),
                unit_variable_cost=float(r.get("unit_variable_cost", 0.0)),
                fixed_monthly_cost=float(r.get("fixed_monthly_cost", 0.0)),
                step_threshold=float(r.get("step_threshold", 0.0)),
                step_increment=float(r.get("step_increment", 0.0)),
            )
            for _, r in cp_src.iterrows()
            if pd.notna(r.get("name"))
        ]
    else:
        cost_pools = base_inputs.cost_pools

    oi_src = st.session_state.get("assump_data_other_income")
    if isinstance(oi_src, pd.DataFrame):
        oi_src = _non_empty_rows(oi_src)
        other_income_items = [
            OtherIncomeItem(
                other_income_name=str(r.get("other_income_name")),
                amount=float(r.get("amount", 0.0)),
                active=_coerce_bool(r.get("active", True)),
                category=None if pd.isna(r.get("category")) else str(r.get("category")),
                notes=None if pd.isna(r.get("notes")) else str(r.get("notes")),
            )
            for _, r in oi_src.iterrows()
            if pd.notna(r.get("other_income_name"))
        ]
    else:
        other_income_items = base_inputs.other_income_items

    capex_src = st.session_state.get("assump_data_capex")
    if isinstance(capex_src, pd.DataFrame):
        capex_src = _non_empty_rows(capex_src)
        capex_items = [
            CapexItem(
                name=str(r["name"]),
                amount=float(r["amount"]),
                capex_month=int(r["capex_month"]),
                depreciation_years=float(r["depreciation_years"]),
            )
            for _, r in capex_src.iterrows()
            if pd.notna(r.get("name")) and pd.notna(r.get("amount")) and pd.notna(r.get("capex_month"))
        ]
    else:
        capex_items = base_inputs.capex_items

    debt_src = st.session_state.get("assump_data_debt")
    if isinstance(debt_src, pd.DataFrame):
        debt_src = _non_empty_rows(debt_src)
        debt_facilities = [
            DebtFacility(
                name=str(r["name"]),
                principal=float(r["principal"]),
                annual_interest_rate=float(r["annual_interest_rate"]),
                draw_month=int(r["draw_month"]),
                grace_months=int(r["grace_months"]),
                term_months=int(r["term_months"]),
                repayment_type=str(r["repayment_type"]),
            )
            for _, r in debt_src.iterrows()
            if pd.notna(r.get("name")) and pd.notna(r.get("principal"))
        ]
    else:
        debt_facilities = base_inputs.debt_facilities

    equity_src = st.session_state.get("assump_data_equity")
    if isinstance(equity_src, pd.DataFrame):
        equity_src = _non_empty_rows(equity_src)
        equity_injections = {
            int(r["month"]): float(r["equity_injection"])
            for _, r in equity_src.iterrows()
            if pd.notna(r.get("month")) and pd.notna(r.get("equity_injection"))
        }
    else:
        equity_injections = base_inputs.equity_injections

    direct_labor_schedule = st.session_state.get("assump_data_direct_labor")
    if isinstance(direct_labor_schedule, pd.DataFrame):
        direct_labor_schedule = _non_empty_rows(direct_labor_schedule).copy()
    else:
        direct_labor_schedule = base_inputs.direct_labor_schedule

    indirect_labor_schedule = st.session_state.get("assump_data_indirect_labor")
    if isinstance(indirect_labor_schedule, pd.DataFrame):
        indirect_labor_schedule = _non_empty_rows(indirect_labor_schedule).copy()
    else:
        indirect_labor_schedule = base_inputs.indirect_labor_schedule

    inventory_schedule = st.session_state.get("assump_data_inventory")
    if isinstance(inventory_schedule, pd.DataFrame):
        inventory_schedule = _non_empty_rows(inventory_schedule).copy()
    else:
        inventory_schedule = base_inputs.inventory_schedule

    receivables_schedule = st.session_state.get("assump_data_receivables")
    if isinstance(receivables_schedule, pd.DataFrame):
        receivables_schedule = _non_empty_rows(receivables_schedule).copy()
    else:
        receivables_schedule = base_inputs.receivables_schedule

    payables_schedule = st.session_state.get("assump_data_payables")
    if isinstance(payables_schedule, pd.DataFrame):
        payables_schedule = _non_empty_rows(payables_schedule).copy()
    else:
        payables_schedule = base_inputs.payables_schedule

    sku_operations = st.session_state.get("assump_data_sku_operations")
    if isinstance(sku_operations, pd.DataFrame):
        sku_operations = _non_empty_rows(sku_operations).copy()
    else:
        sku_operations = base_inputs.sku_operations

    brewhouse_schedule = st.session_state.get("assump_data_brewhouse")
    if isinstance(brewhouse_schedule, pd.DataFrame):
        brewhouse_schedule = _non_empty_rows(brewhouse_schedule).copy()
    else:
        brewhouse_schedule = base_inputs.brewhouse_schedule

    cellar_schedule = st.session_state.get("assump_data_cellar")
    if isinstance(cellar_schedule, pd.DataFrame):
        cellar_schedule = _non_empty_rows(cellar_schedule).copy()
    else:
        cellar_schedule = base_inputs.cellar_schedule

    packaging_schedule = st.session_state.get("assump_data_packaging")
    if isinstance(packaging_schedule, pd.DataFrame):
        packaging_schedule = _non_empty_rows(packaging_schedule).copy()
    else:
        packaging_schedule = base_inputs.packaging_schedule

    return ModelInputs(
        skus=skus,
        channels=channels,
        sales_plan=sales_plan,
        sales_plan_frequency=str(sales_plan_frequency),
        sales_plan_propagation_mode=str(sales_plan_propagation_mode),
        sales_plan_propagation_growth_annual=float(sales_plan_propagation_growth_annual),
        cost_pools=cost_pools,
        other_income_items=other_income_items,
        direct_labor_schedule=direct_labor_schedule,
        indirect_labor_schedule=indirect_labor_schedule,
        inventory_schedule=inventory_schedule,
        receivables_schedule=receivables_schedule,
        payables_schedule=payables_schedule,
        sku_operations=sku_operations,
        brewhouse_schedule=brewhouse_schedule,
        cellar_schedule=cellar_schedule,
        packaging_schedule=packaging_schedule,
        capex_items=capex_items,
        debt_facilities=debt_facilities,
        equity_injections=equity_injections,
    )


def _sales_plan_assumption_editor(
    base_sales_plan: pd.DataFrame,
    base_frequency: str,
    cfg: ModelConfig,
    base_propagation_mode: str = "manual",
    base_growth_annual: float = 0.0,
) -> pd.DataFrame:
    canonical_key = "assump_data_sales_plan_base"
    frequency_options = {"Monthly": "monthly", "Quarterly": "quarterly", "Yearly": "yearly"}
    reverse_options = {v: k for k, v in frequency_options.items()}
    propagation_options = {
        "Manual by period": "manual",
        "Repeat first year": "repeat_first_year",
        "Grow first year": "grow_first_year",
    }
    reverse_propagation_options = {v: k for k, v in propagation_options.items()}

    if "assump_sales_plan_frequency" not in st.session_state:
        st.session_state["assump_sales_plan_frequency"] = base_frequency
    if "assump_sales_plan_propagation_mode" not in st.session_state:
        st.session_state["assump_sales_plan_propagation_mode"] = base_propagation_mode
    if "assump_sales_plan_propagation_growth_annual" not in st.session_state:
        st.session_state["assump_sales_plan_propagation_growth_annual"] = float(base_growth_annual)
    current_frequency = str(st.session_state["assump_sales_plan_frequency"])
    if current_frequency not in reverse_options:
        current_frequency = "monthly"
    current_propagation_mode = str(st.session_state["assump_sales_plan_propagation_mode"])
    if current_propagation_mode not in reverse_propagation_options:
        current_propagation_mode = "manual"

    idx = pd.date_range(cfg.start_date, periods=cfg.months, freq="MS")
    template_idx = _sales_plan_template_idx(idx)
    control_col1, control_col2 = st.columns([1, 1])
    selected_label = control_col1.selectbox(
        "Sales plan frequency",
        options=list(frequency_options.keys()),
        index=list(frequency_options.values()).index(current_frequency),
        key="assump_sales_plan_frequency_selector",
        help="Sales plan schedule granularity shown in the table and used by the model.",
    )
    selected_propagation_label = control_col2.selectbox(
        "Sales plan propagation",
        options=list(propagation_options.keys()),
        index=list(propagation_options.values()).index(current_propagation_mode),
        key="assump_sales_plan_propagation_selector",
        help="Use the first production year as a reusable demand template instead of entering every later period manually.",
    )
    selected_frequency = frequency_options[selected_label]
    selected_propagation_mode = propagation_options[selected_propagation_label]
    selected_growth_annual = float(st.session_state.get("assump_sales_plan_propagation_growth_annual", base_growth_annual))
    if selected_propagation_mode == "grow_first_year":
        selected_growth_annual = st.number_input(
            "Annual sales growth after Year 1",
            value=selected_growth_annual,
            step=0.01,
            format="%.4f",
            key="assump_sales_plan_propagation_growth_input",
            help="Applied to the first-year sales template when propagating into later production years.",
        )
    st.session_state["assump_sales_plan_propagation_growth_annual"] = float(selected_growth_annual)

    previous_frequency = str(st.session_state.get("assump_prev_sales_plan_frequency", current_frequency))
    previous_propagation_mode = str(st.session_state.get("assump_prev_sales_plan_propagation_mode", current_propagation_mode))
    if canonical_key not in st.session_state and "assump_data_sales_plan" in st.session_state:
        st.session_state[canonical_key] = st.session_state.get("assump_data_sales_plan")
    canonical_plan = _non_empty_rows(st.session_state.get(canonical_key, base_sales_plan)).copy()
    if canonical_plan is None:
        canonical_plan = base_sales_plan.copy()

    fallback_monthly = _expand_sales_plan_to_monthly(base_sales_plan, base_frequency)
    fallback_monthly = propagate_sales_plan_monthly(
        fallback_monthly,
        idx,
        mode=base_propagation_mode,
        annual_growth_pct=float(base_growth_annual),
    )

    if selected_frequency != previous_frequency or selected_propagation_mode != previous_propagation_mode:
        monthly = _expand_sales_plan_to_monthly(canonical_plan, previous_frequency)
        if previous_propagation_mode == "manual":
            monthly = _fill_monthly_sales_plan_to_timeline(monthly, idx, fallback_monthly)
        else:
            monthly = _rebase_monthly_sales_plan_to_idx(monthly, template_idx)
            monthly = _fill_monthly_sales_plan_to_timeline(monthly, template_idx, fallback_monthly[fallback_monthly["date"].isin(template_idx)].copy())
        if selected_propagation_mode == "manual":
            canonical_plan = _compress_monthly_sales_plan(monthly, selected_frequency)
        else:
            canonical_plan = _compress_monthly_sales_plan(
                _rebase_monthly_sales_plan_to_idx(monthly, template_idx),
                selected_frequency,
            )
        st.session_state[canonical_key] = canonical_plan.copy()
        st.session_state["assump_data_sales_plan"] = _sales_plan_editor_view(canonical_plan, selected_frequency)
        st.session_state["assump_saved_sales_plan"] = st.session_state["assump_data_sales_plan"].copy()
        st.session_state["assump_work_sales_plan"] = st.session_state["assump_data_sales_plan"].copy()
        st.session_state["assump_edit_sales_plan"] = False

    st.session_state["assump_sales_plan_frequency"] = selected_frequency
    st.session_state["assump_prev_sales_plan_frequency"] = selected_frequency
    st.session_state["assump_sales_plan_propagation_mode"] = selected_propagation_mode
    st.session_state["assump_prev_sales_plan_propagation_mode"] = selected_propagation_mode

    if selected_propagation_mode == "manual":
        monthly_for_view = _sales_plan_editor_monthly_view(
            canonical_plan,
            selected_frequency,
            idx,
            selected_propagation_mode,
        )
        monthly_for_view = _fill_monthly_sales_plan_to_timeline(monthly_for_view, idx, fallback_monthly)
    else:
        monthly_for_view = _sales_plan_editor_monthly_view(
            canonical_plan,
            selected_frequency,
            idx,
            selected_propagation_mode,
        )
    canonical_plan = _compress_monthly_sales_plan(monthly_for_view, selected_frequency)
    st.session_state[canonical_key] = canonical_plan.copy()

    if selected_propagation_mode == "manual":
        st.caption("Manual mode expects period-by-period sales entries across the full production horizon.")
    elif selected_propagation_mode == "repeat_first_year":
        st.caption("Enter only the first production year below. The model will repeat that demand pattern across the remaining years.")
    else:
        st.caption("Enter only the first production year below. The model will propagate that pattern across later years using the annual growth rate above.")

    view_df = _sales_plan_editor_view(canonical_plan, selected_frequency)
    if not st.session_state.get("assump_edit_sales_plan", False):
        st.session_state["assump_data_sales_plan"] = view_df.copy()
        st.session_state["assump_saved_sales_plan"] = view_df.copy()
        st.session_state["assump_work_sales_plan"] = view_df.copy()
    edited_view = _assumption_editor("Sales plan", "sales_plan", view_df)
    edited_base = _sales_plan_editor_view_to_base(edited_view, selected_frequency)
    st.session_state[canonical_key] = edited_base.copy()
    return edited_base


def _sync_model_timeline_state(cfg: ModelConfig, base_inputs: ModelInputs) -> None:
    timeline_sig = (str(cfg.start_date), int(cfg.months))
    previous_sig = st.session_state.get("assump_timeline_signature")
    schedule_defaults = {
        "direct_labor": base_inputs.direct_labor_schedule,
        "indirect_labor": base_inputs.indirect_labor_schedule,
        "inventory": base_inputs.inventory_schedule,
        "receivables": base_inputs.receivables_schedule,
        "payables": base_inputs.payables_schedule,
        "sku_operations": base_inputs.sku_operations,
        "brewhouse": base_inputs.brewhouse_schedule,
        "cellar": base_inputs.cellar_schedule,
        "packaging": base_inputs.packaging_schedule,
    }
    for key, default_df in schedule_defaults.items():
        if isinstance(default_df, pd.DataFrame):
            _sync_assumption_table_state(key, default_df)
    if previous_sig == timeline_sig:
        return

    freq = str(st.session_state.get("assump_sales_plan_frequency", base_inputs.sales_plan_frequency))
    if freq not in {"monthly", "quarterly", "yearly"}:
        freq = "monthly"
    propagation_mode = str(
        st.session_state.get(
            "assump_sales_plan_propagation_mode",
            getattr(base_inputs, "sales_plan_propagation_mode", "manual"),
        )
    )
    idx = pd.date_range(cfg.start_date, periods=cfg.months, freq="MS")
    template_idx = _sales_plan_template_idx(idx)

    canonical_key = "assump_data_sales_plan_base"
    current_sales = st.session_state.get(canonical_key, base_inputs.sales_plan)
    current_sales = _non_empty_rows(current_sales).copy() if isinstance(current_sales, pd.DataFrame) else base_inputs.sales_plan.copy()

    monthly_current = _expand_sales_plan_to_monthly(current_sales, freq)
    fallback_monthly = _expand_sales_plan_to_monthly(base_inputs.sales_plan, base_inputs.sales_plan_frequency)
    fallback_monthly = propagate_sales_plan_monthly(
        fallback_monthly,
        idx,
        mode=getattr(base_inputs, "sales_plan_propagation_mode", "manual"),
        annual_growth_pct=float(getattr(base_inputs, "sales_plan_propagation_growth_annual", 0.0)),
    )
    if propagation_mode == "manual":
        monthly_current = _fill_monthly_sales_plan_to_timeline(monthly_current, idx, fallback_monthly)
    else:
        monthly_current = _rebase_monthly_sales_plan_to_idx(monthly_current, template_idx)
        monthly_current = _fill_monthly_sales_plan_to_timeline(
            monthly_current,
            template_idx,
            fallback_monthly[fallback_monthly["date"].isin(template_idx)].copy(),
        )

    canonical_sales = _compress_monthly_sales_plan(monthly_current, freq)
    st.session_state[canonical_key] = canonical_sales.copy()
    st.session_state["assump_data_sales_plan"] = _sales_plan_editor_view(canonical_sales, freq)
    st.session_state["assump_saved_sales_plan"] = st.session_state["assump_data_sales_plan"].copy()
    st.session_state["assump_work_sales_plan"] = st.session_state["assump_data_sales_plan"].copy()
    st.session_state["assump_edit_sales_plan"] = False
    st.session_state["assump_timeline_signature"] = timeline_sig


def _valuation_section(result) -> None:
    st.subheader("Valuation summary")
    valuation_df = pd.DataFrame(result.valuation, index=["value"]).T
    valuation_df.index.name = "metric"
    st.dataframe(valuation_df.style.format("{:,.2f}"))


def _statement_section(result) -> None:
    st.subheader("Statements")
    annual = result.annual.copy()
    inventory_stage_cols = [
        c for c in annual.columns
        if c.startswith("inventory_") and c not in {"inventory", "inventory_reserve"}
    ]

    st.markdown("**Annual Financial Performance Statement (Income Statement)**")
    perf_cols = [
        "gross_revenue",
        "trade_spend",
        "returns_allowances",
        "net_revenue",
        "total_revenue",
        "other_income",
        "direct_material_costs",
        "direct_labor_cost",
        "inventory_writeoff",
        "payable_discount_benefit",
        "direct_costs",
        "gross_profit",
        "indirect_labor_cost",
        "bad_debt_expense",
        "opex",
        "ebitda",
        "depreciation",
        "ebit",
        "interest_expense",
        "pre_tax_income",
        "taxes",
        "net_income",
    ]
    perf_existing = [c for c in perf_cols if c in annual.columns]
    if perf_existing:
        st.dataframe(annual[perf_existing])
    else:
        st.info("No annual income-statement columns were found in the current result payload.")

    st.markdown("**Annual Financial Position (Balance Sheet view)**")
    position_cols = [
        "cash",
        "receivables",
        "inventory",
        *inventory_stage_cols,
        "inventory_reserve",
        "other_current_assets",
        "current_assets",
        "net_fixed_assets",
        "total_assets",
        "payables",
        "other_current_liabilities",
        "current_liabilities",
        "debt_ending_balance",
        "revolver_ending_balance",
        "total_debt_ending_balance",
        "total_liabilities",
        "contributed_capital",
        "retained_earnings",
        "equity",
        "balance_sheet_gap",
    ]
    pos_existing = [c for c in position_cols if c in annual.columns]
    if pos_existing:
        st.dataframe(annual[pos_existing])
    else:
        st.info("No annual balance-sheet columns were found in the current result payload.")

    st.markdown("**Annual Cash Flow Statement**")
    cf_cols = [
        "cash_flow_from_operations",
        "change_in_nwc",
        "capex",
        "cash_flow_from_investing",
        "debt_draw",
        "debt_principal_payment",
        "revolver_draw",
        "revolver_principal_payment",
        "equity_injection",
        "dividends",
        "cash_flow_from_financing",
        "net_change_in_cash",
        "debt_service",
        "fcff",
    ]
    cf_existing = [c for c in cf_cols if c in annual.columns]
    if cf_existing:
        st.dataframe(annual[cf_existing])
    else:
        st.info("No annual cash-flow columns were found in the current result payload.")

    st.markdown("**Latest 12 months (monthly)**")
    st.dataframe(
        result.monthly.tail(12)[
            [
                c
                for c in [
                    "gross_revenue",
                    "net_revenue",
                    "total_revenue",
                    "direct_costs",
                    "gross_profit",
                    "opex",
                    "ebitda",
                    "net_income",
                    "cash",
                    "total_debt_ending_balance",
                ]
                if c in result.monthly.columns
            ]
        ]
    )

    with st.expander("Full monthly statements (all columns)"):
        st.dataframe(result.monthly)


def _charts_section(result) -> None:
    st.subheader("Graphs & plots")
    monthly = result.monthly.copy()
    inventory_stage_cols = [
        c for c in monthly.columns
        if c.startswith("inventory_") and c not in {"inventory", "inventory_reserve"}
    ]

    def _plot_if_available(label: str, cols: list[str]) -> None:
        existing = [c for c in cols if c in monthly.columns]
        missing = sorted(set(cols) - set(existing))
        if not existing:
            st.warning(f"Skipped '{label}' chart because columns are missing: {', '.join(cols)}")
            return
        if missing:
            st.info(
                f"Showing '{label}' with available series only; missing columns: {', '.join(missing)}"
            )
        st.line_chart(monthly[existing])

    st.markdown("**Revenue, EBITDA, and Net Income**")
    _plot_if_available(
        "Revenue, EBITDA, and Net Income",
        ["total_revenue", "ebitda", "net_income"],
    )

    st.markdown("**Cash vs. Debt Ending Balance**")
    _plot_if_available("Cash vs. Debt Ending Balance", ["cash", "total_debt_ending_balance"])

    st.markdown("**Operating Cash Flow vs. FCFF**")
    _plot_if_available(
        "Operating Cash Flow vs. FCFF",
        ["cash_flow_from_operations", "fcff"],
    )
    if inventory_stage_cols:
        st.markdown("**Inventory composition by stage**")
        _plot_if_available(
            "Inventory composition by stage",
            ["inventory", *inventory_stage_cols],
        )


def _download_section(result) -> None:
    buffer = _build_professional_excel_output(result)
    st.download_button(
        label="Download Excel output (Comprehensive Pack)",
        data=buffer,
        file_name="brewery_model_output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _schedules_section(result) -> None:
    st.subheader("Detailed schedules and tables")
    inventory_stage_cols = [
        c for c in result.monthly.columns
        if c.startswith("inventory_") and c not in {"inventory", "inventory_reserve"}
    ]

    st.markdown("**Debt schedules (facility-level)**")
    for name, df in result.debt_schedules.items():
        with st.expander(f"Debt schedule – {name}"):
            st.dataframe(df)

    st.markdown("**Pricing matrix (all SKUs x channels)**")
    with st.expander("Show price table"):
        st.dataframe(result.prices)

    st.markdown("**Working capital components**")
    wc_cols = [
        "receivables",
        "inventory",
        *inventory_stage_cols,
        "inventory_reserve",
        "other_current_assets",
        "payables",
        "other_current_liabilities",
        "net_working_capital",
        "change_in_nwc",
    ]
    st.dataframe(result.monthly[wc_cols])

    st.markdown("**CAPEX, depreciation, and net fixed assets**")
    st.dataframe(result.monthly[["capex", "depreciation", "net_fixed_assets"]])

    st.markdown("**Financing and equity flows**")
    st.dataframe(
        result.monthly[
            [
                "debt_draw",
                "debt_principal_payment",
                "revolver_draw",
                "revolver_principal_payment",
                "interest_expense",
                "debt_service",
                "equity_injection",
                "dividends",
                "cash",
                "total_debt_ending_balance",
                "dscr",
                "interest_coverage",
                "leverage_ratio",
            ]
        ]
    )

    inventory_detail = (result.supporting_schedules or {}).get("inventory_detail")
    if isinstance(inventory_detail, pd.DataFrame) and not inventory_detail.empty:
        st.markdown("**Inventory detail schedule**")
        detail_stage_cols = [
            c for c in inventory_detail.columns
            if c.startswith("inventory_") and c not in {"inventory", "inventory_reserve"}
        ]
        detail_cols = detail_stage_cols + [c for c in ["inventory"] if c in inventory_detail.columns]
        st.dataframe(inventory_detail[detail_cols] if detail_cols else inventory_detail)

    if getattr(result, "supporting_schedules", None):
        st.markdown("**Labor and working-capital detail schedules**")
        for label, df in result.supporting_schedules.items():
            if label == "inventory_detail":
                continue
            with st.expander(label.replace("_", " ").title()):
                st.dataframe(df)


def _key_analytics_section(result, inputs: ModelInputs) -> None:
    st.subheader("Key Analytics")
    annual = result.annual.copy()
    monthly = result.monthly.copy()

    st.markdown("### Editable analytics controls")
    var_data_key = "ka_var_data"
    var_saved_key = "ka_var_saved"
    var_work_key = "ka_var_work"
    var_edit_key = "ka_var_edit"
    var_select_key = "ka_var_select"
    if var_data_key not in st.session_state:
        st.session_state[var_data_key] = pd.DataFrame(
            [
                {"variable": "price_uplift_pct", "value": 0.0, "apply_to": "revenue"},
                {"variable": "volume_uplift_pct", "value": 0.0, "apply_to": "revenue"},
                {"variable": "cost_uplift_pct", "value": 0.0, "apply_to": "direct_costs"},
                {"variable": "risk_sigma_pct", "value": 20.0, "apply_to": "monte_carlo"},
            ]
        )
    if var_saved_key not in st.session_state:
        st.session_state[var_saved_key] = st.session_state[var_data_key].copy()
    if var_work_key not in st.session_state:
        st.session_state[var_work_key] = st.session_state[var_saved_key].copy()
    if var_edit_key not in st.session_state:
        st.session_state[var_edit_key] = False

    c1, c2, c3, c4 = st.columns([1, 1, 1, 2])
    if c1.button("Edit", key="ka_edit_button"):
        st.session_state[var_edit_key] = not st.session_state[var_edit_key]
        if st.session_state[var_edit_key]:
            st.session_state[var_work_key] = st.session_state[var_saved_key].copy()
    if c2.button("Add Variable", key="ka_add_variable") and st.session_state[var_edit_key]:
        st.session_state[var_work_key] = pd.concat(
            [
                st.session_state[var_work_key],
                pd.DataFrame([{"variable": "new_variable", "value": 0.0, "apply_to": "revenue"}]),
            ],
            ignore_index=True,
        )
    if c3.button("Delete Variable", key="ka_delete_variable") and st.session_state[var_edit_key]:
        wdf = st.session_state[var_work_key]
        if not wdf.empty:
            sel = st.session_state.get(var_select_key, 0)
            if sel in wdf.index:
                st.session_state[var_work_key] = wdf.drop(index=sel).reset_index(drop=True)
    if c4.button("Reset Changes", key="ka_reset_changes"):
        st.session_state[var_work_key] = st.session_state[var_saved_key].copy()
        st.session_state[var_data_key] = st.session_state[var_saved_key].copy()

    if st.session_state[var_edit_key]:
        wdf = st.session_state[var_work_key]
        if not wdf.empty:
            selected_row = st.selectbox(
                "Variable selector",
                options=list(wdf.index),
                key=var_select_key,
                format_func=lambda row_idx: _row_selection_label(wdf, row_idx, prefix="Variable"),
            )
            st.markdown("**Editable variable fields**")
            var_name = st.text_input(
                "variable",
                value="" if pd.isna(wdf.loc[selected_row, "variable"]) else str(wdf.loc[selected_row, "variable"]),
                key=f"ka_field_name_{selected_row}",
            )
            var_value = st.number_input(
                "value",
                value=0.0 if pd.isna(wdf.loc[selected_row, "value"]) else float(wdf.loc[selected_row, "value"]),
                key=f"ka_field_val_{selected_row}",
            )
            apply_to = st.text_input(
                "apply_to",
                value="" if pd.isna(wdf.loc[selected_row, "apply_to"]) else str(wdf.loc[selected_row, "apply_to"]),
                key=f"ka_field_apply_{selected_row}",
            )
            b1, b2, b3, b4 = st.columns(4)
            if b1.button("Apply Variable Changes", key="ka_apply_var"):
                st.session_state[var_work_key].loc[selected_row, "variable"] = var_name
                st.session_state[var_work_key].loc[selected_row, "value"] = var_value
                st.session_state[var_work_key].loc[selected_row, "apply_to"] = apply_to
            if b2.button("Save Changes", key="ka_save_var"):
                st.session_state[var_saved_key] = st.session_state[var_work_key].copy()
                st.session_state[var_data_key] = st.session_state[var_saved_key].copy()
            if b3.button("Discard Edits", key="ka_discard_var"):
                st.session_state[var_work_key] = st.session_state[var_saved_key].copy()
            if b4.button("Close Editor", key="ka_close_var"):
                st.session_state[var_edit_key] = False
                st.session_state[var_data_key] = st.session_state[var_saved_key].copy()
        else:
            st.info("No variables left. Use 'Add Variable' to create one.")
        st.dataframe(st.session_state[var_work_key], use_container_width=True)
    else:
        st.dataframe(st.session_state[var_data_key], use_container_width=True)

    var_df = st.session_state[var_data_key].copy()
    var_lookup = {
        str(r["variable"]): float(r["value"])
        for _, r in var_df.iterrows()
        if pd.notna(r.get("variable")) and pd.notna(r.get("value"))
    }

    # Sensitivity Analysis
    st.markdown("### Sensitivity Analysis")
    base_revenue_raw = float(annual["total_revenue"].iloc[0]) if "total_revenue" in annual.columns else 0.0
    base_ebitda_raw = float(annual["ebitda"].iloc[0]) if "ebitda" in annual.columns else 0.0
    price_uplift = var_lookup.get("price_uplift_pct", 0.0) / 100.0
    volume_uplift = var_lookup.get("volume_uplift_pct", 0.0) / 100.0
    cost_uplift = var_lookup.get("cost_uplift_pct", 0.0) / 100.0
    base_revenue = base_revenue_raw * (1 + price_uplift) * (1 + volume_uplift)
    base_ebitda = base_ebitda_raw
    sens_defaults = pd.DataFrame(
        [
            {"case": "Price -10%", "revenue_factor": 0.90, "ebitda_factor": 0.75},
            {"case": "Base", "revenue_factor": 1.00, "ebitda_factor": 1.00},
            {"case": "Price +10%", "revenue_factor": 1.10, "ebitda_factor": 1.25},
            {"case": "Volume -10%", "revenue_factor": 0.90, "ebitda_factor": 0.80},
            {"case": "Volume +10%", "revenue_factor": 1.10, "ebitda_factor": 1.20},
        ]
    )
    sens_input = _dynamic_table_editor("Sensitivity variables", "ka_sensitivity", sens_defaults, label="case")
    sens_df = sens_input.copy()
    sens_df["revenue"] = base_revenue * pd.to_numeric(sens_df["revenue_factor"], errors="coerce").fillna(1.0)
    sens_df["ebitda"] = base_ebitda * pd.to_numeric(sens_df["ebitda_factor"], errors="coerce").fillna(1.0)
    st.dataframe(sens_df)
    st.bar_chart(sens_df.set_index("case")[["revenue", "ebitda"]])

    # Monte Carlo simulation
    st.markdown("### Monte Carlo simulation")
    mc_defaults = pd.DataFrame([{"simulations": 500, "sigma_pct": max(var_lookup.get("risk_sigma_pct", 20.0), 0.0)}])
    mc_cfg = _dynamic_table_editor("Monte Carlo variables", "ka_mc", mc_defaults, label="config")
    sims = int(pd.to_numeric(mc_cfg.iloc[0].get("simulations", 500), errors="coerce")) if not mc_cfg.empty else 500
    sims = int(np.clip(sims, 100, 5000))
    sigma_pct = (float(pd.to_numeric(mc_cfg.iloc[0].get("sigma_pct", 20.0), errors="coerce")) / 100.0) if not mc_cfg.empty else 0.2
    sigma_pct = max(sigma_pct, 0.0)
    if base_ebitda > 0:
        draws = np.random.normal(loc=base_ebitda, scale=base_ebitda * sigma_pct, size=sims)
        mc_df = pd.DataFrame({"ebitda_sim": draws})
        st.line_chart(mc_df.sort_values("ebitda_sim").reset_index(drop=True))
        mc_mean = float(mc_df["ebitda_sim"].mean())
        mc_p05 = float(mc_df["ebitda_sim"].quantile(0.05))
        mc_p50 = float(mc_df["ebitda_sim"].quantile(0.50))
        mc_p95 = float(mc_df["ebitda_sim"].quantile(0.95))
        st.markdown(
            (
                f"Based on **{sims:,}** simulations, average EBITDA is **{mc_mean:,.2f}**. "
                f"The distribution ranges from **{mc_p05:,.2f}** at the 5th percentile "
                f"to **{mc_p95:,.2f}** at the 95th percentile, with a median of **{mc_p50:,.2f}**."
            )
        )

    # What-ifs analysis
    st.markdown("### What ifs analysis")
    what_if_defaults = pd.DataFrame([{"price_change_pct": 0.0, "volume_change_pct": 0.0, "cost_change_pct": 0.0}])
    what_if_cfg = _dynamic_table_editor("What-if variables", "ka_whatif", what_if_defaults, label="set")
    what_if_price = float(pd.to_numeric(what_if_cfg.iloc[0].get("price_change_pct", 0.0), errors="coerce")) if not what_if_cfg.empty else 0.0
    what_if_volume = float(pd.to_numeric(what_if_cfg.iloc[0].get("volume_change_pct", 0.0), errors="coerce")) if not what_if_cfg.empty else 0.0
    what_if_cost = float(pd.to_numeric(what_if_cfg.iloc[0].get("cost_change_pct", 0.0), errors="coerce")) if not what_if_cfg.empty else 0.0
    what_if_revenue = base_revenue * (1 + what_if_price / 100.0) * (1 + what_if_volume / 100.0)
    base_direct = (float(annual["direct_costs"].iloc[0]) if "direct_costs" in annual.columns else 0.0) * (1 + cost_uplift)
    what_if_direct = base_direct * (1 + what_if_cost / 100.0)
    what_if_ebitda = what_if_revenue - what_if_direct - float(annual["opex"].iloc[0])
    m1, m2, m3 = st.columns(3)
    m1.metric("What-if Revenue", f"{what_if_revenue:,.2f}")
    m2.metric("What-if Direct Costs", f"{what_if_direct:,.2f}")
    m3.metric("What-if EBITDA", f"{what_if_ebitda:,.2f}")
    st.bar_chart(
        pd.DataFrame(
            {
                "base": [base_revenue, base_direct, base_ebitda],
                "what_if": [what_if_revenue, what_if_direct, what_if_ebitda],
            },
            index=["revenue", "direct_costs", "ebitda"],
        )
    )

    # Break-even analysis by product
    st.markdown("### Break-even analysis by product")
    prices = result.prices
    weight_defaults = pd.DataFrame(
        [{"channel": "Wholesale", "weight": 0.40}, {"channel": "Retail", "weight": 0.30}, {"channel": "E-Commerce", "weight": 0.15}, {"channel": "On-Premise", "weight": 0.10}, {"channel": "Export", "weight": 0.05}]
    )
    weight_df = _dynamic_table_editor("Break-even channel weights", "ka_breakeven_weights", weight_defaults, label="weight row")
    channel_weights = {
        str(r["channel"]): float(r["weight"])
        for _, r in weight_df.iterrows()
        if pd.notna(r.get("channel")) and pd.notna(r.get("weight"))
    }
    be_rows = []
    fixed_pool = float(annual["opex"].iloc[0]) if "opex" in annual.columns else 0.0
    for _, sku in inputs.skus.iterrows():
        sid = sku["sku_id"]
        weighted_price = 0.0
        for ch, w in channel_weights.items():
            if (sid, ch) in prices.columns:
                weighted_price += float(prices[(sid, ch)].iloc[0]) * w
        unit_margin = max(weighted_price - float(sku["direct_cost_per_unit"]), 1e-6)
        be_rows.append(
            {"sku_id": sid, "name": sku["name"], "unit_margin": unit_margin, "break_even_units": fixed_pool / unit_margin}
        )
    st.dataframe(pd.DataFrame(be_rows))
    be_df = pd.DataFrame(be_rows).set_index("name")
    st.bar_chart(be_df[["break_even_units"]])

    # Scenario planning
    st.markdown("### Scenario planning")
    scen_defaults = pd.DataFrame(
        [
            {"scenario": "Downside", "revenue_multiplier": 0.85, "ebitda_multiplier": 0.70},
            {"scenario": "Base", "revenue_multiplier": 1.00, "ebitda_multiplier": 1.00},
            {"scenario": "Upside", "revenue_multiplier": 1.15, "ebitda_multiplier": 1.30},
        ]
    )
    scen_df = _dynamic_table_editor("Scenario planning variables", "ka_scenarios", scen_defaults, label="scenario")
    scen_df["revenue"] = base_revenue * pd.to_numeric(scen_df["revenue_multiplier"], errors="coerce").fillna(1.0)
    scen_df["ebitda"] = base_ebitda * pd.to_numeric(scen_df["ebitda_multiplier"], errors="coerce").fillna(1.0)
    st.dataframe(scen_df)
    st.line_chart(scen_df.set_index("scenario")[["revenue", "ebitda"]])

    # Goal seek
    st.markdown("### Goal seek")
    goal_defaults = pd.DataFrame([{"target_ebitda": max(base_ebitda, 1.0)}])
    goal_df = _dynamic_table_editor("Goal-seek variables", "ka_goalseek", goal_defaults, label="target row")
    target_ebitda = float(pd.to_numeric(goal_df.iloc[0].get("target_ebitda", max(base_ebitda, 1.0)), errors="coerce")) if not goal_df.empty else max(base_ebitda, 1.0)
    margin_rate = max((base_ebitda / base_revenue) if base_revenue else 0.0, 1e-6)
    required_revenue = target_ebitda / margin_rate
    implied_uplift = (required_revenue / base_revenue - 1) * 100 if base_revenue else 0.0
    st.markdown(
        (
            f"To reach a target EBITDA of **{target_ebitda:,.2f}**, the model implies required revenue "
            f"of **{required_revenue:,.2f}**, which is an uplift of **{implied_uplift:,.2f}%** "
            "versus the current base revenue."
        )
    )
    st.bar_chart(
        pd.DataFrame(
            {"value": [base_revenue, required_revenue]},
            index=["current_revenue", "required_revenue"],
        )
    )

    # Debt service coverage ratio
    st.markdown("### Debt service coverage ratio")
    dscr_cfg = _dynamic_table_editor(
        "DSCR variables",
        "ka_dscr",
        pd.DataFrame([{"min_dscr_threshold": 1.20}]),
        label="threshold row",
    )
    min_dscr = float(pd.to_numeric(dscr_cfg.iloc[0].get("min_dscr_threshold", 1.20), errors="coerce")) if not dscr_cfg.empty else 1.2
    if "dscr" in monthly.columns:
        dscr_df = pd.DataFrame({"DSCR": monthly["dscr"], "min_threshold": min_dscr}, index=monthly.index)
        st.line_chart(dscr_df)
    elif {"debt_service", "ebitda"}.issubset(monthly.columns):
        dscr = np.where(monthly["debt_service"] > 0, monthly["ebitda"] / monthly["debt_service"], np.nan)
        dscr_df = pd.DataFrame({"DSCR": dscr, "min_threshold": min_dscr}, index=monthly.index)
        st.line_chart(dscr_df)
    else:
        st.info("DSCR uses the model's debt service and EBITDA outputs when available.")

    # Predictive analytics
    st.markdown("### Predictive analytics")
    pred_cfg = _dynamic_table_editor(
        "Predictive variables",
        "ka_predictive",
        pd.DataFrame([{"forecast_steps": 1}]),
        label="config row",
    )
    forecast_steps = int(pd.to_numeric(pred_cfg.iloc[0].get("forecast_steps", 1), errors="coerce")) if not pred_cfg.empty else 1
    forecast_steps = max(forecast_steps, 1)
    if "total_revenue" in annual.columns and len(annual) >= 3:
        y = annual["total_revenue"].values.astype(float)
        x = np.arange(len(y))
        coeff = np.polyfit(x, y, 1)
        pred = coeff[0] * (x + forecast_steps) + coeff[1]
        pred_df = pd.DataFrame({"actual_revenue": y, f"prediction_step_{forecast_steps}": pred}, index=annual.index)
        st.dataframe(pred_df)
        st.line_chart(pred_df)
    else:
        st.info("Need at least 3 annual revenue points for trend-based prediction.")

    # Return diagnostics
    st.markdown("### Return diagnostics")
    ret_cfg = _dynamic_table_editor(
        "Return diagnostics variables",
        "ka_returns",
        pd.DataFrame([{"target_irr_pct": 20.0}]),
        label="target row",
    )
    target_irr = float(pd.to_numeric(ret_cfg.iloc[0].get("target_irr_pct", 20.0), errors="coerce")) / 100.0 if not ret_cfg.empty else 0.2
    val = result.valuation
    keys = ["enterprise_value", "equity_value", "investor_irr_annual", "investor_moic"]
    st.markdown(
        (
            f"Valuation diagnostics show enterprise value of **{float(val.get('enterprise_value', 0.0)):,.2f}** "
            f"and equity value of **{float(val.get('equity_value', 0.0)):,.2f}**. "
            f"Investor annual IRR is **{float(val.get('investor_irr_annual', 0.0)):.2%}** "
            f"against a target IRR of **{target_irr:.2%}**, and investor MOIC is "
            f"**{float(val.get('investor_moic', 0.0)):.2f}x**."
        )
    )
    diag_vals = {k: val.get(k, np.nan) for k in keys if isinstance(val.get(k, np.nan), (int, float))}
    if diag_vals:
        st.bar_chart(pd.DataFrame({"value": diag_vals}))

    # Coverage & resilience
    st.markdown("### Coverage & resilience")
    cov_cfg = _dynamic_table_editor(
        "Coverage & resilience variables",
        "ka_coverage",
        pd.DataFrame([{"min_cash_to_debt": 1.0, "min_months_opex": 3.0}]),
        label="threshold row",
    )
    min_cash_to_debt = float(pd.to_numeric(cov_cfg.iloc[0].get("min_cash_to_debt", 1.0), errors="coerce")) if not cov_cfg.empty else 1.0
    min_months_opex = float(pd.to_numeric(cov_cfg.iloc[0].get("min_months_opex", 3.0), errors="coerce")) if not cov_cfg.empty else 3.0
    cov = {}
    if {"cash", "total_debt_ending_balance"}.issubset(monthly.columns):
        cov["cash_to_debt_ratio_last_month"] = float(
            monthly["cash"].iloc[-1] / max(monthly["total_debt_ending_balance"].iloc[-1], 1e-6)
        )
    if {"cash", "opex"}.issubset(monthly.columns):
        cov["months_of_opex_covered_last_month"] = float(
            monthly["cash"].iloc[-1] / max(monthly["opex"].iloc[-1], 1e-6)
        )
    if cov:
        cov_parts = []
        if "cash_to_debt_ratio_last_month" in cov:
            cov_parts.append(
                f"cash-to-debt ratio is **{cov['cash_to_debt_ratio_last_month']:.2f}x** "
                f"(threshold **{min_cash_to_debt:.2f}x**)"
            )
        if "months_of_opex_covered_last_month" in cov:
            cov_parts.append(
                f"cash coverage equals **{cov['months_of_opex_covered_last_month']:.2f}** months of opex "
                f"(threshold **{min_months_opex:.2f}** months)"
            )
        st.markdown("Latest resilience view: " + " and ".join(cov_parts) + ".")
        st.markdown(f"Reference target IRR for this diagnostic set is **{target_irr:.2%}**.")
    else:
        st.info("Coverage metrics require cash, debt, and opex monthly columns.")
    if cov:
        st.bar_chart(pd.DataFrame({"value": cov}))


def _internal_findings_snapshot(result, inputs: ModelInputs, cfg: ModelConfig, div: DividendPolicy) -> dict:
    annual = result.annual.copy()
    latest_annual = annual.iloc[-1].to_dict() if len(annual) else {}
    latest_monthly = result.monthly.iloc[-1].to_dict() if len(result.monthly) else {}
    ratios = {}
    if latest_annual.get("ebitda") not in [None, 0] and latest_monthly.get("total_debt_ending_balance") is not None:
        ratios["debt_to_ebitda"] = float(latest_monthly.get("total_debt_ending_balance", 0.0)) / max(float(latest_annual.get("ebitda", 0.0)), 1e-9)
    if latest_monthly.get("cash") is not None and latest_monthly.get("total_debt_ending_balance") not in [None, 0]:
        ratios["cash_to_debt"] = float(latest_monthly.get("cash", 0.0)) / max(float(latest_monthly.get("total_debt_ending_balance", 0.0)), 1e-9)
    if latest_annual.get("dscr") is not None:
        ratios["dscr"] = float(latest_annual.get("dscr", 0.0))
    if latest_annual.get("interest_coverage") is not None:
        ratios["interest_coverage"] = float(latest_annual.get("interest_coverage", 0.0))
    if latest_annual.get("leverage_ratio") is not None:
        ratios["leverage_ratio"] = float(latest_annual.get("leverage_ratio", 0.0))
    if latest_monthly.get("current_assets") is not None and latest_monthly.get("current_liabilities") not in [None, 0]:
        ratios["current_ratio"] = float(latest_monthly.get("current_assets", 0.0)) / max(float(latest_monthly.get("current_liabilities", 0.0)), 1e-9)
    return {
        "model_governance": {
            "start_date": cfg.start_date,
            "months": cfg.months,
            "wacc": cfg.wacc_annual,
            "tax_rate": cfg.tax_rate,
            "dividend_policy": {
                "enabled": div.enabled,
                "model": div.model,
                "start_month": div.start_month,
                "minimum_cash_position": div.minimum_cash_position,
                "payout_ratio": div.payout_ratio,
            },
        },
        "production_and_revenues": {
            "sku_count": int(len(inputs.skus)),
            "channel_count": int(len(inputs.channels)),
            "latest_total_revenue": latest_annual.get("total_revenue"),
            "latest_ebitda": latest_annual.get("ebitda"),
        },
        "financial_statements": {
            "latest_net_income": latest_annual.get("net_income"),
            "latest_cash": latest_monthly.get("cash"),
            "latest_debt_balance": latest_monthly.get("total_debt_ending_balance"),
        },
        "advanced_analytics": {
            "valuation_summary": result.valuation,
        },
        "latest_annual": latest_annual,
        "latest_monthly": latest_monthly,
        "ratios": ratios,
    }


def _best_practice_sources_by_topic(question: str) -> list[dict]:
    q = question.lower()
    sources = []
    if any(k in q for k in ["governance", "control", "risk", "model governance"]):
        sources += [
            {"name": "SR 11-7 Supervisory Guidance", "url": "https://www.federalreserve.gov/supervisionreg/srletters/sr1107.htm"},
            {"name": "OCC 2011-12 Model Risk Management", "url": "https://occ.treas.gov/news-issuances/bulletins/2011/bulletin-2011-12.html"},
        ]
    if any(k in q for k in ["financial statement", "ifrs", "gaap", "reporting"]):
        sources += [
            {"name": "IFRS Foundation", "url": "https://www.ifrs.org/"},
            {"name": "FASB US GAAP", "url": "https://www.fasb.org/"},
        ]
    if any(k in q for k in ["dscr", "debt service", "coverage"]):
        sources += [
            {"name": "CFI DSCR overview", "url": "https://corporatefinanceinstitute.com/resources/commercial-lending/debt-service-coverage-ratio/"},
            {"name": "SBA Loan Basics", "url": "https://www.sba.gov/funding-programs/loans"},
        ]
    if any(k in q for k in ["monte carlo", "simulation", "sensitivity", "scenario"]):
        sources += [
            {"name": "NIST Engineering Statistics Handbook", "url": "https://www.itl.nist.gov/div898/handbook/"},
            {"name": "OECD Corporate Governance (risk context)", "url": "https://www.oecd.org/corporate/"},
        ]
    if not sources:
        sources += [
            {"name": "OECD Corporate Governance", "url": "https://www.oecd.org/corporate/"},
            {"name": "IMF Data and Financial Soundness", "url": "https://www.imf.org/en/Data"},
        ]
    # de-duplicate by url
    seen = set()
    deduped = []
    for s in sources:
        if s["url"] not in seen:
            deduped.append(s)
            seen.add(s["url"])
    return deduped[:6]


def _live_web_headlines(query: str) -> list[dict]:
    # lightweight optional web lookup via DuckDuckGo instant answer API
    url = "https://api.duckduckgo.com/?" + urllib.parse.urlencode({"q": query, "format": "json", "no_redirect": 1, "no_html": 1})
    with urllib.request.urlopen(url, timeout=8) as resp:
        payload = json.loads(resp.read().decode("utf-8"))
    out = []
    if payload.get("AbstractURL"):
        out.append({"title": payload.get("Heading") or "DuckDuckGo Abstract", "url": payload.get("AbstractURL")})
    for item in payload.get("RelatedTopics", []):
        if isinstance(item, dict) and item.get("FirstURL"):
            out.append({"title": item.get("Text", "Related Topic"), "url": item.get("FirstURL")})
        if len(out) >= 5:
            break
    return out


def _ai_decision_making_page(result, inputs: ModelInputs, cfg: ModelConfig, div: DividendPolicy) -> None:
    st.subheader("AI Decision Making")
    st.caption(
        "Context-aware model Q&A: answers assumptions, calculations, outputs, schedules, logic, dependencies, and scenario follow-ups."
    )
    if "decision_memory" not in st.session_state:
        st.session_state["decision_memory"] = DecisionSessionMemory()
    memory: DecisionSessionMemory = st.session_state["decision_memory"]

    c1, c2 = st.columns([1, 1])
    with c1:
        if st.button("Clear Q&A context"):
            st.session_state["decision_memory"] = DecisionSessionMemory()
            memory = st.session_state["decision_memory"]
            st.success("Conversation context was cleared.")
    with c2:
        st.caption(f"Questions in context: {len(memory.turns)}")

    question = st.text_area("Ask a decision question", placeholder="Example: Is our dividend policy and DSCR resilient under downside scenarios?")
    run = st.button("Analyze question")
    if not run or not question.strip():
        if memory.turns:
            st.markdown("### Previous Q&A context")
            for i, turn in enumerate(memory.turns[-5:], start=max(1, len(memory.turns) - 4)):
                st.markdown(f"**Q{i}:** {turn.question}")
                st.markdown(f"**A{i}:** {turn.answer}")
        return

    internal = _internal_findings_snapshot(result, inputs, cfg, div)
    qtype = classify_question(question, memory)
    benchmark_query_map = {
        QuestionType.VALUATION: "craft beverage EV EBITDA multiples lower middle market",
        QuestionType.PROFITABILITY: "beverage industry EBITDA margin benchmark",
        QuestionType.LIQUIDITY: "small business current ratio benchmark debt liquidity",
        QuestionType.LEVERAGE: "Debt EBITDA covenant range DSCR benchmark",
        QuestionType.GROWTH: "beverage market revenue growth benchmark",
        QuestionType.PRICING: "beverage pricing power gross margin benchmark",
        QuestionType.EFFICIENCY: "beverage COGS opex benchmark",
        QuestionType.RISK: "brewery downside scenario debt coverage benchmark",
    }
    benchmark_query = benchmark_query_map.get(qtype, question)
    curated_sources = _best_practice_sources_by_topic(f"{question} {benchmark_query}")
    live_sources = []
    live_error = None
    try:
        live_sources = _live_web_headlines(benchmark_query)
    except Exception as e:
        live_error = str(e)

    st.markdown("### Decision answer")
    turn = build_structured_answer(question, internal, memory, web_sources=live_sources)
    st.markdown(turn.answer)

    st.markdown("### Internal model findings")
    mg = internal.get("model_governance", {})
    dp = mg.get("dividend_policy", {})
    pr = internal.get("production_and_revenues", {})
    fs = internal.get("financial_statements", {})
    va = internal.get("advanced_analytics", {}).get("valuation_summary", {})
    st.write(
        f"The internal model is configured with a start date of **{mg.get('start_date')}**, a horizon of "
        f"**{mg.get('months')} months**, **WACC {mg.get('wacc'):.2%}**, and **tax rate {mg.get('tax_rate'):.2%}**. "
        f"Dividend policy is **{dp.get('model')}** (enabled: {dp.get('enabled')}), starting at month "
        f"**{dp.get('start_month')}**, with minimum cash of **{dp.get('minimum_cash_position'):,.0f}** and "
        f"payout ratio **{dp.get('payout_ratio'):.2%}**."
    )
    st.write(
        f"Operations currently span **{pr.get('sku_count')} SKUs** across **{pr.get('channel_count')} channels**. "
        f"Latest annual revenue is **{pr.get('latest_total_revenue'):,.2f}** and latest EBITDA is "
        f"**{pr.get('latest_ebitda'):,.2f}**."
    )
    st.write(
        f"Financial statement outputs show latest net income of **{fs.get('latest_net_income'):,.2f}**, "
        f"cash of **{fs.get('latest_cash'):,.2f}**, and ending debt of **{fs.get('latest_debt_balance'):,.2f}**."
    )
    if va:
        st.write(
            f"Valuation analytics indicate terminal value **{va.get('terminal_value', 0):,.2f}**, enterprise value "
            f"**{va.get('enterprise_value_dcf', 0):,.2f}**, equity value at exit **{va.get('equity_value_exit', 0):,.2f}**, "
            f"annual IRR **{va.get('equity_irr_annual', 0):.2%}**, and MOIC **{va.get('equity_moic', 0):,.2f}x**."
        )

    st.markdown("### Web-based benchmark references")
    st.markdown("**Curated authoritative references**")
    for s in curated_sources:
        st.markdown(f"- [{s['name']}]({s['url']})")
    if live_sources:
        st.markdown("**Live web matches (supplementary)**")
        for s in live_sources:
            st.markdown(f"- [{s['title']}]({s['url']})")
    elif live_error:
        st.info(f"Live web lookup was limited in this run: {live_error}")

    st.markdown("### Key gaps, risks, or strengths")
    st.markdown(
        "- **Strength**: Internal model has integrated schedules and diagnostics (valuation, DSCR, scenarios).\n"
        "- **Gap**: Governance and benchmark checks should be codified as explicit thresholds and alerts.\n"
        "- **Risk**: Policy assumptions (e.g., dividends/cash floors) may be optimistic without stress-case guardrails."
    )

    st.markdown("### Professional recommendation")
    st.markdown(
        "1. Formalize benchmark hurdles (DSCR, leverage, runway) as hard controls.\n"
        "2. Tie dividend activation to downside-case resilience, not base case only.\n"
        "3. Add periodic benchmark refresh against external references shown below."
    )

    st.markdown("### Conversation continuity")
    st.markdown(
        f"- Question type detected: **{turn.question_type.value}**\n"
        f"- Topic tags: **{', '.join(turn.topics)}**\n"
        f"- Context window size: **{len(memory.turns)} turn(s)**"
    )

    st.markdown("### Sources")
    for s in curated_sources:
        st.markdown(f"- {s['name']}: {s['url']}")
    for s in live_sources:
        st.markdown(f"- {s['title']}: {s['url']}")


def main() -> None:
    _inject_app_theme()
    _render_model_hero()

    run_controls = st.container()
    tab_assumptions, tab_results, tab_key_analytics, tab_ai_decision = st.tabs([
        "Core Assumptions",
        "Results",
        "Key Analytics",
        "AI Decision Making",
    ])

    with tab_assumptions:
        st.subheader("Core assumptions")
        c1, c2 = st.columns(2)
        start_year = c1.selectbox(
            "Start year",
            options=list(range(2025, 2036)),
            index=0,
        )
        end_year = c2.selectbox(
            "End year",
            options=list(range(int(start_year), 2041)),
            index=min(9, len(list(range(int(start_year), 2041))) - 1),
        )
        years = int(end_year) - int(start_year) + 1
        months = years * 12
        st.caption(f"Projection length: {years} years ({months} months)")

        c3, c4 = st.columns(2)
        wacc = c3.selectbox(
            "WACC (annual)",
            options=[round(x, 3) for x in np.arange(0.050, 0.205, 0.005)],
            index=14,
            format_func=lambda x: f"{x:.3f}",
        )
        exit_multiple = c4.selectbox(
            "Exit EV/EBITDA multiple",
            options=[round(x, 1) for x in np.arange(4.0, 12.5, 0.5)],
            index=8,
            format_func=lambda x: f"{x:.1f}x",
        )
        c5, c6 = st.columns(2)
        price_inflation = c5.selectbox(
            "Price inflation (annual)",
            options=[round(x, 4) for x in np.arange(0.0, 0.0501, 0.0025)],
            index=6,
            format_func=lambda x: f"{x:.2%}",
        )
        cost_inflation = c6.selectbox(
            "Cost inflation (annual)",
            options=[round(x, 4) for x in np.arange(0.0, 0.0501, 0.0025)],
            index=6,
            format_func=lambda x: f"{x:.2%}",
        )

        c7, c8 = st.columns(2)
        dividend_start = c7.selectbox(
            "Dividend start month",
            options=list(range(0, int(months))),
            index=min(60, int(months) - 1),
        )
        min_cash = c8.selectbox(
            "Minimum cash for sweep",
            options=[0.0, 250_000.0, 500_000.0, 1_000_000.0, 1_500_000.0, 2_000_000.0, 2_500_000.0, 3_000_000.0],
            index=4,
            format_func=lambda x: f"${x:,.0f}",
        )

    cfg = ModelConfig(
        start_date=f"{int(start_year)}-01-01",
        months=months,
        pricing_cost_basis_month=24,
        price_inflation_annual=price_inflation,
        cost_inflation_annual=cost_inflation,
        tax_rate=0.25,
        wacc_annual=wacc,
        exit_ev_ebitda_multiple=exit_multiple,
        initial_cash=0.0,
    )

    div = DividendPolicy(
        enabled=True,
        model="cash_sweep",
        start_month=int(dividend_start),
        minimum_cash_position=float(min_cash),
        payout_ratio=0.25,
    )

    config_table_state = st.session_state.get("assump_data_config")
    if isinstance(config_table_state, pd.DataFrame) and not config_table_state.empty:
        row = config_table_state.iloc[0]
        cfg = replace(
            cfg,
            price_inflation_annual=float(row.get("price_inflation_annual", cfg.price_inflation_annual)),
            cost_inflation_annual=float(row.get("cost_inflation_annual", cfg.cost_inflation_annual)),
            tax_rate=float(row.get("tax_rate", cfg.tax_rate)),
            wacc_annual=float(row.get("wacc_annual", cfg.wacc_annual)),
            exit_ev_ebitda_multiple=float(row.get("exit_ev_ebitda_multiple", cfg.exit_ev_ebitda_multiple)),
            revolver_limit=float(row.get("revolver_limit", cfg.revolver_limit)),
            revolver_interest_annual=float(row.get("revolver_interest_annual", cfg.revolver_interest_annual)),
            revolver_target_cash=float(row.get("revolver_target_cash", cfg.revolver_target_cash)),
            min_dscr=float(row.get("min_dscr", cfg.min_dscr)),
            min_interest_coverage=float(row.get("min_interest_coverage", cfg.min_interest_coverage)),
            max_leverage_ratio=float(row.get("max_leverage_ratio", cfg.max_leverage_ratio)),
            temporary_labor_premium_pct=float(
                row.get("temporary_labor_premium_pct", cfg.temporary_labor_premium_pct)
            ),
        )
    dividend_table_state = st.session_state.get("assump_data_dividend")
    if isinstance(dividend_table_state, pd.DataFrame) and not dividend_table_state.empty:
        row = dividend_table_state.iloc[0]
        div = replace(
            div,
            enabled=_coerce_bool(row.get("enabled", div.enabled)),
            model=str(row.get("model", div.model)),
            start_month=int(row.get("start_month", div.start_month)),
            minimum_cash_position=float(row.get("minimum_cash_position", div.minimum_cash_position)),
            payout_ratio=float(row.get("payout_ratio", div.payout_ratio)),
        )

    base_inputs = _build_sample_assumptions(cfg, div)
    _sync_model_timeline_state(cfg, base_inputs)
    inputs = _build_inputs_from_state(base_inputs)
    draft_signature = _build_draft_signature(cfg, div, inputs)

    current_bundle = _current_result_bundle()
    run_label = "Run Model" if current_bundle is None else "Recalculate"
    with run_controls:
        left_col, right_col = st.columns([3, 1])
        status_slot = left_col.empty()
        meta_slot = left_col.empty()
        run_clicked = right_col.button(
            run_label,
            key="brewery_run_model",
            type="primary",
            use_container_width=True,
        )

    if run_clicked:
        with st.spinner("Running microbrewery model..."):
            started = time.perf_counter()
            model = MicrobreweryFinancialModel(cfg, div, deepcopy(inputs))
            result = model.run()
            current_bundle = StoredModelRun(
                signature=draft_signature,
                cfg=cfg,
                div=div,
                inputs=model.inputs,
                result=result,
                computed_at=time.time(),
                duration_seconds=time.perf_counter() - started,
            )
        _activate_result_bundle(current_bundle)

    draft_is_dirty = _draft_is_dirty(draft_signature, current_bundle)
    if current_bundle is None:
        status_slot.info(
            "Draft inputs are ready. Press Run Model to generate the valuation, dashboards, analytics, and exports."
        )
        meta_slot.empty()
    elif draft_is_dirty:
        status_slot.warning(
            "Draft inputs changed. The results below still reflect the last completed run. Press Recalculate to refresh the model."
        )
        meta_slot.caption(
            f"Current results reflect the last completed run, which took {current_bundle.duration_seconds:.2f}s."
        )
    else:
        status_slot.success("Results are current with the latest draft inputs.")
        meta_slot.caption(
            f"Latest run completed in {current_bundle.duration_seconds:.2f}s."
        )

    with tab_results:
        if current_bundle is None:
            st.info(
                "Run Model from the draft controls above to generate the current valuation, statements, schedules, and downloads."
            )
        else:
            if draft_is_dirty:
                _render_stale_banner(
                    "These dashboards and downloads still reflect the last completed run. Press Recalculate when you want them refreshed for the latest draft."
                )
            _valuation_section(current_bundle.result)
            _statement_section(current_bundle.result)
            _charts_section(current_bundle.result)
            _schedules_section(current_bundle.result)
            _driver_based_opex_views_section(current_bundle.result)
            _download_section(current_bundle.result)

            st.caption(
                "The sample assumptions mirror the CLI example in "
                "`brewery_financial_model_all_in_one.py`. Adjust the draft inputs and press Recalculate when you want a new model run."
            )

    with tab_key_analytics:
        if current_bundle is None:
            st.info("Run Model first, then review the analytics based on the stored result set.")
        else:
            if draft_is_dirty:
                _render_stale_banner(
                    "These analytics still reflect the last completed run. Press Recalculate to refresh the base model before reviewing updated analytics."
                )
            _key_analytics_section(current_bundle.result, current_bundle.inputs)

    with tab_ai_decision:
        if current_bundle is None:
            st.info("Run Model first, then use AI Decision Making against the stored result set.")
        else:
            if draft_is_dirty:
                _render_stale_banner(
                    "AI decision-support answers still reflect the last completed run until you press Recalculate."
                )
            _ai_decision_making_page(
                current_bundle.result,
                current_bundle.inputs,
                current_bundle.cfg,
                current_bundle.div,
            )

    with tab_assumptions:
        st.divider()
        st.subheader("Detailed Assumption Tables")
        st.caption("Changes in these tables update the draft only. Save the edits here, then press Recalculate above to refresh the model outputs.")
        capex_df = pd.DataFrame(
            [
                {
                    "name": item.name,
                    "amount": item.amount,
                    "capex_month": item.capex_month,
                    "depreciation_years": item.depreciation_years,
                }
                for item in inputs.capex_items
            ]
        )
        debt_df = pd.DataFrame(
            [
                {
                    "name": fac.name,
                    "principal": fac.principal,
                    "annual_interest_rate": fac.annual_interest_rate,
                    "draw_month": fac.draw_month,
                    "grace_months": fac.grace_months,
                    "term_months": fac.term_months,
                    "repayment_type": fac.repayment_type,
                }
                for fac in inputs.debt_facilities
            ]
        )
        equity_df = pd.DataFrame(
            [{"month": k, "equity_injection": v} for k, v in sorted(inputs.equity_injections.items())]
        )
        config_df = pd.DataFrame(
            [
                {
                    "start_date": cfg.start_date,
                    "months": cfg.months,
                    "price_inflation_annual": cfg.price_inflation_annual,
                    "cost_inflation_annual": cfg.cost_inflation_annual,
                    "tax_rate": cfg.tax_rate,
                    "wacc_annual": cfg.wacc_annual,
                    "exit_ev_ebitda_multiple": cfg.exit_ev_ebitda_multiple,
                    "revolver_limit": cfg.revolver_limit,
                    "revolver_interest_annual": cfg.revolver_interest_annual,
                    "revolver_target_cash": cfg.revolver_target_cash,
                    "min_dscr": cfg.min_dscr,
                    "min_interest_coverage": cfg.min_interest_coverage,
                    "max_leverage_ratio": cfg.max_leverage_ratio,
                    "temporary_labor_premium_pct": cfg.temporary_labor_premium_pct,
                }
            ]
        )
        dividend_df = pd.DataFrame(
            [
                {
                    "enabled": div.enabled,
                    "model": div.model,
                    "start_month": div.start_month,
                    "minimum_cash_position": div.minimum_cash_position,
                    "payout_ratio": div.payout_ratio,
                }
            ]
        )
        cost_pool_df = pd.DataFrame(
            [
                {
                    "name": p.name,
                    "cost_type": p.cost_type,
                    "behavior": p.behavior,
                    "allocation_driver": p.allocation_driver,
                    "scope": p.scope,
                    "channel": p.channel,
                    "unit_variable_cost": p.unit_variable_cost,
                    "fixed_monthly_cost": p.fixed_monthly_cost,
                    "step_threshold": p.step_threshold,
                    "step_increment": p.step_increment,
                }
                for p in (inputs.cost_pools or [])
            ]
        )
        other_income_df = pd.DataFrame(
            [
                {
                    "other_income_name": item.other_income_name,
                    "amount": item.amount,
                    "active": "True" if item.active else "False",
                    "category": item.category,
                    "notes": item.notes,
                }
                for item in (inputs.other_income_items or [])
            ]
        )
        direct_labor_df = inputs.direct_labor_schedule.copy() if isinstance(inputs.direct_labor_schedule, pd.DataFrame) else pd.DataFrame()
        indirect_labor_df = inputs.indirect_labor_schedule.copy() if isinstance(inputs.indirect_labor_schedule, pd.DataFrame) else pd.DataFrame()
        sku_operations_df = inputs.sku_operations.copy() if isinstance(inputs.sku_operations, pd.DataFrame) else pd.DataFrame()
        brewhouse_df = inputs.brewhouse_schedule.copy() if isinstance(inputs.brewhouse_schedule, pd.DataFrame) else pd.DataFrame()
        cellar_df = inputs.cellar_schedule.copy() if isinstance(inputs.cellar_schedule, pd.DataFrame) else pd.DataFrame()
        packaging_df = inputs.packaging_schedule.copy() if isinstance(inputs.packaging_schedule, pd.DataFrame) else pd.DataFrame()
        inventory_df = inputs.inventory_schedule.copy() if isinstance(inputs.inventory_schedule, pd.DataFrame) else pd.DataFrame()
        receivables_df = inputs.receivables_schedule.copy() if isinstance(inputs.receivables_schedule, pd.DataFrame) else pd.DataFrame()
        payables_df = inputs.payables_schedule.copy() if isinstance(inputs.payables_schedule, pd.DataFrame) else pd.DataFrame()

        if not st.session_state.get("assump_edit_config", False):
            st.session_state["assump_data_config"] = config_df.copy()
            st.session_state["assump_saved_config"] = config_df.copy()
            st.session_state["assump_work_config"] = config_df.copy()

        _assumption_editor("Model config assumptions", "config", config_df)
        _assumption_editor("Dividend assumptions", "dividend", dividend_df)
        st.caption("`direct_cost_per_unit` is derived from direct cost pools. Use `direct_cost_per_unit_override` only for explicit manual overrides.")
        _assumption_editor("SKUs", "skus", inputs.skus)
        _assumption_editor("Channels", "channels", inputs.channels)
        _sales_plan_assumption_editor(
            inputs.sales_plan,
            inputs.sales_plan_frequency,
            cfg,
            getattr(inputs, "sales_plan_propagation_mode", "manual"),
            getattr(inputs, "sales_plan_propagation_growth_annual", 0.0),
        )
        st.caption("SKU operations translate demand into packaged output, brew input, finished-goods targets, and batch counts.")
        _assumption_editor("SKU operations", "sku_operations", sku_operations_df)
        st.caption("Brewhouse schedule sets brew-start capacity from batch size, brew days, uptime, and yearly expansion counts.")
        _assumption_editor("Brewhouse schedule", "brewhouse", brewhouse_df)
        st.caption("Cellar schedule constrains throughput through tank count, tank size, utilization, and downtime.")
        _assumption_editor("Cellar schedule", "cellar", cellar_df)
        st.caption("Packaging schedule constrains packaged liters through line rate, run days, uptime, and changeover loss.")
        _assumption_editor("Packaging schedule", "packaging", packaging_df)
        _assumption_editor("Cost pools", "cost_pools", cost_pool_df)
        st.caption("Direct labor schedules drive payroll, capacity, direct COGS, and temporary labor cost when capacity is short.")
        _assumption_editor("Direct labor schedule", "direct_labor", direct_labor_df)
        st.caption("Indirect labor schedules flow to OPEX and covenant-sensitive cash burn.")
        _assumption_editor("Indirect labor schedule", "indirect_labor", indirect_labor_df)
        st.caption("Receivables schedule drives gross-to-net revenue, bad debt expense, and AR balances by channel.")
        _assumption_editor("Receivables aging schedule", "receivables", receivables_df)
        st.caption("Inventory schedule drives stage-based DIO, write-offs, reserves, and balance-sheet inventory.")
        _assumption_editor("Inventory aging schedule", "inventory", inventory_df)
        st.caption("Payables schedule drives supplier-term balances and procurement discount benefits.")
        _assumption_editor("Payables aging schedule", "payables", payables_df)
        _assumption_editor("Other income items", "other_income", other_income_df)
        _assumption_editor("CAPEX schedule", "capex", capex_df)
        _assumption_editor("Debt facilities", "debt", debt_df)
        _assumption_editor("Equity injections", "equity", equity_df)

# ---------------------------------------------------------------------------
# Scenario state hooks — called by the parent NumQuants shell to save/restore
# the full workspace state across sessions.
# ---------------------------------------------------------------------------

_STATE_KEYS = [
    "assump_data_skus",
    "assump_data_channels",
    "assump_data_sales_plan_base",
    "assump_sales_plan_frequency",
    "assump_sales_plan_propagation_mode",
    "assump_sales_plan_propagation_growth_annual",
    "assump_data_sku_operations",
    "assump_data_brewhouse",
    "assump_data_cellar",
    "assump_data_packaging",
    "assump_data_cost_pools",
    "assump_data_direct_labor",
    "assump_data_indirect_labor",
    "assump_data_receivables",
    "assump_data_inventory",
    "assump_data_payables",
    "assump_data_other_income",
    "assump_data_capex",
    "assump_data_debt",
    "assump_data_equity",
    "assump_data_config",
    "assump_data_dividend",
]


def get_state() -> dict:
    """Return a snapshot of all user-editable assumption tables.

    DataFrames are kept as-is; the parent app's serialisation layer converts
    them to JSON-safe records before writing to the database.
    """
    import streamlit as _st
    return {k: _st.session_state[k] for k in _STATE_KEYS if k in _st.session_state}


def set_state(state: dict) -> None:
    """Pre-populate session_state from a previously saved snapshot.

    Must be called *before* main() so that widget initial values are picked up
    on the first render of each assumption editor.
    """
    import streamlit as _st
    for k, v in state.items():
        if k in _STATE_KEYS:
            _st.session_state[k] = v


if __name__ == "__main__":
    main()
